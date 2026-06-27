"""재무제표(Financials) — FnGuide(전자공시 집계) 연결 기준 PL·BS·CF.

연간(YoY%)·분기(QoQ%)를 1회 계산해 디스크에 저장하고, Financials 탭은 저장본을 즉시 서빙한다
(로딩 없음). 실시간 갱신 불필요 — 분기/사업보고서 제출 마감일 저녁 cron(main.py)이 일괄 갱신.

데이터 출처: FnGuide SVD_Finance(키리스). DART OpenAPI(OPENDART_API_KEY)는 로컬 미설정이라
키리스 FnGuide로 통일(EBITDA·추정실적과 동일 정책). 서버에 키가 있으면 향후 DART로 5개년 확장 가능.
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import date
from functools import lru_cache
from io import StringIO

import pandas as pd  # noqa: F401  (bs4가 주 파서지만 환경 일치 위해 유지)
import requests
from bs4 import BeautifulSoup

_log = logging.getLogger("app.financials")
_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "financials")
_FRESH_DAYS = 80          # 분기 주기(≈91일)보다 짧게 — 마감일 cron이 우선 갱신
_UA = {"User-Agent": "Mozilla/5.0"}

# 재무제표 3종 → FnGuide div id (연간, 분기)
_STMTS = [
    ("PL", "손익계산서", "divSonikY", "divSonikQ"),
    ("BS", "재무상태표", "divDaechaY", "divDaechaQ"),
    ("CF", "현금흐름표", "divCashY", "divCashQ"),
]


def _num(s: str):
    s = (s or "").strip().replace(",", "")
    if not s or s in {"-", "N/A", "n/a"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _pct(cur, prev):
    """전기 대비 증감률(%). 전기가 0/None이거나 부호가 바뀌면 None(왜곡 방지)."""
    if cur is None or prev is None or prev == 0:
        return None
    if (cur < 0) != (prev < 0):   # 적자전환/흑자전환은 % 무의미
        return None
    return round((cur - prev) / abs(prev) * 100, 1)


def _parse_div(soup: BeautifulSoup, div_id: str, annual: bool) -> dict:
    """FnGuide 재무제표 table 1개 파싱 → {periods:[yyyy/mm], rows:[{account,bold,values:[]}]}."""
    div = soup.find(id=div_id)
    if div is None:
        return {"periods": [], "rows": []}
    table = div.find("table")
    if table is None:
        return {"periods": [], "rows": []}
    head = table.find("thead")
    cols = [th.get_text(strip=True) for th in head.find_all("th")] if head else []
    # 기간 컬럼만 — 연간은 결산월(/12 등 사업연도말)만, 분기는 모든 분기 컬럼
    period_idx, periods = [], []
    for i, c in enumerate(cols[1:]):   # cols[0]=IFRS(연결)
        if re.match(r"^\d{4}/\d{2}$", c):
            if annual and not c.endswith("/12"):
                continue              # 연간 표의 누적분기(예: 2026/03) 제외
            period_idx.append(i)
            periods.append(c)
    body = table.find("tbody")
    rows = []
    gid, cur_group = 0, None
    for tr in body.find_all("tr"):
        th = tr.find("th")
        if th is None:
            continue
        cls = tr.get("class") or []
        # FnGuide UI 토글 문구 제거 → 순수 계정명
        account = re.sub(r"\s*계산에 참여한 계정\s*(펼치기|접기)\s*$", "",
                         th.get_text(" ", strip=True)).strip()
        if not account:
            continue
        is_parent = "acd_dep_start_close" in cls or "acd_dep_start_open" in cls   # 펼침 가능 부모
        is_child = "acd_dep2_sub" in cls                                          # 기본 숨김 상세
        bold = "rowBold" in cls
        if is_parent:
            gid += 1; cur_group = gid; group = gid
        elif is_child:
            group = cur_group
        else:
            cur_group = None; group = None
        tds = tr.find_all("td")
        values = [(_num(tds[i].get_text(strip=True)) if i < len(tds) else None) for i in period_idx]
        # 전 기간 공란인 비주요 계정(해당사항 없음)은 노이즈 → 생략
        if not bold and not is_parent and all(v is None for v in values):
            continue
        rows.append({"account": account, "bold": bold, "parent": is_parent,
                     "child": is_child, "group": group, "values": values})
    return {"periods": periods, "rows": rows}


def _with_change(parsed: dict) -> dict:
    """각 계정에 기간별 증감률(YoY 또는 QoQ) 배열 추가. change[0]=None(기준기)."""
    for row in parsed["rows"]:
        vals = row["values"]
        row["change"] = [None] + [_pct(vals[i], vals[i - 1]) for i in range(1, len(vals))]
    return parsed


def _add_pl_metrics(period: dict) -> None:
    """손익계산서(PL)에 이익률(%) + EBITDA·EBITDA Margin(%) 하위행을 1회 계산해 삽입.

    - 영업이익률/당기순이익률/매출총이익률 = 항목 / 매출액 × 100 (해당 항목 바로 아래).
    - EBITDA = 영업이익 + 감가상각비 + 무형자산상각비(현금흐름표) → EBITDA Margin = EBITDA / 매출액.
    파생행은 derived=True(들여쓰기 표시), 비율행은 pct=True(% 포맷). 저장본에 들어가 매번 재계산 X.
    """
    pl, cf = period.get("PL"), period.get("CF")
    if not pl or not pl.get("rows"):
        return
    # canon(표준명) 우선 매칭 — DART 원본명이 변형(당기순이익(손실)·연결당기순이익 등)이어도 인식.
    norm = lambda r: (r.get("canon") or r.get("account", "")).replace(" ", "")

    def find(stmt, pred):
        if not stmt:
            return None
        return next((r for r in stmt["rows"] if pred(norm(r))), None)

    rev = find(pl, lambda a: a == "매출액")
    if not rev:
        return
    rv = rev["values"]

    def ratio(num):
        return [(round(x / d * 100, 1) if (x is not None and d) else None) for x, d in zip(num, rv)]

    def mk_pct(label, vals):
        return {"account": label, "canon": label, "values": vals, "change": [None] * len(vals),
                "bold": False, "parent": False, "child": False, "group": None,
                "pct": True, "derived": True}

    # EBITDA = 영업이익 + 감가상각비 + 무형자산상각비(CF, 대손/사채상각 제외)
    op = find(pl, lambda a: a == "영업이익")
    dep = find(cf, lambda a: "감가상각" in a and "대손" not in a)
    amo = find(cf, lambda a: "무형자산" in a and "상각" in a and "대손" not in a)
    ebitda = None
    if op and (dep or amo):     # D&A를 못 찾으면(예: DART 표준 CF) EBITDA=영업이익 오해 방지 → 생략
        def at(row, i):
            return row["values"][i] if (row and i < len(row["values"])) else None
        ev = []
        for i, o in enumerate(op["values"]):
            ev.append(None if o is None else o + (at(dep, i) or 0) + (at(amo, i) or 0))
        if any(v is not None for v in ev):
            ebitda = ev

    new = []
    for r in pl["rows"]:
        new.append(r)
        a = norm(r)
        if a == "매출총이익":
            new.append(mk_pct("매출총이익률(%)", ratio(r["values"])))
        elif a == "영업이익":
            new.append(mk_pct("영업이익률(%)", ratio(r["values"])))
            if ebitda:
                new.append({"account": "EBITDA", "canon": "EBITDA", "values": ebitda, "bold": False,
                            "parent": False, "child": False, "group": None, "derived": True,
                            "change": [None] + [_pct(ebitda[i], ebitda[i - 1]) for i in range(1, len(ebitda))]})
                new.append(mk_pct("EBITDA Margin(%)", ratio(ebitda)))
        elif a == "당기순이익":
            new.append(mk_pct("당기순이익률(%)", ratio(r["values"])))
    pl["rows"] = new


def _graft_sga(fg_pl: dict | None, dart_pl: dict | None) -> None:
    """FnGuide 연간 PL의 판관비 상세(인건비·연구개발비·판매비·관리비 등)를 DART 연간 PL의
    '판매비와관리비' 바로 아래에 이식. DART CIS는 판관비를 단일행으로만 줘 상세가 없다(분기는 FnGuide라 있음)."""
    if not fg_pl or not dart_pl:
        return
    fg_rows = fg_pl.get("rows") or []
    fg_periods = fg_pl.get("periods") or []
    norm = lambda s: (s or "").replace(" ", "")
    names = [norm(r.get("account")) for r in fg_rows]
    if "판매비와관리비" not in names:
        return
    i_sga = names.index("판매비와관리비")
    i_op = next((i for i in range(i_sga + 1, len(names)) if names[i] in ("영업이익", "영업이익(손실)")), len(fg_rows))
    subs = [r for r in fg_rows[i_sga + 1:i_op] if not r.get("pct")]
    d_rows = dart_pl.get("rows") or []
    d_periods = dart_pl.get("periods") or []
    j = next((i for i, r in enumerate(d_rows) if norm(r.get("account")) == "판매비와관리비"), None)
    if j is None or not subs:
        return
    fp_idx = {p: i for i, p in enumerate(fg_periods)}
    grafted = []
    for r in subs:
        rv = r.get("values") or []
        vals = [(rv[fp_idx[p]] if (p in fp_idx and fp_idx[p] < len(rv)) else None) for p in d_periods]
        if all(v is None for v in vals):
            continue
        grafted.append({"account": r.get("account", ""), "values": vals,
                        "change": [None] * len(vals), "bold": False,
                        "parent": False, "child": True, "group": "판매비와관리비"})
    if grafted:
        dart_pl["rows"] = d_rows[:j + 1] + grafted + d_rows[j + 1:]


def _reorder_by_doc(stmt: dict, layout: list) -> None:
    """OpenAPI rows를 원문 보고서 표시순서(layout=[(ckey,sign)...])로 제자리 재정렬 + 부호 적용.
    - 같은 계정명이 여러 번(유동/비유동 매출채권 등)이면 OpenAPI 등장순서대로 보고서 등장순서에 1:1 배정.
    - 부호: 보고서 표기 부호(비용=음수)를 OpenAPI 절대값에 적용 → 보고서 그대로.
    - 매칭 안 되는 계정은 원래 상대순서 유지하며 맨 뒤."""
    from collections import defaultdict
    rows = stmt.get("rows")
    if not rows or not layout:
        return
    from . import dart_doc
    docpos = defaultdict(list)                       # ckey → [(위치, 부호), ...] 등장순
    for i, (ck, sg) in enumerate(layout):
        docpos[ck].append((i, sg))
    used = defaultdict(int)
    keyed = []
    for j, r in enumerate(rows):
        ck = dart_doc._ckey(r.get("account", ""))
        lst = docpos.get(ck)
        if lst:
            k = min(used[ck], len(lst) - 1)
            used[ck] += 1
            idx, sg = lst[k]
            # 보고서가 이 계정을 표시상 부호반전(비용을 음수표기 등)하면 전 기간 반전.
            # 단 OpenAPI 부호와 보고서 부호가 '다를 때만' — 영업이익(손실)처럼 연도별 흑/적이
            # 바뀌는 계정은 OpenAPI 실제값 보존(최신값 부호를 과거에 강제하면 흑자가 적자로 둔갑).
            if sg is not None:
                nz = [v for v in r.get("values", []) if v is not None]
                if nz and (-1 if nz[-1] < 0 else 1) != sg:
                    r["values"] = [None if v is None else -v for v in r["values"]]
            keyed.append((idx, j, r))
        else:
            keyed.append((10 ** 6, j, r))
    keyed.sort(key=lambda x: (x[0], x[1]))           # 보고서 위치, 동률은 OpenAPI 원순서(안정)
    stmt["rows"] = [r for _, _, r in keyed]


_REV_SYN = {"영업수익", "수익(매출액)", "수익", "매출", "매출수익", "매출액", "영업수익(매출액)"}


def _fill_key(name: str) -> str:
    """채움 매칭용 canon — 매출 동의어(매출액↔영업수익↔수익)는 한 키로(계정 개명 회사 대응)."""
    from . import dart_doc
    ns = (name or "").replace(" ", "")
    return "매출액" if ns in _REV_SYN else dart_doc._ckey(name)


def _extend_periods(stmt: dict, oa: dict) -> None:
    """dart-fss 연간(DART 연결 끊김으로 최근 3개년만 온 경우)을 OpenAPI(oa)로 5개년 축까지 확장.
    옛 연도(FY21/22)를 canon 매칭해 채움. 표시 계정·순서는 원문(dart-fss) 그대로 유지."""
    if not stmt or not oa or not oa.get("periods"):
        return
    sp = stmt.get("periods") or []
    op = oa.get("periods") or []
    target = sorted(set(sp) | set(op))[-5:]
    si = {p: i for i, p in enumerate(sp)}
    oi = {p: i for i, p in enumerate(op)}
    oc = {}
    for r in oa.get("rows", []):
        oc.setdefault(_fill_key(r.get("canon") or r.get("account", "")), r)
    for r in stmt.get("rows", []):
        orow = oc.get(_fill_key(r.get("canon") or r.get("account", "")))
        vals = r.get("values", [])
        ovals = orow.get("values", []) if orow else []
        nv = []
        for p in target:
            v = vals[si[p]] if (p in si and si[p] < len(vals)) else None
            if v is None and p in oi and oi[p] < len(ovals):
                v = ovals[oi[p]]                       # 원문이 비면(개명·throttle) OpenAPI로 채움 — 축 완성이어도 행 갭 채움
            nv.append(v)
        r["values"] = nv
    stmt["periods"] = target


def _fetch(code: str) -> dict:
    r = requests.get("https://comp.fnguide.com/SVO2/ASP/SVD_Finance.asp",
                     params={"pGB": 1, "gicode": f"A{code}"}, headers=_UA, timeout=20)
    if r.encoding in (None, "ISO-8859-1"):
        r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "html.parser")
    annual, quarterly = {}, {}
    for key, _label, div_y, div_q in _STMTS:
        annual[key] = _with_change(_parse_div(soup, div_y, annual=True))
        quarterly[key] = _with_change(_parse_div(soup, div_q, annual=False))
    _add_pl_metrics(annual)      # 이익률·EBITDA 하위행 1회 계산·삽입(저장본에 포함)
    _add_pl_metrics(quarterly)
    # 연간 = DART 전자공시 원문 보고서(계정명·순서·값 '그대로'). 분기 = OpenAPI 단일분기를 원문 계정순서로.
    # FnGuide는 폴백(원문 파싱 실패 시).
    try:
        from .config import settings
        if settings.OPENDART_API_KEY:
            from . import dart, dart_fss_fetch
            draw = dart.fetch(code)
            doc = None
            try:
                doc = dart_fss_fetch.fetch(code)   # dart-fss(XBRL) — 보고서 계정·순서·값 그대로(은행 포함 강건)
            except Exception:
                _log.exception("dart-fss 추출 실패 %s", code)
            d_annual = (doc or {}).get("annual") or {}
            oa_annual = (draw or {}).get("annual") or {}
            merged = {}
            for k in ("PL", "BS", "CF"):
                dk, ok = d_annual.get(k), oa_annual.get(k)
                # 원문이 최신연도까지 있으면 원문(보고서 계정·순서 그대로), 아니면 OpenAPI 폴백
                # (은행·지주 등 원문 파싱 실패/구버전만 — 무회귀).
                use_doc = bool(dk and dk.get("rows") and (not (ok and ok.get("periods")) or
                               (dk.get("periods") and dk["periods"][-1] >= ok["periods"][-1])))
                pick = dk if use_doc else (ok or dk)
                if pick:
                    merged[k] = pick
            if merged.get("PL") or merged.get("BS"):
                for k in ("PL", "BS", "CF"):
                    if merged.get(k):
                        _extend_periods(merged[k], oa_annual.get(k))   # 옛 연도(FY21/22)를 OpenAPI로 채워 5개년 축 보장
                        _with_change(merged[k])
                _add_pl_metrics(merged)   # 이익률(·D&A 있으면 EBITDA) 부여. 판관비 상세 graft는 안 함(보고서 그대로).
                annual = merged    # 연간 = 원문 보고서(폴백 시 OpenAPI)
                # 분기: OpenAPI 단일분기 값을 원문(연간) 계정순서로 재배치 — 분기보고서는 같은 계정 구조.
                layout = {k: [(r.get("canon") or r.get("account", ""),
                               (-1 if (nz := [x for x in r.get("values", []) if x is not None]) and nz[-1] < 0
                                else (1 if nz else None)))
                              for r in v.get("rows", [])] for k, v in merged.items()}
                d_quarterly = (draw or {}).get("quarterly")
                if d_quarterly and d_quarterly.get("PL"):
                    for k in ("PL", "BS", "CF"):
                        if d_quarterly.get(k):
                            if layout.get(k):
                                _reorder_by_doc(d_quarterly[k], layout[k])
                            _with_change(d_quarterly[k])
                    _add_pl_metrics(d_quarterly)
                    quarterly = d_quarterly
    except Exception:
        _log.exception("DART 전자공시 병합 실패 — FnGuide 유지 %s", code)
    return {"fetched": date.today().isoformat(), "annual": annual, "quarterly": quarterly}


def _path(code: str) -> str:
    return os.path.join(_DIR, f"{code}.json")


def _save(code: str, data: dict) -> None:
    os.makedirs(_DIR, exist_ok=True)
    tmp = _path(code) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, _path(code))


def refresh(code: str) -> dict:
    """FnGuide에서 재무제표를 받아 증감률 계산·저장. 반환: 저장 데이터."""
    data = _fetch(code)
    try:
        _save(code, data)
    except Exception:
        _log.exception("재무제표 저장 실패 %s", code)
    return data


def _load_or_fetch(code: str) -> dict:
    path = _path(code)
    try:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                cached = json.load(f)
            fetched = cached.get("fetched", "")
            if fetched and (date.today() - date.fromisoformat(fetched)).days < _FRESH_DAYS:
                return cached
    except Exception:
        pass
    try:
        return refresh(code)
    except Exception as e:
        _log.warning("재무제표 조회 실패 %s: %s", code, e)
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"fetched": "", "annual": {}, "quarterly": {}}


@lru_cache(maxsize=1024)
def _cached(code: str, _day: str) -> dict:
    return _load_or_fetch(code)


def financials(code: str) -> dict:
    """Financials 탭용 — 메모리 캐시(프로세스 내 즉시) → 디스크 저장본 → 라이브(최후) 순.

    디스크는 Railway 재시작 시 휘발하므로 (code, 오늘) 메모리 lru_cache로 같은 종목 재요청·재마운트를
    즉시 응답한다. 기동 시 prewarm으로 산업 종목을 미리 데워 첫 진입도 빠르게."""
    return _cached(str(code).strip(), date.today().isoformat())


def clear_cache() -> None:
    """장 마감 후 등 캐시 무효화."""
    _cached.cache_clear()


def to_xlsx(code: str) -> bytes:
    """재무제표(연간·분기 × 손익/재무/현금) .xlsx — 사용자 다운로드용.

    양식(사용자 수정본 기준): A열 여백·B열부터 / B1 로고·헤더 = 네이비(#0F243E)+흰 글자 / 한글 맑은 고딕·
    숫자 Arial / 제목 외 9pt / 눈금선 제거 / 금액=백만원·정수표시(#,##0)이되 셀 값은 소수 풀정밀도 보존 /
    FY헤더·숫자 우측정렬 / yoy(%)·이익률은 표 하단 별도 섹션(전년대비 증감 / Margins)으로 모음."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 색(사용자 수정본 기준). 제목바=네이비, 소제목=연회색채움+네이비글자, 볼드계정=아주연한채움, 밑줄=회색.
    NAVY = "FF0F243E"
    GRAY = "FF646464"
    WHITE = "FFFFFFFF"
    HEAD_BG = "FFD9D9D9"   # 소제목(계정/YoY/Margins) 행 채움
    BOLD_BG = "FFF2F2F2"   # 볼드 소계 계정 채움
    LINE = "FFA6A6A6"      # 밑줄 색
    AMT_FMT = "#,##0;(#,##0);-"   # 백만원 표시는 정수(셀 값은 ×scale 풀정밀도 → F2/수식입력줄에 소수점)
    WON_FMT = '#,##0"원";(#,##0"원");-'   # 주당이익(EPS 등) — 표시는 정수 원(셀 값은 풀정밀도)
    PCT_FMT = '0.0"%"'            # 이익률(값이 이미 %)
    YOY_FMT = "0.0%;(0.0%);-"     # RATE 결과(소수) → %
    title_fill = PatternFill("solid", fgColor=NAVY)
    head_fill = PatternFill("solid", fgColor=HEAD_BG)
    bold_fill = PatternFill("solid", fgColor=BOLD_BG)
    under = Border(bottom=Side(style="thin", color=LINE))

    def fname(text):
        return "Arial" if str(text if text is not None else "").isascii() else "맑은 고딕"

    def plabel(p, quarterly):
        try:
            y, m = p.split("/")
        except ValueError:
            return p
        if quarterly:
            return f"{({'03': 1, '06': 2, '09': 3, '12': 4}).get(m, m)}Q{y[2:]}"
        return f"FY{y[2:]}"

    data = financials(code)
    try:
        from . import kis_master_cache
        name = kis_master_cache.get_name(str(code).strip()) or str(code)
    except Exception:
        name = str(code)
    labels = {"PL": "손익계산서", "BS": "재무상태표", "CF": "현금흐름표"}
    wb = Workbook()
    wb.remove(wb.active)

    for pdata, tag, quarterly in [((data.get("annual") or {}), "연간", False),
                                  ((data.get("quarterly") or {}), "분기", True)]:
        for key in ("PL", "BS", "CF"):
            st = pdata.get(key) or {}
            periods = st.get("periods") or []
            rows = st.get("rows") or []
            if not periods or not rows:
                continue
            ws = wb.create_sheet(f"{tag}_{labels[key]}"[:31])
            ws.sheet_view.showGridLines = False
            np_ = len(periods)
            c0, v0 = 2, 3                       # 계정=B, 첫 값=C (A는 여백)
            last = get_column_letter(v0 + np_ - 1)

            def put(r, c, val, *, font_name=None, bold=False, color=GRAY, size=9,
                    align="left", indent=0, fmt=None, fill=None):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = Font(name=font_name or fname(val), bold=bold, size=size, color=color)
                cell.alignment = Alignment(horizontal=align, vertical="center", indent=indent)
                if fmt:
                    cell.number_format = fmt
                if fill:
                    cell.fill = fill

            def underline(r):              # 밑줄 = B~끝 하단 thin(회색)
                for c in range(c0, v0 + np_):
                    ws.cell(row=r, column=c).border = under

            def section(r, title):         # 소제목 행(연회색 채움 B~끝 + 네이비 볼드)
                put(r, c0, title, bold=True, color=NAVY, fill=head_fill)
                for c in range(v0, v0 + np_):
                    put(r, c, None, bold=True, color=NAVY, align="right", fill=head_fill)

            # R1 여백 / R2 제목바(네이비+흰글자) / R3 단위 부제 / R4 소제목 계정헤더
            ws.row_dimensions[1].height = 11
            ws.merge_cells(f"B2:{last}2")
            put(2, c0, f"{name} - {code}", bold=True, color=WHITE, size=20, fill=title_fill)
            ws.row_dimensions[2].height = 28.5
            ws.merge_cells(f"B3:{last}3")
            put(3, c0, f"{tag} {labels[key]}  (단위: 백만원, 비율: %)", bold=True, color=GRAY, size=9)
            put(4, c0, "계정", bold=True, color=NAVY, fill=head_fill)
            for j, p in enumerate(periods):
                put(4, v0 + j, plabel(p, quarterly), font_name="Arial", bold=True,
                    color=NAVY, align="right", fill=head_fill)

            # 데이터 — 금액 계정(값만). yoy·이익률은 하단 섹션으로 모음.
            # 볼드 소계=연한 채움 + 직전 행에 밑줄(섹션 첫 행 제외). 자식계정만 들여쓰기. 섹션 끝 밑줄.
            ri = 5
            margins = []
            accs = []   # (계정명, 엑셀행, bold, child) — 하단 YoY 섹션용
            prev = None
            for r in rows:
                if r.get("pct"):
                    margins.append(r)
                    continue
                bold = bool(r.get("bold"))
                child = bool(r.get("child"))
                acc = r.get("account", "")
                # 주당이익(EPS)은 원 단위(DART /1e8 저장 → ×1e8 원 환원, 계정명 '주당' 포함).
                is_won = "주당" in "".join(acc.split())
                scale, vfmt = (1e8, WON_FMT) if is_won else (100, AMT_FMT)
                rfill = bold_fill if bold else None
                if bold and prev is not None:
                    underline(prev)
                put(ri, c0, acc, bold=bold, color=GRAY, indent=1 if child else 0, fill=rfill)
                for j, v in enumerate(r.get("values") or []):
                    put(ri, v0 + j, ("-" if v is None else v * scale), font_name="Arial",
                        bold=bold, color=GRAY, align="right",
                        fmt=(None if v is None else vfmt), fill=rfill)
                accs.append((acc, ri, bold, child, (r.get("canon") or acc), r.get("values") or []))
                prev = ri
                ri += 1
            if prev is not None:
                underline(prev)

            # 소계 자동합계 — 하위항목 있는 소계를 SUBTOTAL(9,하위계정)로 합계검증.
            # 하위계정=소계 바로 다음부터 다음 볼드 소계 직전까지(CF는 현금성자산 요약행도 경계). 풀정밀도 보존.
            # CF는 제외 — DART CF 하위계정이 절대값(취득·상환 모두 양수)이라 단순 SUM이
            # 순증감과 불일치. CF 활동소계는 DART 원본값(정확)을 그대로 사용.
            SUB_PARENTS = {
                "BS": {"유동자산", "비유동자산", "유동부채", "비유동부채"},
                "PL": {"판매비와관리비"},
            }.get(key, set())
            for idx, (a, prow, pbold, child, canon, pvals) in enumerate(accs):
                if canon.replace(" ", "") not in SUB_PARENTS:
                    continue
                kids, kvals = [], []
                for (a2, r2, b2, c2, cn2, v2) in accs[idx + 1:]:
                    if b2 or (key == "CF" and "현금및현금성자산" in cn2.replace(" ", "")):
                        break              # 다음 소계 또는 CF 현금성자산 요약행(증감·환율·기초·기말)에서 중단
                    kids.append(r2)
                    kvals.append(v2)
                if not kids:
                    continue
                # 부호 정합: 부모가 음수표기(비용을 음수로 적은 회사의 판관비 등)이고 하위합이
                # 양수면 =-SUBTOTAL. 최신 동시존재 기간으로 판정(아니면 +).
                neg = False
                for t in range(min(len(pvals), np_) - 1, -1, -1):
                    pv = pvals[t]
                    ks = [v[t] for v in kvals if t < len(v) and v[t] is not None]
                    if pv is None or not ks:
                        continue
                    if pv < 0 and sum(ks) > 0:
                        neg = True
                    break
                pre = "-" if neg else ""
                rfill = bold_fill if pbold else None
                for j in range(np_):
                    col = get_column_letter(v0 + j)
                    put(prow, v0 + j, f"={pre}SUBTOTAL(9,{col}{kids[0]}:{col}{kids[-1]})",
                        font_name="Arial", bold=pbold, color=GRAY, align="right",
                        fmt=AMT_FMT, fill=rfill)

            # 하단 ① 전년대비 증감(YoY %) — 전 계정 모아서. RATE(1,0,-전기,당기)=당기/전기-1.
            if accs:
                ri += 1
                section(ri, "전년대비 증감 (YoY %)")
                ri += 1
                prev = None
                for nm, arow, bold, child, _canon, _vals in accs:
                    rfill = bold_fill if bold else None
                    if bold and prev is not None:
                        underline(prev)
                    put(ri, c0, nm, bold=bold, color=GRAY, indent=1 if child else 0, fill=rfill)
                    put(ri, v0, None, bold=bold, color=GRAY, align="right", fill=rfill)  # 첫 기간 YoY 없음
                    for j in range(1, np_):
                        pcol, ccol = get_column_letter(v0 + j - 1), get_column_letter(v0 + j)
                        put(ri, v0 + j, f'=IFERROR(RATE(1,0,-{pcol}{arow},{ccol}{arow}),"n/a")',
                            font_name="Arial", bold=bold, color=GRAY, align="right",
                            fmt=YOY_FMT, fill=rfill)
                    prev = ri
                    ri += 1
                if prev is not None:
                    underline(prev)

            # 하단 ② Margins(손익 — 이익률 %)
            if margins:
                ri += 1
                section(ri, "Margins")
                ri += 1
                prev = None
                for r in margins:
                    put(ri, c0, r.get("account", ""), color=GRAY, indent=1)
                    for j, v in enumerate(r.get("values") or []):
                        put(ri, v0 + j, ("-" if v is None else v), font_name="Arial",
                            color=GRAY, align="right", fmt=(None if v is None else PCT_FMT))
                    prev = ri
                    ri += 1
                if prev is not None:
                    underline(prev)

            ws.column_dimensions["A"].width = 2.4
            ws.column_dimensions["B"].width = 34
            for i in range(np_):
                ws.column_dimensions[get_column_letter(v0 + i)].width = 18
            ws.freeze_panes = "C5"
    if not wb.sheetnames:
        wb.create_sheet("재무제표").append(["재무 데이터가 없습니다."])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def prewarm(codes: list[str]) -> int:
    """기동 시 산업 종목 재무제표를 메모리(+디스크)에 미리 적재 — 첫 탭 진입도 즉시."""
    n = 0
    for c in codes:
        try:
            d = financials(c)
            if d.get("annual"):
                n += 1
        except Exception:
            pass
    return n


def refresh_all(codes: list[str]) -> int:
    """분기/사업보고서 마감일 cron이 호출 — 추적 종목 재무제표 일괄 갱신(병렬). 반환: 성공 수."""
    from concurrent.futures import ThreadPoolExecutor
    ok = 0
    with ThreadPoolExecutor(max_workers=8) as ex:
        for data in ex.map(lambda c: _safe_refresh(c), codes):
            if data:
                ok += 1
    _log.info("재무제표 %d/%d 갱신·저장", ok, len(codes))
    return ok


def _safe_refresh(code: str):
    try:
        d = refresh(code)
        return d if (d.get("annual") or d.get("quarterly")) else None
    except Exception:
        return None
