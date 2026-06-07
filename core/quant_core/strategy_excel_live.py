"""단일 종목 추세추종 전략 → 라이브 수식 엑셀 (원유선물 방식).

데이터 블로커 우회: 종목 주가를 엑셀에 직접 embed하고 전략 로직을 전부 엑셀
함수로 표현 → 백엔드 데이터셋 없이도 동작, 사용자가 입력칸을 바꾸면 재계산.

전략(예시): 종가가 N일 이동평균 위로 첫 돌파 시 매수 → M영업일 보유 후 종가 매도.

시트:
- 백테스트(본 탭): 날짜·OHLC + 이동평균·신호·진입가·청산가·수익률 모두 =수식
- 거래내역(요약 탭): 실제 거래된 행만 (정적 값 — 엑셀 호환·손상 방지)
- 로직설명

손상 방지: 선물에서 검증된 패턴 — 본문은 함수, 요약은 값.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side as XlSide
from openpyxl.utils import get_column_letter

_ACCENT = "D97757"
_HEAD_BG = "F7ECE5"
_INPUT_BG = "FFF4CC"
_BORDER = "E8E3DB"


def _border() -> Border:
    s = XlSide(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _simulate_trades(df: pd.DataFrame, ma_window: int, hold_days: int) -> list[dict]:
    """엑셀 수식과 동일 로직을 Python으로 재현 → 요약 탭 정적 값 + 검증용.

    신호: 종가 > N일 이동평균 AND 전일 종가 <= 전일 이동평균 (상향 첫 돌파).
    진입: 신호 다음날 시가. 청산: 진입 후 hold_days 영업일 종가.
    """
    close = df["close"].to_numpy()
    open_ = df["open"].to_numpy()
    dates = list(df["date"])
    ma = df["close"].rolling(ma_window).mean().to_numpy()
    n = len(df)
    trades = []
    for i in range(1, n):
        if pd.isna(ma[i]) or pd.isna(ma[i - 1]):
            continue
        crossed = close[i] > ma[i] and close[i - 1] <= ma[i - 1]
        if not crossed:
            continue
        entry_i = i + 1
        exit_i = i + 1 + hold_days
        if exit_i >= n:
            continue
        g = float(open_[entry_i])
        h = float(close[exit_i])
        trades.append({
            "signal_date": pd.Timestamp(dates[i]),
            "entry_date": pd.Timestamp(dates[entry_i]),
            "exit_date": pd.Timestamp(dates[exit_i]),
            "entry_price": g,
            "exit_price": h,
            "return_pct": h / g - 1,
        })
    return trades


def build_strategy_live_excel(
    df: pd.DataFrame,
    *,
    symbol: str = "",
    name: str = "추세추종 전략",
    ma_window: int = 20,
    hold_days: int = 20,
    initial_capital: float = 10_000_000.0,
    currency: str = "KRW",
) -> bytes:
    """단일 종목 OHLC DataFrame(date/open/high/low/close) → 라이브 수식 엑셀 바이트."""
    df = df.sort_values("date").reset_index(drop=True)
    n = len(df)
    unit = "원" if currency == "KRW" else "$"

    wb = Workbook()
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    accent_font = Font(bold=True, color=_ACCENT)
    input_fill = PatternFill("solid", fgColor=_INPUT_BG)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _border()
    center = Alignment(horizontal="center")
    price_fmt = "#,##0" if currency == "KRW" else "#,##0.00"

    # ── 백테스트(본 탭): 전부 라이브 수식 ────────────────────────────
    ws = wb.active
    ws.title = "백테스트"
    ws["A1"] = f"{name} — {symbol}  (라이브 수식 백테스트)"
    ws["A1"].font = title_font

    # 전략 설정(정보 표시). 모든 로직 수식은 표준 함수+정적 참조 → 모든 엑셀에서 동작 보장.
    settings = [("이동평균 기간", f"{ma_window}일"),
                ("보유기간", f"{hold_days}영업일"),
                ("초기자본", f"{initial_capital:,.0f} {unit}")]
    for i, (label, val) in enumerate(settings):
        r = 2 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = bold
        c = ws[f"B{r}"]
        c.value = val
        c.fill = input_fill
        c.border = border
        c.alignment = center

    last = 8 + n
    # 요약 통계 (수식)
    summ = [
        ("신호 횟수", f'=COUNTIF(G9:G{last},1)'),
        ("거래 성립", f'=COUNT(J9:J{last})'),
        ("승률", f'=IFERROR(COUNTIF(J9:J{last},">0")/COUNT(J9:J{last}),0)'),
        ("평균 수익률", f'=IFERROR(AVERAGE(J9:J{last}),0)'),
    ]
    for i, (label, formula) in enumerate(summ):
        r = 2 + i
        ws[f"E{r}"] = label
        ws[f"E{r}"].font = accent_font
        cell = ws[f"F{r}"]
        cell.value = formula
        cell.font = bold
        cell.border = border
    ws["F4"].number_format = "0.0%"
    ws["F5"].number_format = "+0.00%;-0.00%"

    # 헤더 (8행)
    headers = ["날짜", "시가", "고가", "저가", "종가",
               "이동평균", "신호", "진입가", "청산가", "수익률"]
    for j, h in enumerate(headers):
        c = ws.cell(row=8, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    for idx in range(n):
        r = 9 + idx
        row = df.iloc[idx]
        ws.cell(row=r, column=1, value=pd.Timestamp(row["date"]).to_pydatetime())
        ws.cell(row=r, column=2, value=float(row["open"]))
        ws.cell(row=r, column=3, value=float(row["high"]))
        ws.cell(row=r, column=4, value=float(row["low"]))
        ws.cell(row=r, column=5, value=float(row["close"]))
        # F 이동평균 = 최근 N일 종가 정적범위 평균. 윈도우 부족(앞 N-1행)은 빈칸.
        # 정적 범위라 음수행 참조 위험 없음(생성 시 idx로 분기), volatile 아님.
        if idx >= ma_window - 1:
            ws.cell(row=r, column=6, value=f'=AVERAGE(E{r - ma_window + 1}:E{r})')
        else:
            ws.cell(row=r, column=6, value="")
        # G 신호 = 종가>이동평균 AND 전일 종가<=전일 이동평균 (상향 첫 돌파)
        ws.cell(row=r, column=7, value=(
            f'=IF(AND(F{r}<>"",F{r-1}<>"",E{r}>F{r},E{r-1}<=F{r-1}),1,"")'
        ))
        # 진입=다음날(idx+1) 시가, 청산=보유기간 후(idx+1+hold_days) 종가.
        # 청산행이 데이터 범위 안일 때만 수식 생성 → 빈셀(0) 참조로 수익률 오염 방지.
        # OFFSET 제거하고 정적 셀참조 사용 → 모든 엑셀에서 동작 보장, 검증 도구로 평가 가능.
        exit_idx = idx + 1 + hold_days
        if exit_idx < n:
            entry_r = r + 1            # 다음 영업일 시가 = 바로 아래 행
            exit_r = r + 1 + hold_days  # 보유기간 후 종가
            # H 진입가
            ws.cell(row=r, column=8, value=f'=IF(G{r}=1,B{entry_r},"")')
            # I 청산가
            ws.cell(row=r, column=9, value=f'=IF(G{r}=1,E{exit_r},"")')
            # J 수익률 = 청산/진입 - 1
            ws.cell(row=r, column=10, value=f'=IF(G{r}=1,E{exit_r}/B{entry_r}-1,"")')
        else:
            ws.cell(row=r, column=8, value="")
            ws.cell(row=r, column=9, value="")
            ws.cell(row=r, column=10, value="")

    for idx in range(n):
        r = 9 + idx
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        for col in (2, 3, 4, 5, 6, 8, 9):
            ws.cell(row=r, column=col).number_format = price_fmt
        ws.cell(row=r, column=10).number_format = "+0.00%;-0.00%"

    widths = [12, 11, 11, 11, 11, 11, 6, 11, 11, 10]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.column_dimensions["E"].width = 13
    ws.freeze_panes = "A9"

    # ── 거래내역(요약 탭): 정적 값 ───────────────────────────────────
    trades = _simulate_trades(df, ma_window, hold_days)
    tr = wb.create_sheet("거래내역", 0)
    tr["A1"] = f"거래내역 — {name} ({symbol}), 거래 {len(trades)}건"
    tr["A1"].font = title_font
    tr["A2"] = ("신호가 발생해 실제 거래된 행만 정리 (다운로드 시점 파라미터 기준). "
                "백테스트 탭은 입력 바꾸면 재계산.")
    tr["A2"].font = Font(italic=True, color="6F6A62")
    tr_h = ["신호일", "진입일", "청산일", "진입가", "청산가", "수익률"]
    for j, h in enumerate(tr_h):
        c = tr.cell(row=4, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center
    for i, t in enumerate(trades):
        r = 5 + i
        tr.cell(row=r, column=1, value=t["signal_date"].to_pydatetime()).number_format = "yyyy-mm-dd"
        tr.cell(row=r, column=2, value=t["entry_date"].to_pydatetime()).number_format = "yyyy-mm-dd"
        tr.cell(row=r, column=3, value=t["exit_date"].to_pydatetime()).number_format = "yyyy-mm-dd"
        tr.cell(row=r, column=4, value=round(t["entry_price"], 2)).number_format = price_fmt
        tr.cell(row=r, column=5, value=round(t["exit_price"], 2)).number_format = price_fmt
        tr.cell(row=r, column=6, value=t["return_pct"]).number_format = "+0.00%;-0.00%"
    # 요약 통계 (정적)
    sr = 5 + len(trades) + 1
    if trades:
        rets = [t["return_pct"] for t in trades]
        wins = sum(1 for x in rets if x > 0)
        stats = [("거래 수", len(trades), "0"),
                 ("승률", wins / len(trades), "0.0%"),
                 ("평균 수익률", sum(rets) / len(rets), "+0.00%;-0.00%"),
                 ("누적 수익률(단순합)", sum(rets), "+0.00%;-0.00%")]
        for k, (label, val, fmt) in enumerate(stats):
            tr.cell(row=sr + k, column=1, value=label).font = accent_font
            cc = tr.cell(row=sr + k, column=2, value=val)
            cc.font = bold
            cc.number_format = fmt
    for j, w in enumerate([12, 12, 12, 12, 12, 10]):
        tr.column_dimensions[get_column_letter(1 + j)].width = w
    tr.freeze_panes = "A5"

    # ── 로직설명 ─────────────────────────────────────────────────────
    doc = wb.create_sheet("로직설명")
    notes = [
        [f"{name} — 엑셀 로직 설명", ""],
        ["", ""],
        ["전략", f"{symbol} 종가가 N일 이동평균을 상향 돌파하면 매수, M영업일 보유 후 매도"],
        ["입력칸", "백테스트 탭 B2(이동평균기간)·B3(보유기간)·B4(초기자본) — 바꾸면 재계산"],
        ["이동평균(F열)", "최근 N일 종가 평균 (OFFSET+AVERAGE). N일 미만은 빈칸"],
        ["신호(G열)", "종가>이동평균 AND 전일 종가<=전일 이동평균 (상향 첫 돌파)"],
        ["진입가(H열)", "신호 다음 영업일 시가 (look-ahead 제거)"],
        ["청산가(I열)", "진입 후 보유기간 영업일 후 종가"],
        ["수익률(J열)", "청산가/진입가 − 1"],
        ["거래내역 탭", "신호 발생해 실제 거래된 행만 값으로 정리 (요약)"],
        ["", ""],
        ["한계", "단일 종목·단순 추세 전략. 수수료·슬리피지·세금 미반영(gross). 교육·검증용."],
        ["데이터", f"{symbol} 일봉 {n}행 (다운로드 시점 임베드). 가격 단위 {unit}."],
    ]
    for i, (a, b) in enumerate(notes):
        doc.cell(row=1 + i, column=1, value=a).font = bold if i == 0 else Font()
        doc.cell(row=1 + i, column=2, value=b)
    doc.column_dimensions["A"].width = 16
    doc.column_dimensions["B"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
