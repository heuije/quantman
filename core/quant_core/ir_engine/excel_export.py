"""인사이트 엔진 분석 → '증빙' 엑셀(.xlsx) 생성.

목적: 챗봇/빌더가 IR 백테스트를 돌리면 *결과만* 주는 게 아니라, **어떤 데이터를 어떤
연산으로** 산출했는지 엑셀로 증빙한다. 선물 분석의 build_oil_excel(라이브 수식)과 같은
취지 — 데이터 + 수식이 든 파일로, 사용자가 직접 검증한다.

엔진은 정수주·현금·마진·지연체결을 가진 **이벤트 구동 NAV 시뮬레이션**이라 셀 수식으로
*정확히* 복제할 수 없다(naive cumprod(가중×수익)은 엔진과 불일치). 그래서 라이브 수식은
엔진의 **정본 자산곡선(equity) 위에서 정의식 변환**(일수익·낙폭·CAGR·샤프·MDD)만 한다 —
이 값들은 엔진 metrics와 *정확히 일치*(증빙)하고, '일일 추가비용(bps)' 노란 칸으로
사후 민감도를 라이브로 본다(0이면 엔진과 동일). 전략 로직/파라미터 변경의 라이브 재계산은
'변수 조정'(/ir/strategy 재실행)이 담당 — 엑셀은 산술·비용 검증 도구다.

시트:
- 백테스트: 엔진 자산곡선 + 라이브 정의식(일수익·조정수익·조정자산·낙폭) + 지표 비교.
- 거래내역: 엔진 trades(정적 값) — 검증 앵커.
- 원자료: 보유종목 종가(정적 값) — '어떤 데이터'.
- 일별비중: 엔진 EOD 기여 비중 패널(정적 값) — '어떤 포지션'.
- 지표·설명: 엔진 metrics 전체 + 전략 정의(IR) + 방법론·한계.
"""
from __future__ import annotations

import io
import json
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side as XlSide
from openpyxl.utils import get_column_letter

from .spec import StrategyIR

# 디자인 토큰 (네이비/골드 — DESIGN.md 계열, 인쇄 가독 우선 라이트)
_ACCENT = "A6791F"     # 골드(진하게 — 텍스트 가독)
_HEAD_BG = "E9EDF3"    # 헤더 배경(라이트 네이비그레이)
_INPUT_BG = "FFF4CC"   # 입력칸(노랑 — 편집 가능 관례)
_SPEC_BG = "EFEAE3"
_BORDER = "D8DEE7"
_NOTE = "6F6A62"

_MAX_SYMS = 40   # 원자료·일별비중 가로 폭 상한(대형 유니버스 — 초과 시 명시 생략)

# 백테스트 시트 레이아웃 행 좌표
_ROW_TITLE = 1
_ROW_INPUT = 3
_ROW_SUMHEAD = 5
_ROW_SUM0 = 6        # 지표 비교 5행: 6~10
_ROW_DHEAD = 12
_ROW_D0 = 13         # 데이터 첫 행

# 엔진 metrics 키 → 한글 라벨 (없는 키는 원문 사용)
_METRIC_LABELS = {
    "total_return": "누적수익률(%)", "cagr": "연복리수익률 CAGR(%)",
    "mdd": "최대낙폭 MDD(%)", "sharpe": "샤프(연)", "sortino": "소르티노(연)",
    "calmar": "칼마", "n_trades": "거래 수", "win_rate": "승률(%)",
    "avg_hold": "평균 보유일", "avg_trade_return": "평균 거래수익률(%)",
    "bench_total": "벤치마크 누적(%)", "bench_cagr": "벤치마크 CAGR(%)",
    "bench_mdd": "벤치마크 MDD(%)", "excess_return": "초과수익(%)",
    "var_95": "VaR 95(%)", "cvar_95": "CVaR 95(%)", "beta": "베타",
    "profit_factor": "손익비 PF", "payoff_ratio": "페이오프",
}


def _thin_border() -> Border:
    s = XlSide(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _num(v: Any):
    """엔진 값 → 엑셀 셀 값. NaN/None은 빈칸으로(엑셀에서 비교 시 노이즈 방지)."""
    if v is None:
        return None
    try:
        f = float(v)
        return None if f != f else f        # NaN → None
    except (TypeError, ValueError):
        return v


def _ir_dict(ir: Any) -> dict:
    if hasattr(ir, "model_dump"):
        return ir.model_dump(mode="json")
    return ir if isinstance(ir, dict) else {}


def _input_symbols_by_role(ir, dataset: dict) -> dict[str, list[str]]:
    """입력 데이터를 역할별로 — 출력(거래종목)이 아니라 '전략이 소비한 데이터 전부'.

    엔진은 dataset을 needed_symbols(=universe ∪ 신호 참조)로 적재하므로 universe를 뺀 나머지
    dataset 키 = 신호로 참조되나 거래하지 않는 심볼(크로스에셋 등 — 기존 누락 부류).
      subject : universe(거래·분석 대상) — 원자료에 OHLC.
      signal  : dataset − subject (신호 참조) — 원자료에 종가.
    (벤치마크는 엔진이 대상에서 파생 → dataset 부재. 펀더멘털 등 ③참조는 후속 Phase.)
    """
    def _in_ds(s: str) -> bool:
        df = dataset.get(s)
        return isinstance(df, pd.DataFrame) and not df.empty and "Close" in df.columns
    subject: list[str] = []
    try:
        uni = (_ir_dict(ir).get("universe") or {}).get("symbols") or []
        subject = [str(s) for s in uni if _in_ds(str(s))]
    except Exception:   # noqa: BLE001 — IR 파싱 실패해도 dataset 기준 진행(증빙은 best-effort)
        subject = []
    sset = set(subject)
    signal = [str(s) for s in dataset if str(s) not in sset and _in_ds(str(s))]
    return {"subject": subject[:_MAX_SYMS], "signal": signal[:_MAX_SYMS]}


def build_strategy_excel(
    ir: StrategyIR | dict,
    dataset: dict[str, pd.DataFrame],
    result: dict,
    *,
    name: str | None = None,
) -> bytes:
    """인사이트 엔진 분석 결과 → 증빙 .xlsx 바이트. **결과 형상별 디스패치.**

    ir: 분석한 StrategyIR(또는 dict). dataset: {symbol: OHLCV}. result: run_query/
    strategy_from_spec 의 *원본* dict(pandas 객체 포함). 형상(SIMULATE·SELECT·DESCRIBE·
    SWEEP·PERIOD·EXTREMIZE·RELATE·EVENT·SIGNAL)에 따라 전용 시트 구성으로 직렬화한다.
    """
    disp = name or (_ir_dict(ir).get("name") or "전략")
    q, axis = result.get("query"), result.get("axis")
    report, reduction = result.get("report"), result.get("reduction")
    # SIMULATE 1회 백테스트 — equity 보유 & 펼침/리서치 아님(axis 우선 판정: condition도 equity를
    # 들고 다니므로 axis가 있으면 simulate로 오인하지 않는다 — #169 동일 교훈).
    if isinstance(result.get("equity"), pd.Series) and not axis and not q:
        return _build_simulate(ir, dataset, result, disp)
    if q == "select":
        return _build_select(ir, dataset, result, disp)
    if report == "single":
        return _build_describe_single(ir, dataset, result, disp)
    if report == "portfolio":
        return _build_describe_portfolio(ir, dataset, result, disp)
    if reduction == "extremize":          # axis=parameter/asset + 최적화 — axis 분기보다 먼저
        return _build_extremize(ir, dataset, result, disp)
    if axis == "period_split":
        return _build_period_split(ir, dataset, result, disp)
    if axis in ("parameter", "asset"):
        return _build_sweep(ir, dataset, result, disp)
    if axis == "condition":
        return _build_condition(ir, dataset, result, disp)
    if axis == "signal":
        return _build_signal_dist(ir, dataset, result, disp)
    if axis == "relation":                # ic / regression
        return _build_relation(ir, dataset, result, disp)
    if axis == "time":                    # event study
        return _build_event(ir, dataset, result, disp)
    raise ValueError(f"build_strategy_excel: 미지원 결과 형상 (query={q}, axis={axis}, report={report})")


# ── 신호 트리 → Excel 수식 컴파일러 (L2 — 신호·결정 라이브 패널) ────────────────
# 엔진 indicators.py 정의와 1:1인 공통 지표만 셀 수식으로 재현. 미지원 op/지표·다중대상은 패널
# 생략(폴백) — 틀린 "라이브 수식"은 없느니만 못함(4원칙 검증). 체결 지연·청산 룰은 엔진 소관이라
# 패널은 '신호→포지션 결정'까지만 재현하고, 실제 체결은 거래내역이 정본임을 명시한다.
_IND_LAG = {"price_level": 0, "Close": 0, "close": 0, "Open": 0, "High": 0, "Low": 0,
            "pct_change_1d": 1, "pct_change_5d": 5, "pct_change_20d": 20,
            "pct_change_252d": 252, "log_return_1d": 1,
            # 윈도우 지표(F2) — 룩백 = window-1 (rolling min_periods=window). 셀 수식은 원자료 Close 윈도우.
            "ma_dev_20d": 19, "ma_dev_60d": 59, "ma_dev_200d": 199,
            "bb_width": 19, "bb_pct": 19, "ma_gap_20_60": 59}
_CMP_EXCEL = {"<=": "<=", ">=": ">=", "<": "<", ">": ">", "==": "=", "!=": "<>",
              "le": "<=", "ge": ">=", "lt": "<", "gt": ">", "eq": "=", "ne": "<>"}
_AR_EXCEL = {"add": "+", "sub": "-", "mul": "*", "div": "/"}
# 원자료에 그대로 있는 가격 레벨(윈도우 ts_* 연산이 셀 범위로 재현 가능한 입력).
_RAW_LEVELS = {"price_level", "Close", "close", "Open", "High", "Low"}
# 윈도우 시계열 op(F3) → Excel 범위함수. 엔진 expression_parser 정의와 1:1(std=ddof1=STDEV.S).
_TS_FN = {"ts_mean": "AVERAGE", "ts_sum": "SUM", "ts_std": "STDEV.S", "ts_max": "MAX", "ts_min": "MIN"}


def _split_ref(ref: str, subjects: list[str]) -> tuple[str | None, str | None]:
    """'SYM.지표'/'__SELF__.지표'/'지표' → (sym, ind). 심볼 없음·__SELF__→첫 subject. 미인식 지표→(None,None)."""
    s = str(ref)
    if "." in s:
        sym, _, tail = s.rpartition(".")
        if sym == "__SELF__":   # 엔진 자기참조 토큰 — 패널 resolver도 단일대상 subject로 해석(없으면 무효)
            sym = subjects[0] if subjects else None
        return (sym, tail) if (tail in _IND_LAG and sym) else (None, None)
    if s in _IND_LAG:
        return (subjects[0] if subjects else None), s
    return None, None


def _ind_excel(ind: str, col: str, r: int) -> str | None:
    """지원 지표를 원자료 col열·행 r 기준 Excel 수식('=' 포함)으로. 엔진 indicators.py와 동일 정의."""
    raw = f"원자료!{col}"
    if ind in ("price_level", "Close", "close", "Open", "High", "Low"):
        return f"={raw}{r}"
    if ind in ("pct_change_1d", "pct_change_5d", "pct_change_20d", "pct_change_252d"):
        return f"=({raw}{r}/{raw}{r - _IND_LAG[ind]}-1)*100"
    if ind == "log_return_1d":
        return f"=LN({raw}{r}/{raw}{r - 1})*100"
    # ── 윈도우 지표(F2): 원자료 Close 윈도우로 엔진 indicators.py와 1:1 재현 ──
    if ind in ("ma_dev_20d", "ma_dev_60d", "ma_dev_200d"):
        w = {"ma_dev_20d": 20, "ma_dev_60d": 60, "ma_dev_200d": 200}[ind]
        avg = f"AVERAGE({raw}{r - w + 1}:{raw}{r})"
        return f"=({raw}{r}/{avg}-1)*100"                       # (Close-MA)/MA*100
    if ind in ("bb_width", "bb_pct"):                            # window=20·k=2
        avg = f"AVERAGE({raw}{r - 19}:{raw}{r})"
        std = f"STDEV.S({raw}{r - 19}:{raw}{r})"
        if ind == "bb_width":
            return f"=(4*{std})/{avg}*100"                       # 2k·std/MA*100, k=2
        return f"=({raw}{r}-{avg}+2*{std})/(4*{std})"           # (Close-lower)/(upper-lower)
    if ind == "ma_gap_20_60":
        a20 = f"AVERAGE({raw}{r - 19}:{raw}{r})"
        a60 = f"AVERAGE({raw}{r - 59}:{raw}{r})"
        return f"=({a20}-{a60})/{a60}*100"
    return None


def _collect_data_refs(node, out: list[str]) -> None:
    if not isinstance(node, dict):
        return
    if node.get("op") == "data":
        out.append(str((node.get("params") or {}).get("ref")))
    for v in (node.get("inputs") or {}).values():
        _collect_data_refs(v, out)


def _signal_to_excel(node, ref_to_col, raw_of, r):
    """signal 트리 → 행 r의 Excel 수식 본문(셀 참조 확정). data 잎→패널 신호열·윈도우 op→원자료
    범위. 미지원 op/잎→None(패널 폴백). r은 정수 행번호(윈도우·전일 참조 위해 행별 컴파일)."""
    if not isinstance(node, dict):
        return None
    op = node.get("op")
    p = node.get("params") or {}
    ins = node.get("inputs") or {}
    if op == "const":
        try:
            return repr(float(p.get("value")))
        except (TypeError, ValueError):
            return None
    if op == "data":
        col = ref_to_col.get(str(p.get("ref")))
        return f"{col}{r}" if col else None
    if op == "compare":
        lt = _signal_to_excel(ins.get("left"), ref_to_col, raw_of, r)
        rt = _signal_to_excel(ins.get("right"), ref_to_col, raw_of, r)
        ox = _CMP_EXCEL.get(str(p.get("op")))
        return f"({lt}{ox}{rt})" if (lt and rt and ox) else None
    if op == "select":
        c = _signal_to_excel(ins.get("cond"), ref_to_col, raw_of, r)
        a = _signal_to_excel(ins.get("a"), ref_to_col, raw_of, r)
        b = _signal_to_excel(ins.get("b"), ref_to_col, raw_of, r)
        return f"IF({c},{a},{b})" if (c and a and b) else None
    if op in _AR_EXCEL:
        a = _signal_to_excel(ins.get("a") or ins.get("left"), ref_to_col, raw_of, r)
        b = _signal_to_excel(ins.get("b") or ins.get("right"), ref_to_col, raw_of, r)
        return f"({a}{_AR_EXCEL[op]}{b})" if (a and b) else None
    if op == "logic":                                  # F3 — AND/OR (엔진: 슬롯 자연정렬 결합)
        keys = sorted(ins.keys(), key=lambda k: (len(k), k))
        parts = [_signal_to_excel(ins[k], ref_to_col, raw_of, r) for k in keys]
        if not parts or any(x is None for x in parts):
            return None
        fn = "OR" if str(p.get("logic", "AND")).upper() == "OR" else "AND"
        return f"{fn}({','.join(parts)})"
    if op == "cross":                                  # F3 — 돌파(오늘 교차 + 어제 반대편)
        cl = _signal_to_excel(ins.get("left"), ref_to_col, raw_of, r)
        cr = _signal_to_excel(ins.get("right"), ref_to_col, raw_of, r)
        pl = _signal_to_excel(ins.get("left"), ref_to_col, raw_of, r - 1)
        pr = _signal_to_excel(ins.get("right"), ref_to_col, raw_of, r - 1)
        if not (cl and cr and pl and pr):
            return None
        if str(p.get("direction", "up")).lower() == "down":
            return f"AND({cl}<{cr},{pl}>{pr})"
        return f"AND({cl}>{cr},{pl}<{pr})"
    if op in _TS_FN:                                   # F3 — 윈도우 시계열(MA·채널 등)
        rng = _ts_range(ins.get("signal"), raw_of, r, int(p.get("window", 20)))
        return f"{_TS_FN[op]}({rng})" if rng else None
    if op in ("ts_delay", "ts_delta"):
        info = _ts_raw(ins.get("signal"), raw_of)
        if not info:
            return None
        col, w = info[0], int(p.get("window", 20))
        return f"원자료!{col}{r - w}" if op == "ts_delay" else f"(원자료!{col}{r}-원자료!{col}{r - w})"
    return None


def _ts_raw(inner, raw_of):
    """ts_* 의 inner가 원자료 가격 레벨 data 잎이면 (원자료 col, ind), 아니면 None(폴백)."""
    if not isinstance(inner, dict) or inner.get("op") != "data":
        return None
    info = raw_of.get(str((inner.get("params") or {}).get("ref")))
    if not info or info[2] not in _RAW_LEVELS:
        return None
    return info[0], info[2]


def _ts_range(inner, raw_of, r, w):
    """윈도우 [r-w+1, r]의 원자료 범위 문자열 (inner가 가격 레벨일 때만)."""
    info = _ts_raw(inner, raw_of)
    return f"원자료!{info[0]}{r - w + 1}:원자료!{info[0]}{r}" if info else None


def _signal_lag(node, raw_of) -> int:
    """신호 트리가 요구하는 최대 룩백(행) — 윈도우·전일 참조가 유효하도록 패널 시작행을 결정."""
    if not isinstance(node, dict):
        return 0
    op = node.get("op")
    ins = node.get("inputs") or {}
    p = node.get("params") or {}
    if op == "data":
        info = raw_of.get(str((p or {}).get("ref")))
        return info[1] if info else 0
    if op == "const":
        return 0
    if op == "cross":
        return 1 + max([_signal_lag(ins.get("left"), raw_of),
                        _signal_lag(ins.get("right"), raw_of)], default=0)
    if op in _TS_FN:
        return (int(p.get("window", 20)) - 1) + _signal_lag(ins.get("signal"), raw_of)
    if op in ("ts_delay", "ts_delta"):
        return int(p.get("window", 20)) + _signal_lag(ins.get("signal"), raw_of)
    return max([_signal_lag(v, raw_of) for v in ins.values()], default=0)


def _signal_panel(wb: "Workbook", ird: dict, equity, roles: dict,
                  raw_col_map: dict, disp: str) -> str | None:
    """신호·결정 라이브 패널 — 신호값·포지션을 입력(원자료)에서 셀로 독립 재현.

    렌더 성공→None, 불가→**사유 문자열**(F4: 지표·설명에 구체 표면화). 단일 대상 전략·지원 지표·
    지원 op만(다중대상·미지원 지표/연산은 셀 재현 불가 → 생략). 틀린 수식은 없느니만 못함(원칙1)."""
    sig = ird.get("signal")
    subjects = list(roles.get("subject") or [])
    if not isinstance(sig, dict):
        return "신호 정의 없음"
    if len(subjects) != 1:
        return f"다중 대상({len(subjects)}종목) — 패널은 단일 종목 전략만 재현(종목별 결정을 1열로 표현 불가)"
    refs: list[str] = []
    _collect_data_refs(sig, refs)
    refs = list(dict.fromkeys(refs))
    if not refs:
        return "신호에 데이터 참조 없음"
    raw_of: dict[str, tuple[str, int, str]] = {}   # ref → (원자료 col, lag, 지표명)
    for ref in refs:
        sym, ind = _split_ref(ref, subjects)
        if ind is None:
            return f"미지원 지표 — '{ref}'는 셀 수식 재현 대상 아님(RSI·ATR·변동성·모멘텀·펀더멘털 등)"
        field = ind if ind in ("Open", "High", "Low") else "Close"
        col = raw_col_map.get((sym, field))
        if col is None:
            return f"원자료 열 없음 — '{ref}'"
        raw_of[ref] = (col, _IND_LAG.get(ind, 0), ind)
    ref_to_col = {ref: get_column_letter(2 + i) for i, ref in enumerate(refs)}
    n = len(equity)
    first = 6
    maxlag = _signal_lag(sig, raw_of)
    if n <= maxlag:
        return f"룩백 부족 — 데이터 {n}행 < 필요 {maxlag + 1}행"
    if _signal_to_excel(sig, ref_to_col, raw_of, first + maxlag) is None:
        return "미지원 연산 — 신호 트리에 셀 재현 불가 연산 포함(ts_rank·횡단순위·라벨·modifier 등)"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    italic = Font(italic=True, color=_NOTE)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _thin_border()
    center = Alignment(horizontal="center")
    idx_list = list(equity.index)
    pos_col_i = 2 + len(refs)
    ps = wb.create_sheet("신호·결정")
    ps.cell(1, 1, f"신호·결정 (라이브 재현) — {disp}").font = title_font
    ps.cell(2, 1, "신호값·포지션을 입력(원자료)에서 셀 수식으로 독립 재현 — 엔진과 별개 검산. "
                  "실제 체결(신호→주문 지연·청산 룰)은 '거래내역'이 정본. 점수 양수=롱·음수=숏·0=관망.").font = italic
    headers = ["날짜", *[f"{ref} (신호값)" for ref in refs], "포지션 점수(룰)"]
    for j, h in enumerate(headers):
        c = ps.cell(5, 1 + j, h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center
    for i in range(n):
        r = first + i
        ps.cell(r, 1, pd.Timestamp(idx_list[i]).to_pydatetime()).number_format = "yyyy-mm-dd"
        if r - maxlag < first:
            continue   # 룩백 부족(엔진 NaN 구간) → 공란
        for j, ref in enumerate(refs):
            col, _lg, ind = raw_of[ref]
            ps.cell(r, 2 + j, _ind_excel(ind, col, r)).number_format = "0.0000"
        ps.cell(r, pos_col_i, "=" + _signal_to_excel(sig, ref_to_col, raw_of, r)).number_format = "0.00"
    ps.column_dimensions["A"].width = 12
    for j in range(len(refs) + 1):
        ps.column_dimensions[get_column_letter(2 + j)].width = 18
    ps.freeze_panes = "B6"
    return None


def _simulate_sheets(wb: "Workbook", ir, dataset, result, disp: str) -> None:
    """SIMULATE 증빙 시트(백테스트 라이브수식·거래내역·원자료·신호·결정·지표·설명)를 wb에 추가.

    period_split(연도별)도 재사용 — 연도 분석은 단일 자산곡선을 연도로 자른 것이라 원자료+라이브
    수식+거래 검증 시트가 동일하게 성립(파라미터 스윕 같은 N개 독립 시뮬과 다름)."""
    equity = result.get("equity")
    if not isinstance(equity, pd.Series) or equity.empty:
        raise ValueError("_simulate_sheets: result['equity'] (pd.Series) 필요.")
    metrics = result.get("metrics") or {}

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    accent_font = Font(bold=True, color=_ACCENT)
    italic = Font(italic=True, color=_NOTE)
    input_fill = PatternFill("solid", fgColor=_INPUT_BG)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _thin_border()
    center = Alignment(horizontal="center")

    # ── 시트: 백테스트 (엔진 자산곡선 + 라이브 정의식 수식) ──────────────────────
    ws = wb.create_sheet("백테스트")
    n = len(equity)
    last = _ROW_D0 + n - 1

    ws.cell(_ROW_TITLE, 1, f"{disp} — 백테스트 증빙 (라이브 수식)").font = title_font

    # 입력칸: 일일 추가비용(bps). 0이면 엔진 그대로, >0이면 사후 비용 민감도.
    ws.cell(_ROW_INPUT, 1, "일일 추가비용 (bps)").font = bold
    inp = ws.cell(_ROW_INPUT, 2, 0)
    inp.fill = input_fill
    inp.border = border
    inp.alignment = center
    note = ws.cell(_ROW_INPUT, 3,
                   "← 노란 칸을 바꾸면 '조정' 열·지표가 라이브 재계산 (0 = 엔진과 동일)")
    note.font = italic

    # 지표 비교: 엔진(정본 값) vs 엑셀(자산곡선 위 정의식 수식). bps=0이면 일치 = 증빙.
    for j, h in enumerate(("지표", "엔진", "엑셀(라이브수식)", "비고")):
        c = ws.cell(_ROW_SUMHEAD, 1 + j, h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center
    sums = [
        ("누적수익률(%)", _num(metrics.get("total_return")),
         f"=(E{last}/E{_ROW_D0}-1)*100", "bps=0이면 엔진과 일치"),
        ("연복리수익률 CAGR(%)", _num(metrics.get("cagr")),
         f"=IFERROR(((E{last}/E{_ROW_D0})^(252/COUNT(E{_ROW_D0}:E{last}))-1)*100,\"\")", ""),
        ("샤프(연)", _num(metrics.get("sharpe")),
         f"=IFERROR(AVERAGE(D{_ROW_D0 + 1}:D{last})/STDEV.S(D{_ROW_D0 + 1}:D{last})*SQRT(252),\"\")",
         "엔진=일별 표본 std"),
        ("최대낙폭 MDD(%)", _num(metrics.get("mdd")),
         f"=MIN(F{_ROW_D0}:F{last})*100", ""),
        ("연변동성(%)", "—",
         f"=IFERROR(STDEV.S(D{_ROW_D0 + 1}:D{last})*SQRT(252)*100,\"\")", "엔진 미보고"),
    ]
    for i, (label, eng, formula, memo) in enumerate(sums):
        r = _ROW_SUM0 + i
        ws.cell(r, 1, label).font = accent_font
        ce = ws.cell(r, 2, eng)
        ce.number_format = "0.00"
        ce.border = border
        cf = ws.cell(r, 3, formula)
        cf.number_format = "0.00"
        cf.border = border
        ws.cell(r, 4, memo).font = italic

    # 데이터 + 수식 헤더
    dheaders = ["날짜", "자산(엔진)", "일수익률", "조정수익률", "조정자산", "낙폭"]
    for j, h in enumerate(dheaders):
        c = ws.cell(_ROW_DHEAD, 1 + j, h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    idx_list = list(equity.index)
    val_list = [float(v) for v in equity.to_numpy()]
    for i in range(n):
        r = _ROW_D0 + i
        ws.cell(r, 1, pd.Timestamp(idx_list[i]).to_pydatetime())  # A 날짜
        ws.cell(r, 2, val_list[i])                                # B 자산(엔진, 정본 값)
        if i == 0:
            # 첫 행: 수익률 없음. 조정자산 시작 = 엔진 자산(겹쳐 그려져 bps=0 검증).
            ws.cell(r, 5, f"=B{r}")                                # E 조정자산 시작
            ws.cell(r, 6, f"=E{r}/MAX(E${_ROW_D0}:E{r})-1")       # F 낙폭(=0)
        else:
            ws.cell(r, 3, f"=B{r}/B{r - 1}-1")                    # C 일수익률(엔진 자산)
            ws.cell(r, 4, f"=C{r}-$B${_ROW_INPUT}/10000")         # D 조정수익률(−bps)
            ws.cell(r, 5, f"=E{r - 1}*(1+D{r})")                  # E 조정자산(복리)
            ws.cell(r, 6, f"=E{r}/MAX(E${_ROW_D0}:E{r})-1")       # F 낙폭

    for i in range(n):
        r = _ROW_D0 + i
        ws.cell(r, 1).number_format = "yyyy-mm-dd"
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "0.00%"
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 5).number_format = "#,##0"
        ws.cell(r, 6).number_format = "0.00%"
    for col, w in zip("ABCDEF", (12, 14, 11, 11, 14, 10)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["C"].width = 18   # 비고 노트 가독
    ws.freeze_panes = f"A{_ROW_D0}"

    # ── 시트 2: 거래내역 (엔진 trades — 정적 값, 검증 앵커) ──────────────────────
    tdf = result.get("trades")
    tr = wb.create_sheet("거래내역")
    n_tr = len(tdf) if isinstance(tdf, pd.DataFrame) else 0
    tr.cell(1, 1, f"거래내역 — {disp} (거래 {n_tr}건, 엔진 결과 정적 값)").font = title_font
    tr.cell(2, 1, "거래 수").font = accent_font
    tr.cell(2, 2, n_tr).font = bold
    tr.cell(2, 3, "← 엔진 백테스트가 산출한 확정 거래(라이브 수식 아님 — 검증 기준)").font = italic
    t_cols = ["종목", "진입일", "청산일", "보유일", "진입가", "청산가", "수익률(%)", "청산사유"]
    for j, h in enumerate(t_cols):
        c = tr.cell(4, 1 + j, h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center
    if isinstance(tdf, pd.DataFrame) and n_tr:
        for i, (_, row) in enumerate(tdf.iterrows()):
            r = 5 + i
            tr.cell(r, 1, str(row.get("종목", "")))
            tr.cell(r, 2, pd.Timestamp(row["진입일"]).to_pydatetime()
                    if pd.notna(row.get("진입일")) else None).number_format = "yyyy-mm-dd"
            tr.cell(r, 3, pd.Timestamp(row["청산일"]).to_pydatetime()
                    if pd.notna(row.get("청산일")) else None).number_format = "yyyy-mm-dd"
            tr.cell(r, 4, int(row.get("보유일")) if pd.notna(row.get("보유일")) else None)
            tr.cell(r, 5, round(float(row["진입가"]), 4)
                    if pd.notna(row.get("진입가")) else None).number_format = "#,##0.00"
            tr.cell(r, 6, round(float(row["청산가"]), 4)
                    if pd.notna(row.get("청산가")) else None).number_format = "#,##0.00"
            tr.cell(r, 7, round(float(row["수익률(%)"]), 4)
                    if pd.notna(row.get("수익률(%)")) else None).number_format = "+0.00;-0.00"
            tr.cell(r, 8, str(row.get("청산사유", "")))
    for col, w in zip("ABCDEFGH", (10, 12, 12, 8, 11, 11, 10, 14)):
        tr.column_dimensions[col].width = w
    tr.freeze_panes = "A5"

    # ── 시트 3: 원자료 (입력 데이터 — 대상 OHLC + 신호 종가, 역할 태깅) ──────────
    # ★ 재설계(excel-export-redesign): 출력(거래종목 _held_symbols)이 아니라 '전략이 쓴 데이터
    #   전부'(입력). 신호 참조 종목(크로스에셋 등)은 거래 안 해도 입력이므로 포함 — 신호 검증 가능.
    roles = _input_symbols_by_role(ir, dataset)
    raw = wb.create_sheet("원자료")
    raw.cell(1, 1, f"원자료 — 입력 데이터 ({disp})").font = title_font
    raw.cell(2, 1, "전략이 소비한 데이터 전부. 대상=OHLC(거래·마킹 기준)·신호=종가(결정 입력). 거래 안 해도 신호면 포함.").font = italic
    # 컬럼 계획: 날짜 | 대상 OHLC… | 신호 종가… (역할 라벨)
    plan: list[tuple[str, str, str]] = []   # (헤더라벨, 심볼, OHLC 필드)
    for s in roles["subject"]:
        for field, fl in (("Open", "시가"), ("High", "고가"), ("Low", "저가"), ("Close", "종가")):
            plan.append((f"{s} {fl}(대상)", s, field))
    for s in roles["signal"]:
        plan.append((f"{s} 종가(신호)", s, "Close"))
    raw.cell(5, 1, "날짜").font = bold
    raw.cell(5, 1).fill = head_fill
    raw.cell(5, 1).border = border
    series: dict[int, pd.Series] = {}
    for j, (label, s, field) in enumerate(plan):
        c = raw.cell(5, 2 + j, label)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center
        df = dataset.get(s)
        if isinstance(df, pd.DataFrame) and field in df.columns:
            series[j] = df[field].reindex(equity.index)
    for i in range(n):
        r = 6 + i
        raw.cell(r, 1, pd.Timestamp(idx_list[i]).to_pydatetime()).number_format = "yyyy-mm-dd"
        for j in range(len(plan)):
            ser = series.get(j)
            v = ser.iloc[i] if ser is not None else None
            cell = raw.cell(r, 2 + j, None if v is None or pd.isna(v) else float(v))
            cell.number_format = "#,##0.00"
    raw.column_dimensions["A"].width = 12
    for j in range(len(plan)):
        raw.column_dimensions[get_column_letter(2 + j)].width = 12
    raw.freeze_panes = "B6"

    # ── 시트: 신호·결정 (L2 — 신호값·포지션을 원자료에서 라이브 재현, 공통 지표·단일대상 한정) ──
    raw_col_map = {(s, field): get_column_letter(2 + j)
                   for j, (_lab, s, field) in enumerate(plan)}
    panel_reason = _signal_panel(wb, _ir_dict(ir), equity, roles, raw_col_map, disp)

    # ── 시트 4: 일별비중 (엔진 EOD 기여 비중) · 패널 있을 때만 생성 ───────────────
    # 단일종목/방향성 백테스트는 엔진이 weight 패널을 내지 않음 → 빈 시트 대신 생략(죽은 시트 제거).
    wpanel = result.get("weight")
    if isinstance(wpanel, pd.DataFrame) and not wpanel.empty:
        wp_syms = [str(c) for c in wpanel.columns][:_MAX_SYMS]
        wp = wb.create_sheet("일별비중")
        wp.cell(1, 1, f"일별 비중 — 엔진 EOD 기여 패널 ({disp})").font = title_font
        wp.cell(2, 1, "양수=롱·음수=숏 (자기자본 대비 비중). 엔진이 기록한 일별 포지션.").font = italic
        wp.cell(4, 1, "날짜").font = bold
        wp.cell(4, 1).fill = head_fill
        wp.cell(4, 1).border = border
        for j, s in enumerate(wp_syms):
            c = wp.cell(4, 2 + j, s)
            c.font = bold
            c.fill = head_fill
            c.border = border
            c.alignment = center
        wp_al = wpanel.reindex(equity.index)
        for i in range(n):
            r = 5 + i
            wp.cell(r, 1, pd.Timestamp(idx_list[i]).to_pydatetime()).number_format = "yyyy-mm-dd"
            for j, s in enumerate(wp_syms):
                v = wp_al[s].iloc[i] if s in wp_al.columns else None
                cell = wp.cell(r, 2 + j, None if v is None or pd.isna(v) else float(v))
                cell.number_format = "0.00%"
        wp.column_dimensions["A"].width = 12
        for j in range(len(wp_syms)):
            wp.column_dimensions[get_column_letter(2 + j)].width = 11
        wp.freeze_panes = "B5"

    # ── 시트 5: 지표·설명 (엔진 metrics 전체 + 전략 정의 + 방법론) ───────────────
    info = wb.create_sheet("지표·설명")
    info.cell(1, 1, f"{disp} — 엔진 지표 · 전략 정의 · 방법론").font = title_font
    info.cell(3, 1, "엔진 지표 (정본)").font = accent_font
    rr = 4
    for k, v in metrics.items():
        info.cell(rr, 1, _METRIC_LABELS.get(k, k)).font = bold
        val = _num(v)
        cell = info.cell(rr, 2, val)
        if isinstance(val, float):
            cell.number_format = "0.0000"
        rr += 1

    rr += 1
    info.cell(rr, 1, "전략 정의 (IR)").font = accent_font
    rr += 1
    ird = _ir_dict(ir)
    for key in ("name", "query", "universe", "signal", "position", "simulation", "study"):
        if key not in ird or ird[key] in (None, {}, []):
            continue
        info.cell(rr, 1, key).font = bold
        v = ird[key]
        text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, separators=(",", ":"))
        if len(text) > 900:
            text = text[:900] + " …"
        info.cell(rr, 2, text).alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    rr += 1
    info.cell(rr, 1, "방법론 · 한계").font = accent_font
    rr += 1
    notes = [
        ("라이브 수식", "엔진 정본 자산곡선(B열) 위에서 일수익·낙폭·CAGR·샤프·MDD를 정의식으로 재계산 — bps=0이면 엔진 지표와 일치(증빙)."),
        ("일일 추가비용(bps)", "백테스트 시트 B3. 일별 수익률에서 차감하는 사후 비용 민감도(라이브). 엔진 비용모델 변경이 아님."),
        ("거래내역", "엔진이 산출한 확정 거래(정적 값). 진입가·청산가·수익률은 엔진과 byte 일치 — 검증 앵커."),
        ("원자료(입력 데이터)", "전략이 소비한 데이터 전부 — 대상은 OHLC, 신호 참조 종목(거래 안 해도)은 종가. 신호종목 가격으로 신호 로직을 직접 수동검증."),
        ("신호·결정 패널", "신호값·포지션 룰을 원자료에서 셀 수식으로 독립 재현(공통 지표·단일대상 한정). "
         "엔진 실집행(신호→체결 지연·청산 룰)은 거래내역이 정본 — 패널은 결정 로직 검산용. ▸ " + (
             "이 전략은 패널 생성됨(아래 '신호·결정' 시트)." if panel_reason is None
             else f"이 전략은 패널 생략 — {panel_reason}.")),
        ("엑셀이 복제 못 하는 것", "엔진은 정수주·현금·마진콜·지연체결의 이벤트 NAV 시뮬레이션 — 셀 수식으로 NAV 경로를 정확히 재현하지 않는다(자산곡선은 엔진 값 그대로 사용)."),
        ("전략 로직 변경", "임계값·기간·종목 등 전략 파라미터를 바꿔 다시 보려면 '변수 조정'(엔진 재실행) — 엑셀은 산술·비용 검증 도구."),
    ]
    for k, v in notes:
        info.cell(rr, 1, k).font = bold
        info.cell(rr, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    for w in (result.get("warnings") or []):     # 데이터 품질·실행 경고도 엑셀에 표면화(재설계 원칙 #5)
        msg = w.get("message") if isinstance(w, dict) else str(w)
        if msg:
            rr += 1
            info.cell(rr, 1, "⚠").font = bold
            info.cell(rr, 2, msg).alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 90


def _build_simulate(ir, dataset, result, disp: str) -> bytes:
    """SIMULATE(1회 백테스트) → 증빙 .xlsx (P1 — 데이터+라이브수식)."""
    wb = Workbook()
    wb.remove(wb.active)                      # 기본 빈 시트 제거 — 시트는 _simulate_sheets가 생성
    _simulate_sheets(wb, ir, dataset, result, disp)
    return _save(wb)


# ═════════════════════════════════════════════════════════════════════════════
# P2 — 형상별 빌더 (select·describe·sweep·period·extremize·relate·event·signal)
#
# 메타분석(스윕·기간·최적화·관계·이벤트·분포)은 **'감사표'**: 엔진이 산출한 버킷/윈도
# 통계를 그대로 직렬화하고 사용 IR·방법론을 함께 실어 '무엇을 어떤 파라미터로 계산했는지'
# 증빙한다(라이브 수식은 한 시트로 재현 불가 — 각 버킷이 별도 시뮬/추정량). DESCRIBE 종목·
# 포트폴리오는 데이터→통계를 셀로 재현 가능한 곳(52주 고저·연변동성, HHI·가중밸류)만 라이브.
# ═════════════════════════════════════════════════════════════════════════════

# PERF_KEYS 표시 순서 (key, 라벨, 숫자서식) — 모든 성과 감사표 공용.
_PERF_COLS = [
    ("n", "표본수", "#,##0"), ("cagr", "CAGR(%)", "0.00"), ("cum_return", "누적수익(%)", "0.00"),
    ("sharpe", "샤프", "0.00"), ("sortino", "소르티노", "0.00"), ("mdd", "MDD(%)", "0.00"),
    ("mean", "평균일수익(%)", "0.0000"), ("std", "변동(%)", "0.0000"), ("win_rate", "승률(%)", "0.0"),
    ("payoff_ratio", "페이오프", "0.00"), ("profit_factor", "손익비", "0.00"),
    ("var_95", "VaR95(%)", "0.00"), ("cvar_95", "CVaR95(%)", "0.00"),
]


def _styles() -> dict:
    return {
        "bold": Font(bold=True), "title": Font(bold=True, size=14, color="20201D"),
        "accent": Font(bold=True, color=_ACCENT), "italic": Font(italic=True, color=_NOTE),
        "head_fill": PatternFill("solid", fgColor=_HEAD_BG), "border": _thin_border(),
        "center": Alignment(horizontal="center"), "wrap": Alignment(wrap_text=True, vertical="top"),
    }


def _title(ws, text: str, S: dict) -> None:
    ws.cell(1, 1, text).font = S["title"]


def _header(ws, row: int, headers: list, S: dict) -> None:
    for j, h in enumerate(headers):
        c = ws.cell(row, 1 + j, h)
        c.font = S["bold"]
        c.fill = S["head_fill"]
        c.border = S["border"]
        c.alignment = S["center"]


def _perf_table(ws, row0: int, label_header: str, rows: list, S: dict) -> int:
    """rows=[(label, perf_dict)]. PERF_KEYS 열로 성과 감사표. 다음 행 반환."""
    _header(ws, row0, [label_header] + [c[1] for c in _PERF_COLS], S)
    r = row0 + 1
    for label, perf in rows:
        ws.cell(r, 1, str(label)).font = S["bold"]
        for j, (k, _lbl, fmt) in enumerate(_PERF_COLS):
            cell = ws.cell(r, 2 + j, _num((perf or {}).get(k)))
            cell.number_format = fmt
        r += 1
    ws.column_dimensions["A"].width = 18
    for j in range(len(_PERF_COLS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    return r


def _warning_block(info, rr: int, result, S: dict) -> int:
    """결과의 데이터 품질·실행 경고(Phase 0.5·무거래 등)를 설명 시트에 표면화. 다음 행 반환.

    ★ 재설계 원칙 #5: 데이터 품질 계약이 챗 응답뿐 아니라 엑셀(오프라인 검토 surface)에도
    닿아야 한다 — 전 연산유형 공통. 경고 없으면 무동작.
    """
    warns = (result or {}).get("warnings") or []
    msgs = [w.get("message") if isinstance(w, dict) else str(w) for w in warns]
    msgs = [m for m in msgs if m]
    if not msgs:
        return rr
    info.cell(rr, 1, "⚠ 데이터 품질·실행 경고").font = S["accent"]
    rr += 1
    for m in msgs:
        info.cell(rr, 1, "⚠").font = S["bold"]
        info.cell(rr, 2, m).alignment = S["wrap"]
        rr += 1
    return rr + 1


def _methodology(wb, ir, title: str, notes: list, S: dict, result=None):
    """'설명' 시트 — (데이터 경고) + 전략 정의(IR) + 방법론·한계. 전 형상 공용 증빙 꼬리표."""
    info = wb.create_sheet("설명")
    info.cell(1, 1, title).font = S["title"]
    rr = _warning_block(info, 3, result, S)   # 경고 먼저(있으면) — 데이터 품질 계약 표면화
    info.cell(rr, 1, "전략 정의 (IR)").font = S["accent"]
    rr += 1
    ird = _ir_dict(ir)
    for key in ("name", "query", "universe", "signal", "position", "simulation", "study", "select"):
        if key not in ird or ird[key] in (None, {}, []):
            continue
        info.cell(rr, 1, key).font = S["bold"]
        v = ird[key]
        text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, separators=(",", ":"))
        info.cell(rr, 2, text[:900] + (" …" if len(text) > 900 else "")).alignment = S["wrap"]
        rr += 1
    rr += 1
    info.cell(rr, 1, "방법론 · 한계").font = S["accent"]
    rr += 1
    for k, v in notes:
        info.cell(rr, 1, k).font = S["bold"]
        info.cell(rr, 2, v).alignment = S["wrap"]
        rr += 1
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 92
    return info


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_select(ir, dataset, result, disp: str) -> bytes:
    """SELECT(스크리닝 랭킹) — as-of 스냅샷 순위표(값)."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "스크리닝결과"
    _title(ws, f"{disp} — 스크리닝 결과 (as-of {result.get('as_of', '')})", S)
    ws.cell(2, 1, f"유니버스 {result.get('universe_size', '?')}종목 · 자격통과 "
                  f"{result.get('eligible_size', '?')}종목 · 상위 {len(result.get('results', []))} 표시").font = S["italic"]
    results = result.get("results", []) or []
    disp_cols = list((results[0].get("metrics") or {}).keys()) if results else []
    _header(ws, 4, ["순위", "종목", "점수(score)", "섹터"] + disp_cols, S)
    r = 5
    for i, row in enumerate(results, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, str(row.get("symbol", "")))
        ws.cell(r, 3, _num(row.get("score"))).number_format = "#,##0.0000"
        ws.cell(r, 4, str(row.get("sector", "")))
        for j, c in enumerate(disp_cols):
            ws.cell(r, 5 + j, _num((row.get("metrics") or {}).get(c))).number_format = "#,##0.0000"
        r += 1
    for col, w in zip("ABCD", (6, 12, 14, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    _methodology(wb, ir, f"{disp} — 스크리닝 방법론", [
        ("점수(score)", "신호 블록의 as-of(기준일) 횡단 값. descending=참이면 큰 순, 거짓이면 작은 순(예: 저PBR)."),
        ("선별", f"top_n/top_pct 상위만 표시. as-of={result.get('as_of', '')} 단면 스냅샷(시계열 아님)."),
        ("값 정직성", "점수·표시지표는 엔진이 산출한 그 시점 값. 팩터 산식 자체는 IR signal 트리(설명 시트) 참조."),
    ], S, result=result)
    return _save(wb)


def _build_describe_single(ir, dataset, result, disp: str) -> bytes:
    """DESCRIBE 단일종목(360) — 요약값 + 원자료 종가로 52주·변동성 라이브 검증."""
    S = _styles()
    sym = result.get("symbol", "")
    wb = Workbook()
    ws = wb.active
    ws.title = "종목요약"
    _title(ws, f"{disp} — {sym} 종목 분석 (as-of {result.get('as_of', '')})", S)
    price = result.get("price", {}) or {}
    rets = price.get("returns", {}) or {}
    risk = result.get("risk", {}) or {}
    fund = result.get("fundamentals", {}) or {}
    rows = [
        ("섹터", result.get("sector", ""), "@"), ("데이터 일수", result.get("data_points"), "#,##0"),
        ("현재가", price.get("last"), "#,##0.00"), ("1개월 수익률(%)", rets.get("1m"), "0.00"),
        ("3개월(%)", rets.get("3m"), "0.00"), ("6개월(%)", rets.get("6m"), "0.00"),
        ("12개월(%)", rets.get("12m"), "0.00"), ("52주 고가", price.get("high_52w"), "#,##0.00"),
        ("52주 저가", price.get("low_52w"), "#,##0.00"), ("52주 고가대비(%)", price.get("pct_from_52w_high"), "0.00"),
        ("연변동성(%)", risk.get("vol_annualized"), "0.00"), ("최대낙폭(%)", risk.get("max_drawdown"), "0.00"),
        ("PBR", fund.get("pb_ratio"), "#,##0.00"), ("PER", fund.get("trailing_pe"), "#,##0.00"),
        ("EV/EBITDA", fund.get("ev_ebitda"), "#,##0.00"),
    ]
    r = 3
    for k, v, fmt in rows:
        ws.cell(r, 1, k).font = S["bold"]
        c = ws.cell(r, 2, v if fmt == "@" else _num(v))
        if fmt != "@":
            c.number_format = fmt
        r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    df = dataset.get(sym)
    if isinstance(df, pd.DataFrame) and "Close" in df.columns:
        raw = wb.create_sheet("원자료")
        _title(raw, f"{sym} 일별 종가 — 52주 고저·연변동성 라이브 검증", S)
        _header(raw, 3, ["날짜", "종가", "일수익률"], S)
        closes = df["Close"].dropna()
        idx, vals = list(closes.index), [float(v) for v in closes.to_numpy()]
        for i in range(len(idx)):
            rr = 4 + i
            raw.cell(rr, 1, pd.Timestamp(idx[i]).to_pydatetime()).number_format = "yyyy-mm-dd"
            raw.cell(rr, 2, vals[i]).number_format = "#,##0.00"
            if i > 0:
                raw.cell(rr, 3, f"=B{rr}/B{rr - 1}-1").number_format = "0.00%"
        last = 4 + len(idx) - 1
        w0 = max(last - 251, 4)                      # 최근 ~252영업일(52주)
        ws.cell(r + 1, 1, "라이브 검증 (원자료 수식)").font = S["accent"]
        live = [
            ("52주 고가 =MAX", f"=MAX(원자료!B{w0}:B{last})", "#,##0.00"),
            ("52주 저가 =MIN", f"=MIN(원자료!B{w0}:B{last})", "#,##0.00"),
            ("연변동성(%) =STDEV.S×√252", f"=STDEV.S(원자료!C{w0 + 1}:C{last})*SQRT(252)*100", "0.00"),
        ]
        for i, (lbl, formula, fmt) in enumerate(live):
            ws.cell(r + 2 + i, 1, lbl).font = S["bold"]
            ws.cell(r + 2 + i, 2, formula).number_format = fmt
        raw.column_dimensions["A"].width = 12
        raw.freeze_panes = "A4"
    _methodology(wb, ir, f"{disp} — 종목 분석 방법론", [
        ("값/수식", "요약 지표는 엔진 산출값. '원자료' 종가로 52주 고저·연변동성을 라이브 수식으로 재계산해 대조(엔진값과 일치 확인용)."),
        ("수익률", "1/3/6/12개월 = 해당 영업일수 전 종가 대비 변화(%)."),
        ("펀더멘털", "데이터에 있으면 표기, 없으면 빈칸 — 정직(가짜 채움 없음)."),
    ], S, result=result)
    return _save(wb)


def _build_describe_portfolio(ir, dataset, result, disp: str) -> bytes:
    """DESCRIBE 포트폴리오 — 보유표 + HHI·가중밸류 라이브 수식."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "보유진단"
    _title(ws, f"{disp} — 포트폴리오 진단 (as-of {result.get('as_of', '')}, "
               f"{result.get('n_holdings', '?')}종목)", S)
    holdings = result.get("holdings", []) or []
    _header(ws, 3, ["종목", "비중", "섹터", "PBR(as-of)"], S)
    r = 4
    for h in holdings:
        sym = str(h.get("symbol", ""))
        ws.cell(r, 1, sym).font = S["bold"]
        ws.cell(r, 2, _num(h.get("weight"))).number_format = "0.00%"
        ws.cell(r, 3, str(h.get("sector", "")))
        pb = None
        df = dataset.get(sym)
        if isinstance(df, pd.DataFrame) and "pb_ratio" in df.columns:
            s = df["pb_ratio"].dropna()
            pb = float(s.iloc[-1]) if len(s) else None
        ws.cell(r, 4, _num(pb)).number_format = "#,##0.00"
        r += 1
    last = r - 1
    conc = result.get("concentration", {}) or {}
    val = result.get("valuation", {}) or {}
    rsk = result.get("risk", {}) or {}
    r += 1
    _header(ws, r, ["집중도·밸류 지표", "엑셀(라이브수식)", "엔진"], S)
    r += 1
    live = [
        ("HHI =SUMPRODUCT(비중²)", f"=SUMPRODUCT(B4:B{last},B4:B{last})", conc.get("hhi"), "0.0000"),
        ("유효종목수 =1/HHI", f"=IFERROR(1/SUMPRODUCT(B4:B{last},B4:B{last}),\"\")", conc.get("effective_n"), "0.00"),
        ("최대비중 =MAX", f"=MAX(B4:B{last})", conc.get("top_weight"), "0.00%"),
        ("가중PBR =SUMPRODUCT", f"=IFERROR(SUMPRODUCT(B4:B{last},D4:D{last})/SUM(B4:B{last}),\"\")",
         val.get("weighted_pb"), "#,##0.00"),
    ]
    for lbl, formula, eng, fmt in live:
        ws.cell(r, 1, lbl).font = S["accent"]
        ws.cell(r, 2, formula).number_format = fmt
        ws.cell(r, 3, _num(eng)).number_format = fmt
        r += 1
    r += 1
    ws.cell(r, 1, "포트폴리오 연변동성(%)").font = S["bold"]
    ws.cell(r, 2, _num(rsk.get("portfolio_vol_annualized"))).number_format = "0.00"
    r += 1
    ws.cell(r, 1, "평균 쌍상관").font = S["bold"]
    ws.cell(r, 2, _num(rsk.get("avg_pairwise_corr"))).number_format = "0.00"
    sx = result.get("sector_exposure", {}) or {}
    if sx:
        se = wb.create_sheet("섹터노출")
        _header(se, 1, ["섹터", "비중"], S)
        rr = 2
        for k, v in sx.items():
            se.cell(rr, 1, str(k))
            se.cell(rr, 2, _num(v)).number_format = "0.00%"
            rr += 1
        se.column_dimensions["A"].width = 18
    for col, w in zip("ABCD", (24, 16, 14, 12)):
        ws.column_dimensions[col].width = w
    _methodology(wb, ir, f"{disp} — 포트폴리오 진단 방법론", [
        ("라이브 수식", "HHI·유효종목수·최대비중·가중PBR을 보유 비중/PBR로 라이브 재계산(엔진값과 대조). 비중 셀 바꾸면 갱신."),
        ("PBR(as-of)", "각 보유종목 데이터의 마지막 PBR. 없으면 빈칸."),
        ("위험(값)", "포트 연변동성·평균 쌍상관은 일별수익 공분산 기반 엔진 산출값(셀 복제 범위 외)."),
    ], S, result=result)
    return _save(wb)


def _build_sweep(ir, dataset, result, disp: str) -> bytes:
    """SWEEP parameter/entity — 버킷별 성과 감사표(값)."""
    S = _styles()
    axis = result.get("axis")
    label_h = {"parameter": "파라미터값", "asset": "종목"}.get(axis, "버킷")
    wb = Workbook()
    ws = wb.active
    ws.title = "스윕결과"
    sub = " · ".join(a.get("path", "") for a in result.get("axes", [])) if axis == "parameter" else ""
    _title(ws, f"{disp} — {label_h}별 성과 스윕" + (f" ({sub})" if sub else ""), S)
    ws.cell(2, 1, f"각 {label_h} = 독립 백테스트. 아래는 엔진 산출 성과(값).").font = S["italic"]
    rows = list((result.get("buckets") or {}).items())
    _perf_table(ws, 4, label_h, rows, S)
    ws.freeze_panes = "A5"
    _methodology(wb, ir, f"{disp} — 스윕 방법론", [
        ("스윕", f"{label_h} 축을 펼쳐 각 값마다 독립 백테스트를 돌린 성과 비교."),
        ("값 only(감사표)", "각 버킷은 별도 시뮬이라 한 시트 라이브 수식으로 재현 불가 — 엔진 산출 성과를 그대로 표기."),
        ("성과 정의", "CAGR·샤프·MDD 등은 각 버킷 일별수익에서 산출(PERF 정규지표)."),
    ], S, result=result)
    return _save(wb)


def _build_condition(ir, dataset, result, disp: str) -> bytes:
    """SWEEP label(조건 대조) — 전체+라벨별 성과 + 쌍대 t검정 감사표."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "조건별성과"
    _title(ws, f"{disp} — 조건(라벨)별 성과 대조", S)
    rows = [("전체(포트폴리오)", result.get("overall"))] + list((result.get("buckets") or {}).items())
    nxt = _perf_table(ws, 3, "조건", rows, S)
    pw = ((result.get("compare") or {}).get("pairwise")) or {}
    if pw:
        ws.cell(nxt + 1, 1, "쌍대 비교 (2-표본 t검정)").font = S["accent"]
        _header(ws, nxt + 2, ["비교쌍", "평균A(%)", "평균B(%)", "차이(%)", "t값", "p값", "nA", "nB"], S)
        r = nxt + 3
        for pair, t in pw.items():
            ws.cell(r, 1, str(pair))
            for j, k in enumerate(("mean_a", "mean_b", "mean_diff", "t_stat", "p_value", "n_a", "n_b")):
                ws.cell(r, 2 + j, _num((t or {}).get(k))).number_format = "0.0000" if k == "p_value" else "0.00"
            r += 1
    ws.freeze_panes = "A4"
    _methodology(wb, ir, f"{disp} — 조건 대조 방법론", [
        ("조건 대조", "라벨(섹터·국면 등)로 일별 기여를 분할해 그룹별 성과를 비교. '전체'=포트폴리오 합산."),
        ("쌍대 t검정", "각 라벨 쌍의 일별수익 평균 차이 유의성(2-표본 t)."),
        ("값 only(감사표)", "엔진 산출 버킷 성과·검정통계를 그대로 표기."),
    ], S, result=result)
    return _save(wb)


def _year_of(key) -> "int | None":
    """버킷 키('2021'·'2021Q1'·'2010-01-01~…')에서 연도 추출. 실패 None — 라이브 수식 행매핑용."""
    import re
    m = re.search(r"\d{4}", str(key))
    return int(m.group()) if m else None


def _build_period_split(ir, dataset, result, disp: str) -> bytes:
    """PERIOD_SPLIT(연도별) → **원자료+라이브수식** 증빙 .xlsx.

    연도 분석 = 단일 자산곡선을 연도로 자른 것 → simulate 증빙 시트(원자료·백테스트 라이브수식·
    거래내역·일별비중·지표·설명)를 재사용 + **연도별 집계 라이브 수식** 시트. 각 연도 수익을
    '백테스트' 시트 조정자산 비율(=E[연말]/E[직전]−1)로 재계산해 엔진값과 대조 → 원자료→일별→연도를
    셀 수식으로 수동 검증(IB/PE급). (P2 감사표[값만]를 대체 — 파라미터 스윕 등 N개 독립 시뮬과 달리
    단일 곡선이라 라이브 성립.)
    """
    from .run import run_strategy_ir
    S = _styles()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("연도별성과")        # 헤드라인 시트(먼저)
    _title(ws, f"{disp} — 연도별 성과 (단일 자산곡선 분할 · 라이브 수식)", S)
    ws.cell(2, 1, "엑셀 연수익 = '백테스트' 시트 조정자산 비율(라이브 수식). 엔진값과 일치 = 증빙.").font = S["italic"]

    # base 백테스트 재실행 → 일별 자산곡선(원자료·수식 시트 + 연도 집계 수식의 기준).
    base = None
    try:
        sir = ir if hasattr(ir, "model_dump") else StrategyIR.model_validate(ir)
        base = run_strategy_ir(sir, dataset)
    except Exception:
        base = None
    eq = base.get("equity") if isinstance(base, dict) and base.get("success") else None
    year_rows: dict = {}
    if isinstance(eq, pd.Series) and not eq.empty:
        for i, ts in enumerate(eq.index):
            year_rows.setdefault(pd.Timestamp(ts).year, []).append(_ROW_D0 + i)

    _header(ws, 4, ["연도", "엔진 누적(%)", "엑셀 연수익(%) 라이브", "CAGR(%)", "샤프", "MDD(%)", "표본", "비고"], S)
    r = 5
    for key, b in (result.get("buckets") or {}).items():
        b = b or {}
        ws.cell(r, 1, str(key)).font = S["bold"]
        ws.cell(r, 2, _num(b.get("cum_return"))).number_format = "0.00"
        rows = year_rows.get(_year_of(key))
        if rows:
            prior = rows[0] - 1 if rows[0] - 1 >= _ROW_D0 else _ROW_D0
            ws.cell(r, 3, f"=(백테스트!E{rows[-1]}/백테스트!E{prior}-1)*100").number_format = "0.00"
        else:
            ws.cell(r, 3, "—")
        ws.cell(r, 4, _num(b.get("cagr"))).number_format = "0.00"
        ws.cell(r, 5, _num(b.get("sharpe"))).number_format = "0.00"
        ws.cell(r, 6, _num(b.get("mdd"))).number_format = "0.00"
        ws.cell(r, 7, _num(b.get("n"))).number_format = "#,##0"
        if b.get("n") and not b.get("std") and not b.get("cum_return"):
            ws.cell(r, 8, "무거래(데이터·신호 결손 가능)").font = S["italic"]
        r += 1
    cons = result.get("consistency", {}) or {}
    ws.cell(r + 1, 1, f"일관성: 양(+) {cons.get('positive_folds', '?')}/{cons.get('n_folds', '?')} 구간").font = S["accent"]
    for w_ in (result.get("warnings") or []):
        r += 1
        ws.cell(r + 1, 1, "⚠ " + (w_.get("message", str(w_)) if isinstance(w_, dict) else str(w_))).font = S["italic"]
    for col, wd in zip("ABCDEFGH", (10, 13, 22, 10, 8, 10, 8, 28)):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A5"

    if eq is not None:
        _simulate_sheets(wb, ir, dataset, base, disp)     # 원자료+백테스트 라이브수식+거래+비중+지표·설명
    else:
        _methodology(wb, ir, f"{disp} — 연도별 방법론", [
            ("연수익(라이브)", "각 연도 = 조정자산[연말]/조정자산[직전]−1. base 백테스트 재실행 실패로 일별 시트 생략."),
            ("값(엔진)", "엔진 산출 연도 성과를 그대로 표기(대조 기준)."),
        ], S, result=result)
    return _save(wb)


def _build_extremize(ir, dataset, result, disp: str) -> bytes:
    """EXTREMIZE — 후보 랭킹 + 최적 성과 + OOS 가드 감사표."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "최적화결과"
    obj = result.get("objective", {}) or {}
    best = result.get("best", {}) or {}
    _title(ws, f"{disp} — 최적화 (목적: {obj.get('metric', 'sharpe')} {obj.get('direction', 'max')})", S)
    ws.cell(2, 1, f"최적: {best.get('label', '')} (목적값 "
                  f"{_num(best.get('metric_value'))})").font = S["accent"]
    _header(ws, 4, ["순위", "후보", f"목적값({obj.get('metric', '')})"], S)
    r = 5
    for i, row in enumerate(result.get("ranked", []) or [], 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, str(row.get("label", "")))
        ws.cell(r, 3, _num(row.get("metric_value"))).number_format = "0.0000"
        r += 1
    r += 1
    ws.cell(r, 1, "최적 후보 성과").font = S["accent"]
    nxt = _perf_table(ws, r + 1, "최적", [(best.get("label", ""), best.get("perf"))], S)
    og = result.get("oos_guard", {}) or {}
    if og.get("buckets"):
        ws2 = wb.create_sheet("OOS검증")
        _title(ws2, "과최적화 가드 — 시간폴드 OOS 일관성", S)
        _perf_table(ws2, 3, "구간", list(og["buckets"].items()), S)
        c2 = og.get("consistency", {}) or {}
        ws2.cell(3 + len(og["buckets"]) + 2, 1,
                 f"일관성(양구간 비율): {_num(c2.get('consistency'))}").font = S["bold"]
    elif og.get("error"):
        ws.cell(nxt + 1, 1, f"OOS 가드: {og['error']}").font = S["italic"]
    for col, w in zip("ABC", (6, 16, 16)):
        ws.column_dimensions[col].width = w
    _methodology(wb, ir, f"{disp} — 최적화 방법론", [
        ("최적화", f"검색공간(axis={result.get('axis')})을 펼쳐 목적함수({obj.get('metric', '')} "
                 f"{obj.get('direction', '')})를 최대/최소화하는 후보 선택."),
        ("과최적화 가드", "in-sample 최적을 시간폴드 OOS 일관성으로 교차검증(oos_guard)."),
        ("값 only(감사표)", "엔진 산출 랭킹·성과를 그대로 표기."),
    ], S, result=result)
    return _save(wb)


def _build_signal_dist(ir, dataset, result, disp: str) -> bytes:
    """DESCRIBE signal-dist — 신호값 분포(통계·분위수·히스토그램) 감사표."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "신호분포"
    _title(ws, f"{disp} — 신호값 분포", S)
    ov = result.get("overall", {}) or {}
    r = 3
    for k, lbl, fmt in [("n", "표본수", "#,##0"), ("mean", "평균", "0.0000"), ("std", "표준편차", "0.0000"),
                        ("skew", "왜도", "0.00"), ("kurtosis", "첨도", "0.00")]:
        ws.cell(r, 1, lbl).font = S["bold"]
        ws.cell(r, 2, _num(ov.get(k))).number_format = fmt
        r += 1
    q = ov.get("quantiles", {}) or {}
    if q:
        ws.cell(r, 1, "분위수").font = S["accent"]
        r += 1
        for k in ("q05", "q10", "q25", "q50", "q75", "q90", "q95"):
            ws.cell(r, 1, k).font = S["bold"]
            ws.cell(r, 2, _num(q.get(k))).number_format = "0.0000"
            r += 1
    hc, he = ov.get("hist_counts"), ov.get("hist_edges")
    if hc and he:
        hs = wb.create_sheet("히스토그램")
        _header(hs, 1, ["구간 하한", "구간 상한", "빈도"], S)
        for i, c in enumerate(hc):
            hs.cell(2 + i, 1, _num(he[i])).number_format = "0.0000"
            hs.cell(2 + i, 2, _num(he[i + 1] if i + 1 < len(he) else None)).number_format = "0.0000"
            hs.cell(2 + i, 3, _num(c)).number_format = "#,##0"
        for col in ("A", "B", "C"):
            hs.column_dimensions[col].width = 12
    ws.column_dimensions["A"].width = 14
    _methodology(wb, ir, f"{disp} — 신호 분포 방법론", [
        ("신호 분포", "분석 노드(target_node)의 횡단·시계열 값 전체 분포 — 평균·분위수·히스토그램."),
        ("값 only(감사표)", "엔진 산출 분포 통계를 그대로 표기."),
    ], S, result=result)
    return _save(wb)


def _build_relation(ir, dataset, result, disp: str) -> bytes:
    """RELATE ic/regression — 윈도별 IC 또는 Fama-MacBeth 회귀 감사표."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    bw = result.get("by_window", {}) or {}
    if result.get("relation") == "ic":
        ws.title = "IC분석"
        _title(ws, f"{disp} — IC 분석 (신호↔미래수익 횡단 상관)", S)
        _header(ws, 3, ["윈도(일)", "표본수", "평균IC(%)", "t값", "p값", "양(+)비율(%)", "IR(평균/표준편차)"], S)
        r = 4
        for w in result.get("windows", []) or []:
            ovw = (bw.get(w) or {}).get("overall", {}) or {}
            ws.cell(r, 1, str(w))
            for j, (k, fmt) in enumerate([("n", "#,##0"), ("mean", "0.0000"), ("t_stat", "0.00"),
                                          ("p_value", "0.0000"), ("prob_positive", "0.0")]):
                ws.cell(r, 2 + j, _num(ovw.get(k))).number_format = fmt
            ws.cell(r, 7, _num((bw.get(w) or {}).get("ir"))).number_format = "0.00"
            r += 1
        notes = [("IC", "각 기준일의 신호와 forward수익의 횡단 상관계수. 평균IC>0 & t유의 = 예측력."),
                 ("IR", "평균IC/IC표준편차 — 신호 안정성."),
                 ("값 only(감사표)", "Fama-MacBeth t검정 등 엔진 산출값을 그대로 표기.")]
        ws.freeze_panes = "A4"
    else:  # regression
        ws.title = "회귀분석"
        _title(ws, f"{disp} — 다중팩터 횡단 회귀 (Fama-MacBeth)", S)
        ws.cell(2, 1, "설명변수: " + ", ".join(result.get("factor_names", []) or [])).font = S["italic"]
        _header(ws, 4, ["윈도(일)", "구간수", "팩터", "계수(coef)", "표준오차(se)", "t값", "95%CI 하한", "95%CI 상한"], S)
        r = 5
        for w in result.get("windows", []) or []:
            wd = bw.get(w) or {}
            for f in (wd.get("factors") or []):
                ws.cell(r, 1, str(w))
                ws.cell(r, 2, _num(wd.get("n_periods"))).number_format = "#,##0"
                ws.cell(r, 3, str(f.get("name", "")))
                for j, k in enumerate(("coef", "se", "t_stat", "ci_low", "ci_high")):
                    ws.cell(r, 4 + j, _num(f.get(k))).number_format = "0.0000"
                r += 1
        notes = [("Fama-MacBeth", "각 기준일 횡단 OLS(forward수익 ~ 팩터들) → 시계열 평균 계수·t."),
                 ("계수 해석", "양(+)·t유의 = 그 팩터가 미래수익과 양의 관계."),
                 ("값 only(감사표)", "엔진 산출 회귀통계를 그대로 표기.")]
        ws.freeze_panes = "A5"
    _methodology(wb, ir, f"{disp} — 관계분석 방법론", notes, S, result=result)
    return _save(wb)


def _build_event(ir, dataset, result, disp: str) -> bytes:
    """RELATE event-study — 윈도별 이벤트 후 수익 분포(MAE/MFE 포함) 감사표."""
    S = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "이벤트분석"
    _title(ws, f"{disp} — 이벤트 스터디 (기준 {result.get('basis', 'close')}, "
               f"{result.get('n_events', '?')}건)", S)
    cols = [("n", "표본", "#,##0"), ("mean", "평균수익(%)", "0.00"), ("t_stat", "t값", "0.00"),
            ("p_value", "p값", "0.0000"), ("prob_positive", "양(+)비율(%)", "0.0"),
            ("mean_mae", "평균MAE(%)", "0.00"), ("worst_mae", "최악MAE(%)", "0.00"),
            ("mean_mfe", "평균MFE(%)", "0.00"), ("payoff_ratio", "페이오프", "0.00")]
    _header(ws, 3, ["윈도(일)"] + [c[1] for c in cols], S)
    r = 4
    for w in result.get("windows", []) or []:
        ev = (result.get("overall") or {}).get(w, {}) or {}
        ws.cell(r, 1, str(w))
        for j, (k, _l, fmt) in enumerate(cols):
            ws.cell(r, 2 + j, _num(ev.get(k))).number_format = fmt
        r += 1
    ws.freeze_panes = "A4"
    _methodology(wb, ir, f"{disp} — 이벤트 스터디 방법론", [
        ("이벤트 스터디", "이벤트(조건 충족) 발생 후 forward 윈도 수익 분포. MAE=구간내 최대손실, MFE=최대이익."),
        ("값 only(감사표)", "엔진 산출 이벤트 통계를 그대로 표기."),
    ], S, result=result)
    return _save(wb)
