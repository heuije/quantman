"""선물 분석 → 라이브 수식 엑셀(.xlsx) 생성 (종목 일반화).

목적: 대시보드의 백테스트 로직을 *엑셀 함수로 연결된* 파일로 내보내,
사용자가 엑셀에서 임계값·보유기간·비용·롤비용을 바꾸면 즉시 재계산되게 한다.

설계:
- 입력칸(B2~B7): 방향/임계값/보유기간/롤비용·슬리피지·수수료 → 셀 참조로 전 수식 반영.
- 계약사양(H2·H3): multiplier·tick — 종목별 ContractSpec 주입(앱과 동일 계수).
- 데이터+수식 시트: 신호(장중 고저점 첫 터치)·진입(다음날 시가)·청산(보유기간
  후 종가)·롤횟수(보유 중 통과 만기 수)·수익률·PnL을 전부 =수식으로.
- 요약(D2~E6): COUNTIF/AVERAGE/SUM.
- 거래내역 시트: 다운로드 시점 파라미터의 실제 run_backtest 결과를 정적 값으로
  (수식 손상 방지). 라이브 실험은 백테스트 시트, 확정 내역은 이 시트.

비용모델: 앱(run_backtest)과 동일 — 슬리피지=slippage_ticks×tick(절대 $),
수수료=commission_per_contract($/계약, 진입+청산 양레그), 롤비용=roll_cost_pct
(%/롤, 양수=콘탱고 비용). 엑셀 숫자가 앱 백테스트와 일치한다.

한계(엑셀 단순화): SL/TP 조기청산·MAE/MFE·walk-forward는 미반영(보유기간 고정).
임계·보유기간·비용·롤 전략의 라이브 재계산 도구다.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side as XlSide
from openpyxl.utils import get_column_letter

from .backtest import (
    ContractSpec,
    CostModel,
    ExitRules,
    RollModel,
    WTI_SPEC,
    run_backtest,
    wti_expiry_dates,
)
from .signals import generate_signals


# 디자인 토큰 (대시보드 팔레트와 통일)
_ACCENT = "D97757"
_HEAD_BG = "F7ECE5"
_INPUT_BG = "FFF4CC"
_SPEC_BG = "EFEAE3"
_BORDER = "E8E3DB"


def _thin_border() -> Border:
    s = XlSide(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def build_oil_excel(
    df: pd.DataFrame,
    *,
    side: str = "short",
    threshold: float = 100.0,
    horizon_days: int = 120,
    spec: ContractSpec = WTI_SPEC,
    commission_per_contract: float = 2.5,
    slippage_ticks: int = 1,
    roll_cost_pct: float = 0.0,
    min_gap_days: int = 0,
    name: str = "선물",
    currency: str = "USD",
    unit: str = "",
) -> bytes:
    """선물 OHLC DataFrame → 라이브 수식 .xlsx 바이트.

    df: prepare_wti 출력 (date ASC, open/high/low/close/volume).
    spec: 종목 계약사양(tick/multiplier) — 앱과 동일 손익 계수.
    side/threshold/... : 엑셀 입력칸 초기값 (열어서 바꾸면 재계산).
    name/currency/unit: 표시 라벨(종목명·손익 통화·가격 단위).
    """
    df = df.sort_values("date").reset_index(drop=True)
    mult = float(spec.multiplier)
    tick = float(spec.tick)
    cur_sym = {"USD": "$", "KRW": "₩"}.get(currency, currency)
    unit_sfx = f"/{unit}" if unit else ""    # 가격 단위 라벨 (예: $/배럴)
    disp = name if "선물" in name else f"{name} 선물"   # 이름에 '선물' 중복 방지

    wb = Workbook()
    ws = wb.active
    ws.title = "백테스트"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    accent_font = Font(bold=True, color=_ACCENT)
    italic = Font(italic=True, color="6F6A62")
    input_fill = PatternFill("solid", fgColor=_INPUT_BG)
    spec_fill = PatternFill("solid", fgColor=_SPEC_BG)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _thin_border()
    center = Alignment(horizontal="center")

    # ── 제목 ─────────────────────────────────────────────────────────
    ws["A1"] = f"{disp} — 라이브 백테스트"
    ws["A1"].font = title_font

    # ── 입력칸 (B2~B7) — 노란 칸: 바꾸면 전체 재계산 ─────────────────
    inputs = [
        ("방향 (short/long)", side),
        (f"임계값 ({cur_sym}{unit_sfx})", threshold),
        ("보유기간 (영업일)", horizon_days),
        ("롤비용 (%/롤, 예 0.5)", roll_cost_pct * 100),
        ("슬리피지 (틱/체결)", slippage_ticks),
        (f"수수료 ({cur_sym}/계약·레그)", commission_per_contract),
    ]
    for i, (label, val) in enumerate(inputs):
        r = 2 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = bold
        c = ws[f"B{r}"]
        c.value = val
        c.fill = input_fill
        c.border = border
        c.alignment = center
    ws["C2"] = "← 노란 칸을 바꾸면 전체가 재계산됩니다"
    ws["C2"].font = italic

    # ── 계약사양 (G2~H4) — 수식이 참조하는 상수(종목별 ContractSpec) ──
    spec_rows = [
        ("계약배수", mult),
        (f"틱 크기 ({cur_sym})", tick),
        ("손익 통화", currency),
    ]
    for i, (label, val) in enumerate(spec_rows):
        r = 2 + i
        ws[f"G{r}"] = label
        ws[f"G{r}"].font = bold
        cell = ws[f"H{r}"]
        cell.value = val
        cell.fill = spec_fill
        cell.border = border
        cell.alignment = center

    # ── 쿨타임 입력 (노란 칸 $H$5) — 신호 후 N영업일 같은 임계 신호 무시 ──
    ws["G5"] = "쿨타임 (영업일, 0=없음)"
    ws["G5"].font = bold
    cd = ws["H5"]
    cd.value = min_gap_days
    cd.fill = input_fill
    cd.border = border
    cd.alignment = center

    # ── 요약 (D2~E6) ─────────────────────────────────────────────────
    n = len(df)
    last = 8 + n            # 데이터는 9행부터 → 마지막 데이터 행
    summary = [
        ("신호 횟수", f"=COUNTIF(F9:F{last},1)", "0"),
        ("거래 성립", f"=COUNT(L9:L{last})", "0"),
        ("승률", f'=IFERROR(COUNTIF(L9:L{last},">0")/COUNT(L9:L{last}),0)', "0.0%"),
        ("평균 수익률", f"=IFERROR(AVERAGE(L9:L{last}),0)", "+0.00%;-0.00%"),
        (f"누적 PnL ({cur_sym}, 1계약)", f"=SUM(M9:M{last})", "#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(summary):
        r = 2 + i
        ws[f"D{r}"] = label
        ws[f"D{r}"].font = accent_font
        cell = ws[f"E{r}"]
        cell.value = formula
        cell.font = bold
        cell.number_format = fmt
        cell.border = border

    # ── 데이터 + 수식 헤더 (8행) ─────────────────────────────────────
    # A날짜 B시가 C고가 D저가 E종가 F신호 | G진입일 H청산일 I진입가 J청산가
    # K롤횟수 | L수익률 M PnL N유효진입가 O유효청산가
    headers = ["날짜", "시가", "고가", "저가", "종가", "신호",
               "진입일", "청산일", "진입가", "청산가", "롤횟수",
               "수익률", f"PnL({cur_sym})", "유효진입가", "유효청산가"]
    for j, h in enumerate(headers):
        c = ws.cell(row=8, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    # slip(절대 $) = 슬리피지틱($B$6) × tick($H$3). mult=$H$2.
    slip = "$B$6*$H$3"
    for idx in range(n):
        r = 9 + idx
        row = df.iloc[idx]
        ws.cell(row=r, column=1, value=pd.Timestamp(row["date"]).to_pydatetime())
        ws.cell(row=r, column=2, value=float(row["open"]))
        ws.cell(row=r, column=3, value=float(row["high"]))
        ws.cell(row=r, column=4, value=float(row["low"]))
        ws.cell(row=r, column=5, value=float(row["close"]))

        # F 신호: 첫터치(전일 비교 히스테리시스) ∧ 쿨타임($H$5) — 직전 (쿨타임−1)영업일 내
        #   유효신호 있으면 무시. F가 곧 '유효신호'라 G~J(진입/청산)는 그대로 F 참조.
        #   윈도우=쿨−1 (generate_signals 'i−last>=min_gap'와 일치, off-by-one 주의).
        #   OFFSET 높이는 MIN(쿨−1, r−1)로 클램프(row1 위 #REF 방지; 헤더행은 신호 0이라 무해).
        raw = (f'IF($B$2="short",AND(C{r}>=$B$3,C{r-1}<$B$3),'
               f'AND(D{r}<=$B$3,D{r-1}>$B$3))')
        win = f'MIN($H$5-1,{r - 1})'
        cool = (f'IF($H$5<=1,1,IF(COUNTIF(OFFSET(F{r},-{win},0,{win},1),1)=0,1,""))')
        ws.cell(row=r, column=6, value=f'=IF({raw},{cool},"")')
        # G 진입일 = 신호 다음 영업일 날짜
        ws.cell(row=r, column=7,
                value=f'=IF(F{r}=1,IFERROR(OFFSET(A{r},1,0),""),"")')
        # H 청산일 = 진입 후 보유기간 영업일 날짜
        ws.cell(row=r, column=8,
                value=f'=IF(F{r}=1,IFERROR(OFFSET(A{r},1+$B$4,0),""),"")')
        # I 진입가 = 다음날 시가
        ws.cell(row=r, column=9,
                value=f'=IF(F{r}=1,IFERROR(OFFSET(B{r},1,0),""),"")')
        # J 청산가 = 진입 후 보유기간 종가
        ws.cell(row=r, column=10,
                value=f'=IF(F{r}=1,IFERROR(OFFSET(E{r},1+$B$4,0),""),"")')
        # K 롤횟수 = 진입일(G)~청산일(H) 사이 만기일 수 (진입<만기≤청산) — 앱과 동일
        ws.cell(row=r, column=11, value=(
            f'=IF(OR(G{r}="",H{r}=""),"",IFERROR(COUNTIFS('
            f'만기일!$A:$A,">"&G{r},'
            f'만기일!$A:$A,"<="&H{r}'
            f'),""))'
        ))
        # N 유효진입가 (앱: long=entry+slip, short=entry−slip)
        ws.cell(row=r, column=14, value=(
            f'=IF(I{r}="","",IF($B$2="long",I{r}+{slip},I{r}-{slip}))'
        ))
        # O 유효청산가 (앱: long=exit−slip, short=exit+slip)
        ws.cell(row=r, column=15, value=(
            f'=IF(J{r}="","",IF($B$2="long",J{r}-{slip},J{r}+{slip}))'
        ))
        # M PnL = sign·(유효청산−유효진입)·mult − 2·수수료 + 롤비용
        #   롤비용(≤0) = −(K·롤%·진입대금) − K·(2·수수료 + 2·slip·mult)   [롤%≠0·K>0]
        ws.cell(row=r, column=13, value=(
            f'=IF(OR(N{r}="",O{r}=""),"",'
            f'(IF($B$2="long",O{r}-N{r},N{r}-O{r}))*$H$2-2*$B$7'
            f'+IF(OR($B$5=0,K{r}=""),0,'
            f'-(K{r}*($B$5/100)*(I{r}*$H$2))-K{r}*(2*$B$7+2*{slip}*$H$2)))'
        ))
        # L 수익률 (앱 Trade.return_pct 정의): 원가격 수익률(sign) + 롤비용/진입대금.
        #   수수료·슬리피지는 PnL($)에만 반영(앱과 동일 — 수익률은 gross+롤).
        ws.cell(row=r, column=12, value=(
            f'=IF(OR(I{r}="",J{r}=""),"",'
            f'IF($B$2="long",J{r}/I{r}-1,-(J{r}/I{r}-1))'
            f'+IF(OR($B$5=0,K{r}=""),0,'
            f'(-(K{r}*($B$5/100)*(I{r}*$H$2))-K{r}*(2*$B$7+2*{slip}*$H$2))/(I{r}*$H$2)))'
        ))

    # 포맷
    for idx in range(n):
        r = 9 + idx
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"      # A 날짜
        ws.cell(row=r, column=7).number_format = "yyyy-mm-dd"      # G 진입일
        ws.cell(row=r, column=8).number_format = "yyyy-mm-dd"      # H 청산일
        ws.cell(row=r, column=12).number_format = "+0.00%;-0.00%"  # L 수익률
        ws.cell(row=r, column=13).number_format = "#,##0"          # M PnL
        ws.cell(row=r, column=14).number_format = "#,##0.00"       # N 유효진입가
        ws.cell(row=r, column=15).number_format = "#,##0.00"       # O 유효청산가
        for col in (2, 3, 4, 5, 9, 10):                            # 가격 컬럼
            ws.cell(row=r, column=col).number_format = "#,##0.00"

    # ── 열 너비 ──────────────────────────────────────────────────────
    #          A   B  C  D  E  F  G   H   I   J   K  L   M    N    O
    widths = [12, 10, 10, 10, 10, 7, 12, 12, 11, 11, 8, 10, 13, 12, 12]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A9"   # 헤더 고정

    # ── 거래내역 시트 (실제 run_backtest 결과 — 정적 값) ─────────────
    sigs = generate_signals(
        df,
        short_thresholds=[threshold] if side == "short" else [],
        long_thresholds=[threshold] if side == "long" else [],
        min_gap_days=min_gap_days,
    )
    bt = run_backtest(
        df, sigs, horizon_days,
        CostModel(commission_per_contract, slippage_ticks),
        ExitRules(),
        RollModel(roll_cost_pct=roll_cost_pct),
        spec=spec,
    )
    trades = bt.trades

    tr = wb.create_sheet("거래내역")
    tr["A1"] = (f"거래내역 — {side} {cur_sym}{threshold:g} × {horizon_days}일 "
                f"(거래 {len(trades)}건, 다운로드 시점 파라미터 기준)")
    tr["A1"].font = title_font
    tr["A2"] = "거래 수"
    tr["A2"].font = accent_font
    tr["B2"] = len(trades)
    tr["B2"].font = bold
    tr["B2"].border = border
    tr["C2"] = "← 라이브 실험은 '백테스트' 시트, 확정 내역은 이 시트 (앱 결과와 일치)"
    tr["C2"].font = italic

    tr_headers = ["신호일", "진입일", "청산일", "진입가", "청산가",
                  "롤횟수", "수익률", f"PnL({cur_sym})"]
    for j, h in enumerate(tr_headers):
        c = tr.cell(row=4, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    for i, t in enumerate(trades):
        r_out = 5 + i
        tr.cell(row=r_out, column=1, value=t.signal.date.to_pydatetime())
        tr.cell(row=r_out, column=2, value=t.entry_date.to_pydatetime())
        tr.cell(row=r_out, column=3, value=t.exit_date.to_pydatetime())
        tr.cell(row=r_out, column=4, value=round(t.entry_price, 4))
        tr.cell(row=r_out, column=5, value=round(t.exit_price, 4))
        tr.cell(row=r_out, column=6, value=t.num_rollovers)
        tr.cell(row=r_out, column=7, value=t.return_pct)
        tr.cell(row=r_out, column=8, value=round(t.net_pnl_usd, 0))

    fmt = {1: "yyyy-mm-dd", 2: "yyyy-mm-dd", 3: "yyyy-mm-dd",
           4: "#,##0.00", 5: "#,##0.00", 6: "0",
           7: "+0.00%;-0.00%", 8: "#,##0"}
    for i in range(len(trades)):
        r_out = 5 + i
        for col_i, nf in fmt.items():
            tr.cell(row=r_out, column=col_i).number_format = nf
    tr_widths = [12, 12, 12, 11, 11, 8, 10, 13]
    for j, w in enumerate(tr_widths):
        tr.column_dimensions[get_column_letter(1 + j)].width = w
    tr.freeze_panes = "A5"

    # ── 만기일 시트 (롤횟수 COUNTIFS 참조용) ─────────────────────────
    # 앱 run_backtest도 전 종목에 동일 WTI(CL) 만기 스케줄을 사용(연속물 단일
    # 시계열 단순화) — 엑셀 롤횟수가 앱과 일치한다.
    exp_ws = wb.create_sheet("만기일")
    exp_ws["A1"] = "선물 월물 만기일 (CME WTI 규칙 근사)"
    exp_ws["A1"].font = bold
    exp_ws["B1"] = "(인도월 전월 25일의 3영업일 전. 백테스트!K열이 COUNTIFS로 참조 — 앱과 동일 스케줄)"
    exp_ws["B1"].font = italic
    expiries = wti_expiry_dates(
        pd.Timestamp(df["date"].iloc[0]), pd.Timestamp(df["date"].iloc[-1])
    ) if n else []
    for i, e in enumerate(expiries):
        cell = exp_ws.cell(row=2 + i, column=1, value=pd.Timestamp(e).to_pydatetime())
        cell.number_format = "yyyy-mm-dd"
    exp_ws.column_dimensions["A"].width = 14
    exp_ws.column_dimensions["B"].width = 62

    # ── 로직 설명 시트 ───────────────────────────────────────────────
    doc = wb.create_sheet("로직설명")
    notes = [
        [f"{disp} 백테스트 — 엑셀 로직 설명", ""],
        ["", ""],
        ["입력칸", "백테스트 시트 B2~B7의 노란 칸을 바꾸면 전체 재계산"],
        ["방향(B2)", "short=고가가 임계값 위로 첫 터치 시 매도, long=저가가 아래로 첫 터치 시 매수"],
        ["임계값(B3)", f"신호 발생 가격선 ({cur_sym}{unit_sfx})"],
        ["보유기간(B4)", "진입 후 청산까지 영업일 수"],
        ["롤비용(B5)", "만기 롤오버 1회당 % (양수=콘탱고 비용/음수=backwardation 이익)"],
        ["슬리피지(B6)", "체결 1회당 N틱 불리하게 체결 (앱과 동일 — 1틱 = H3). long 진입가↑·청산가↓"],
        ["수수료(B7)", f"체결 1레그당 {cur_sym} (진입+청산 = 2레그). 앱 CostModel.commission_per_contract"],
        ["계약배수(H2)", f"1계약 = {spec.multiplier:g}{unit or '단위'} 명목 (가격 1포인트 = {cur_sym}{spec.multiplier:g}). 종목별 ContractSpec"],
        ["틱 크기(H3)", f"최소 호가단위 ({cur_sym}). 슬리피지 환산 계수"],
        ["신호(F열)", "장중 고가/저가 기준 전일 대비 첫 돌파 (히스테리시스 — 연속 돌파는 1회만)"],
        ["진입일(G열)", "신호 다음 영업일 날짜"],
        ["청산일(H열)", "진입일 + 보유기간 영업일 날짜"],
        ["진입가(I열)", "진입일 시가 (look-ahead 제거)"],
        ["청산가(J열)", "청산일 종가"],
        ["롤횟수(K열)", "진입일(G)~청산일(H) 사이 '만기일' 시트의 만기 수 COUNTIFS (진입<만기≤청산) — 앱과 동일"],
        ["유효진입가(N열)", "진입가에 슬리피지 반영 (=실제 체결가)"],
        ["유효청산가(O열)", "청산가에 슬리피지 반영"],
        ["PnL(M열)", f"[부호×(유효청산−유효진입)×계약배수 − 2×수수료] + 롤비용. 1계약 {cur_sym}. 앱 net_pnl과 일치"],
        ["수익률(L열)", "부호×(청산가/진입가−1) + 롤비용/진입대금. 앱 Trade.return_pct 정의(gross+롤)"],
        ["거래내역 시트", "다운로드 시점 파라미터의 실제 run_backtest 결과(정적 값) — 앱 백테스트와 일치"],
        ["만기일 시트", "롤횟수 계산 기준 만기일 (전 종목 WTI 스케줄 근사 — 앱과 동일)"],
        ["", ""],
        ["한계", "SL/TP 조기청산·MAE/MFE·walk-forward는 미반영 (보유기간 고정 전략). 임계·보유기간·비용·롤의 라이브 재계산 도구."],
    ]
    for i, (a, b) in enumerate(notes):
        doc.cell(row=1 + i, column=1, value=a).font = bold if i == 0 else Font()
        doc.cell(row=1 + i, column=2, value=b)
    doc.column_dimensions["A"].width = 16
    doc.column_dimensions["B"].width = 76

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_oil_trend_excel(
    df: pd.DataFrame,      # 전체 OHLCV (date ASC) — 라이브 수식이 참조할 raw 데이터
    events: list,          # list[TrendEvent] — 현재 디클러스터 결과(정적 스냅샷·교차검증용)
    raw_n: int,
    *,
    lookback: int,
    horizon: int,
    price_lo: float,
    price_hi: float,
    change_lo: float,
    change_hi: float,
    gap: int,
    name: str,
    price_sym: str,
) -> bytes:
    """추세 탐색기 '향후 종가 증감율'을 **라이브 수식 .xlsx** 로.

    시트1 '데이터+계산': raw OHLCV + Excel 수식(과거/향후 증감율·종가·증감율 밴드 매칭·
    gap 디클러스터·요약·회귀). 노란 칸(종가범위·증감율범위·과거L·향후H·간격)을 바꾸면
    전체 재계산 — 백테스트 엑셀과 동일한 라이브 방식. 시트2 '이벤트(현재)': 다운로드
    시점 결과의 정적 스냅샷(수식 결과 교차검증·앱 화면과 일치).
    """
    df = df.sort_values("date").reset_index(drop=True)
    n = len(df)

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    accent_font = Font(bold=True, color=_ACCENT)
    italic = Font(italic=True, color="6F6A62")
    input_fill = PatternFill("solid", fgColor=_INPUT_BG)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _thin_border()
    center = Alignment(horizontal="center")

    HROW = 10               # 헤더 행
    D0 = HROW + 1           # 첫 데이터 행
    last = HROW + n         # 마지막 데이터 행 (n=0이면 HROW)
    psym = price_sym or ""

    def _cv(v: float) -> float:
        # ±무한대 경계(±1e9)는 표시용 유한 sentinel(±100000; %수익률엔 사실상 무한대).
        return -100000.0 if v <= -1e8 else (100000.0 if v >= 1e8 else float(v))

    wb = Workbook()
    ws = wb.active
    ws.title = "데이터+계산"
    disp = name if "선물" in name else f"{name} 선물"
    ws["A1"] = f"{disp} — 향후 종가 증감율 (라이브 수식)"
    ws["A1"].font = title_font

    # ── 입력칸(노란) B2~B8 — 바꾸면 전체 재계산 ───────────────────────
    inputs = [
        (f"종가 하한 ({psym})", _cv(price_lo)),
        (f"종가 상한 ({psym})", _cv(price_hi)),
        ("과거 증감율 하한 (%)", _cv(change_lo)),
        ("과거 증감율 상한 (%)", _cv(change_hi)),
        ("과거 기간 L (영업일)", int(lookback)),
        ("향후 기간 H (영업일)", int(horizon)),
        ("이벤트 최소 간격 (영업일, 0=원시)", int(gap)),
    ]
    for i, (label, val) in enumerate(inputs):
        r = 2 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = bold
        c = ws[f"B{r}"]
        c.value = val
        c.fill = input_fill
        c.border = border
        c.alignment = center
    ws["C2"] = "← 노란 칸을 바꾸면 전체가 재계산됩니다"
    ws["C2"].font = italic

    # ── 요약(전부 수식) D2~E7 ─────────────────────────────────────────
    summary = [
        ("독립 표본 n", f"=COUNT(L{D0}:L{last})", "0"),
        ("원시 매칭 수", f"=SUM(H{D0}:H{last})", "0"),
        ("평균 향후 증감율 (%)", f"=IFERROR(AVERAGE(L{D0}:L{last}),0)", "+0.00;-0.00"),
        ("상승 비율 (%)", f'=IFERROR(COUNTIF(L{D0}:L{last},">0")/COUNT(L{D0}:L{last})*100,0)', "0.0"),
        ("회귀 β (향후~과거)", f'=IFERROR(SLOPE(L{D0}:L{last},K{D0}:K{last}),"")', "+0.000;-0.000"),
        ("회귀 R²", f'=IFERROR(RSQ(L{D0}:L{last},K{D0}:K{last}),"")', "0.000"),
    ]
    for i, (label, formula, fmt) in enumerate(summary):
        r = 2 + i
        ws[f"D{r}"] = label
        ws[f"D{r}"].font = accent_font
        cell = ws[f"E{r}"]
        cell.value = formula
        cell.font = bold
        cell.number_format = fmt
        cell.border = border

    # ── 헤더(HROW) + raw 데이터 + 수식 열 ─────────────────────────────
    # A날짜 B시 C고 D저 E종 | F과거증감율 G향후증감율 H매칭 I채택(디클러스터)
    # J최근채택행(헬퍼) | K채택과거 L채택향후
    headers = ["날짜", "시가", "고가", "저가", "종가",
               "과거증감율(%)", "향후증감율(%)", "매칭", "채택", "_최근채택행",
               "채택 과거(%)", "채택 향후(%)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=HROW, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    for idx in range(n):
        r = D0 + idx
        row = df.iloc[idx]
        ws.cell(row=r, column=1, value=pd.Timestamp(row["date"]).to_pydatetime())
        ws.cell(row=r, column=2, value=float(row["open"]))
        ws.cell(row=r, column=3, value=float(row["high"]))
        ws.cell(row=r, column=4, value=float(row["low"]))
        ws.cell(row=r, column=5, value=float(row["close"]))
        # F 과거증감율% = (종가 ÷ 종가[행−L] − 1)×100. 행−L<첫데이터면 공란.
        ws.cell(row=r, column=6, value=(
            f'=IF(ROW()-$B$6<{D0},"",IFERROR((E{r}/INDEX($E:$E,ROW()-$B$6)-1)*100,""))'
        ))
        # G 향후증감율% = (종가[행+H] ÷ 종가 − 1)×100. 행+H>끝이면 공란.
        ws.cell(row=r, column=7, value=(
            f'=IF(ROW()+$B$7>{last},"",IFERROR((INDEX($E:$E,ROW()+$B$7)/E{r}-1)*100,""))'
        ))
        # H 매칭 = 종가∈[B2,B3] ∧ 과거증감율∈[B4,B5] (양 증감율 존재 시).
        ws.cell(row=r, column=8, value=(
            f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(G{r}),E{r}>=$B$2,E{r}<=$B$3,'
            f'F{r}>=$B$4,F{r}<=$B$5),1,0)'
        ))
        # I 채택 / J 최근채택행 — gap 영업일 이내 연속 매칭은 1건만(그리디 디클러스터).
        if idx == 0:
            ws.cell(row=r, column=9, value=f'=IF(H{r}=1,1,0)')
            ws.cell(row=r, column=10, value=f'=IF(I{r}=1,ROW(),-1000000)')
        else:
            ws.cell(row=r, column=9, value=f'=IF(AND(H{r}=1,ROW()-J{r-1}>=$B$8),1,0)')
            ws.cell(row=r, column=10, value=f'=IF(I{r}=1,ROW(),J{r-1})')
        # K 채택 과거% / L 채택 향후% (회귀 x·y, 요약 입력).
        ws.cell(row=r, column=11, value=f'=IF(I{r}=1,F{r},"")')
        ws.cell(row=r, column=12, value=f'=IF(I{r}=1,G{r},"")')

    for idx in range(n):
        r = D0 + idx
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        for col in (2, 3, 4, 5):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        for col in (6, 7, 11, 12):
            ws.cell(row=r, column=col).number_format = "+0.00;-0.00"

    widths = [12, 10, 10, 10, 10, 14, 14, 8, 8, 12, 13, 13]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.column_dimensions["J"].hidden = True       # 디클러스터 헬퍼 열 숨김
    if n:
        ws.freeze_panes = f"A{D0}"

    # ── 시트2: 현재 결과 정적 스냅샷 (라이브 수식 교차검증·앱 화면과 일치) ──
    snap = wb.create_sheet("이벤트(현재)")
    nn = len(events)
    mean = sum(e.forward_return * 100.0 for e in events) / nn if nn else 0.0
    upr = 100.0 * sum(1 for e in events if e.forward_return > 0) / nn if nn else 0.0
    snap["A1"] = "현재 결과 (정적 스냅샷 — 왼쪽 시트 라이브 수식과 대조용)"
    snap["A1"].font = title_font
    srows: list[tuple[str, object]] = [
        ("종가 범위", f"{psym}{price_lo:g} ~ {psym}{price_hi:g}"),
        ("과거 L / 향후 H", f"{lookback} / {horizon} 영업일"),
        ("이벤트 최소 간격", f"{gap} 영업일"),
        ("독립 표본 n / 원시", f"{nn} / {raw_n}"),
        (f"평균 향후 {horizon}일 증감율 (%)", round(mean, 2)),
        ("상승 비율 (%)", round(upr, 1)),
    ]
    for i, (k, v) in enumerate(srows, start=3):
        snap[f"A{i}"] = k
        snap[f"A{i}"].font = bold
        snap[f"B{i}"] = v
    hr = 10
    hdr2 = ["신호일", "종가", f"과거 {lookback}일 증감율(%)", f"향후 {horizon}일 증감율(%)"]
    for j, h in enumerate(hdr2):
        c = snap.cell(row=hr, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
    for i, e in enumerate(events):
        rr = hr + 1 + i
        snap.cell(row=rr, column=1, value=str(e.date.date()))
        snap.cell(row=rr, column=2, value=round(float(e.close), 2))
        snap.cell(row=rr, column=3, value=round(e.past_return * 100.0, 2))
        snap.cell(row=rr, column=4, value=round(e.forward_return * 100.0, 2))
    for col, w in zip("ABCD", (12, 12, 22, 24)):
        snap.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
