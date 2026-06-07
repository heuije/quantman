"""전략연구소 백테스트 결과 → 엑셀(.xlsx) 생성 (회귀분석 포함).

선물 분석 엑셀과 동일 철학:
- 요약 탭(값) + 일별성과 raw data 탭(수익률은 함수) + 회귀분석 탭(라이브 함수)
  + 거래내역 탭(값).
- 회귀분석: 전략 일별수익률(Y) vs 벤치마크 일별수익률(X)의 CAPM 회귀.
  베타(SLOPE)·알파(INTERCEPT)·R²(RSQ)·상관(CORREL)·정보비율을 엑셀 함수로
  계산 → 사용자가 데이터를 바꾸면 회귀 결과도 재계산.

손상 방지: 거래내역·요약·일별성과 본문은 정적 값. 일별수익률과 회귀 계수만
일반 함수(동적배열 아님). 선물 엑셀에서 검증된 안전 패턴.

입력: run_strategy_backtest 결과 dict
  { success, equity(pd.Series), benchmark(pd.Series), trades(DataFrame), metrics(dict) }
"""
from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side as XlSide
from openpyxl.utils import get_column_letter

_ACCENT = "D97757"
_HEAD_BG = "F7ECE5"
_BORDER = "E8E3DB"
TRADING_DAYS = 252


def _border() -> Border:
    s = XlSide(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def build_strategy_excel(
    result: dict,
    *,
    strategy_name: str = "전략",
    symbol: str = "",
    initial_capital: float = 10_000_000.0,
    start: Optional[str] = None,
    end: Optional[str] = None,
) -> bytes:
    """백테스트 결과 dict → 엑셀 바이트."""
    if not result.get("success"):
        raise ValueError(result.get("error") or "백테스트 실패")

    equity: pd.Series = result["equity"]
    benchmark: pd.Series = result["benchmark"]
    trades: pd.DataFrame = result["trades"]
    metrics: dict = result.get("metrics", {})

    # 벤치마크를 전략 equity 인덱스에 정렬
    eq = equity.dropna()
    bm = benchmark.reindex(eq.index).ffill() if benchmark is not None else None

    wb = Workbook()
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="20201D")
    accent_font = Font(bold=True, color=_ACCENT)
    head_fill = PatternFill("solid", fgColor=_HEAD_BG)
    border = _border()
    center = Alignment(horizontal="center")

    # ── 일별성과 시트 (raw data; 수익률은 함수) ──────────────────────
    ws = wb.active
    ws.title = "일별성과"
    ws["A1"] = f"{strategy_name} — 일별 성과 (raw data)"
    ws["A1"].font = title_font
    headers = ["날짜", "전략 등자산", "벤치마크 등자산",
               "전략 일별수익률", "벤치마크 일별수익률", "초과수익률(전략-벤치)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=1 + j, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = border
        c.alignment = center

    dates = list(eq.index)
    nd = len(dates)
    for i in range(nd):
        r = 3 + i
        d = pd.Timestamp(dates[i])
        ws.cell(row=r, column=1, value=d.to_pydatetime()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=float(eq.iloc[i])).number_format = "#,##0"
        if bm is not None and not pd.isna(bm.iloc[i]):
            ws.cell(row=r, column=3, value=float(bm.iloc[i])).number_format = "#,##0"
        if i == 0:
            continue
        # D 전략 일별수익률 = 오늘/어제 - 1 (함수)
        ws.cell(row=r, column=4,
                value=f"=IFERROR(B{r}/B{r-1}-1,\"\")").number_format = "+0.00%;-0.00%"
        # E 벤치마크 일별수익률
        ws.cell(row=r, column=5,
                value=f"=IFERROR(C{r}/C{r-1}-1,\"\")").number_format = "+0.00%;-0.00%"
        # F 초과수익률
        ws.cell(row=r, column=6,
                value=f"=IFERROR(D{r}-E{r},\"\")").number_format = "+0.00%;-0.00%"

    last = 2 + nd          # 마지막 데이터 행
    rng_d = f"D4:D{last}"  # 수익률은 4행부터 유효 (3행은 첫날, 수익률 없음)
    rng_e = f"E4:E{last}"
    rng_f = f"F4:F{last}"
    for col, w in zip("ABCDEF", [12, 14, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    # ── 회귀분석 시트 (라이브 함수: CAPM) ────────────────────────────
    rg = wb.create_sheet("회귀분석", 0)
    rg["A1"] = "회귀분석 (CAPM) — 전략 수익률 vs 벤치마크 수익률"
    rg["A1"].font = title_font
    rg["A2"] = "Y(종속) = 전략 일별수익률, X(독립) = 벤치마크 일별수익률"
    rg["A2"].font = Font(italic=True, color="6F6A62")

    rows = [
        ("베타 (β, 시장 민감도)", f"=SLOPE(일별성과!{rng_d},일별성과!{rng_e})", "0.000",
         "β>1 시장보다 민감, β<1 둔감, β<0 역방향"),
        ("알파 (α, 일간 초과수익)", f"=INTERCEPT(일별성과!{rng_d},일별성과!{rng_e})", "+0.0000%;-0.0000%",
         "시장으로 설명 안 되는 일간 초과수익 (>0 양호)"),
        ("알파 (연환산)", "=(1+E4)^252-1", "+0.00%;-0.00%",
         "일간 알파를 252영업일 복리 환산"),
        ("R² (설명력)", f"=RSQ(일별성과!{rng_d},일별성과!{rng_e})", "0.000",
         "전략 변동이 시장으로 설명되는 비율 (0~1)"),
        ("상관계수 (ρ)", f"=CORREL(일별성과!{rng_d},일별성과!{rng_e})", "0.000",
         "전략·시장 일별수익률 선형 상관"),
        ("전략 평균 일수익", f"=AVERAGE(일별성과!{rng_d})", "+0.0000%;-0.0000%", ""),
        ("벤치마크 평균 일수익", f"=AVERAGE(일별성과!{rng_e})", "+0.0000%;-0.0000%", ""),
        ("추적오차 (연환산)", f"=STDEV(일별성과!{rng_f})*SQRT(252)", "0.00%",
         "전략-벤치 초과수익률의 변동성"),
        ("정보비율 (IR)", f"=IFERROR(AVERAGE(일별성과!{rng_f})*252/(STDEV(일별성과!{rng_f})*SQRT(252)),\"\")", "0.00",
         "초과수익 / 추적오차 (높을수록 꾸준한 초과성과)"),
    ]
    rg.cell(row=3, column=1, value="지표").font = bold
    rg.cell(row=3, column=2, value="값").font = bold
    rg.cell(row=3, column=3, value="설명").font = bold
    for c in range(1, 4):
        rg.cell(row=3, column=c).fill = head_fill
        rg.cell(row=3, column=c).border = border
    for i, (label, formula, fmt, desc) in enumerate(rows):
        r = 4 + i
        rg.cell(row=r, column=1, value=label).font = accent_font
        cell = rg.cell(row=r, column=2, value=formula)
        cell.font = bold
        cell.number_format = fmt
        cell.border = border
        rg.cell(row=r, column=3, value=desc)
    rg.column_dimensions["A"].width = 22
    rg.column_dimensions["B"].width = 14
    rg.column_dimensions["C"].width = 48
    rg["A15"] = ("※ 알파·베타·R²·상관·정보비율은 '일별성과' 시트를 참조하는 라이브 함수입니다 "
                 "(데이터 바뀌면 재계산).")
    rg["A15"].font = Font(italic=True, color="6F6A62")

    # ── 요약 시트 (값) ───────────────────────────────────────────────
    sm = wb.create_sheet("요약", 0)
    sm["A1"] = f"{strategy_name} — 백테스트 요약"
    sm["A1"].font = title_font
    info = [
        ("종목", symbol or "-"),
        ("초기자본", f"{initial_capital:,.0f}원"),
        ("기간", f"{start or str(dates[0].date()) if nd else '-'} ~ {end or (str(dates[-1].date()) if nd else '-')}"),
        ("거래일수", nd),
    ]
    r = 3
    for label, val in info:
        sm.cell(row=r, column=1, value=label).font = accent_font
        sm.cell(row=r, column=2, value=val).font = bold
        r += 1

    r += 1
    sm.cell(row=r, column=1, value="핵심 성과").font = bold
    r += 1
    pct = lambda v: (v / 100.0) if v is not None else None
    stat_rows = [
        ("총수익률", pct(metrics.get("total_return")), "0.00%"),
        ("CAGR (연복리)", pct(metrics.get("cagr")), "0.00%"),
        ("MDD (최대낙폭)", pct(metrics.get("mdd")), "0.00%"),
        ("Sharpe", metrics.get("sharpe"), "0.00"),
        ("벤치 대비 초과수익", pct(metrics.get("excess_return")), "0.00%"),
        ("거래 수", metrics.get("n_trades"), "0"),
        ("승률", pct(metrics.get("win_rate")), "0.0%"),
        ("평균 보유일", metrics.get("avg_hold"), "0.0"),
        ("평균 거래수익률", pct(metrics.get("avg_trade_return")), "+0.00%;-0.00%"),
    ]
    for label, val, fmt in stat_rows:
        sm.cell(row=r, column=1, value=label).font = accent_font
        cell = sm.cell(row=r, column=2, value=(float(val) if val is not None else None))
        cell.font = bold
        cell.number_format = fmt
        cell.border = border
        r += 1

    r += 1
    sm.cell(row=r, column=1, value="회귀분석 (회귀분석 탭 참조)").font = bold
    r += 1
    for label, ref in [("베타 (β)", "회귀분석!B4"), ("알파(연환산)", "회귀분석!B6"),
                       ("R²", "회귀분석!B7"), ("정보비율", "회귀분석!B12")]:
        sm.cell(row=r, column=1, value=label).font = accent_font
        sm.cell(row=r, column=2, value=f"={ref}").font = bold
        r += 1
    sm.column_dimensions["A"].width = 22
    sm.column_dimensions["B"].width = 16

    # ── 거래내역 시트 (값) ───────────────────────────────────────────
    tr = wb.create_sheet("거래내역")
    tr["A1"] = f"거래내역 ({len(trades)}건)"
    tr["A1"].font = title_font
    if len(trades) > 0:
        cols = list(trades.columns)
        for j, h in enumerate(cols):
            c = tr.cell(row=3, column=1 + j, value=str(h))
            c.font = bold
            c.fill = head_fill
            c.border = border
            c.alignment = center
        for i in range(len(trades)):
            for j, colname in enumerate(cols):
                v = trades.iloc[i][colname]
                cell = tr.cell(row=4 + i, column=1 + j)
                if isinstance(v, (pd.Timestamp,)) or hasattr(v, "to_pydatetime"):
                    cell.value = pd.Timestamp(v).to_pydatetime()
                    cell.number_format = "yyyy-mm-dd"
                elif "수익률" in str(colname):
                    cell.value = float(v) / 100.0
                    cell.number_format = "+0.00%;-0.00%"
                elif "가" in str(colname):
                    cell.value = round(float(v), 2)
                    cell.number_format = "#,##0.00"
                else:
                    try:
                        cell.value = float(v) if str(v).replace(".", "").replace("-", "").isdigit() else str(v)
                    except (ValueError, TypeError):
                        cell.value = str(v)
        for j in range(len(cols)):
            tr.column_dimensions[get_column_letter(1 + j)].width = 13
    tr.freeze_panes = "A4"

    # 시트 순서: 요약 · 회귀분석 · 일별성과 · 거래내역 (삽입 순서대로 정렬됨)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
