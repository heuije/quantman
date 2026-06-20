"""분석 코퍼스 회귀 — 챗봇 정본 경로(strategy_from_spec)로 전 케이스 실행 고정.

analysis_corpus.CASES(13 base + 엣지)를 LLM 없이 IR→엔진→엑셀로 돌려 형상·시트·라이브수식·
데이터품질 경고를 회귀로 잠근다. (엑셀 *구조*만 보는 test_ir_excel_export_shapes는 run_query
경로, 여기는 strategy_from_spec=챗봇 정본 경로 — 검증·경고·전달까지 포함.)
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))   # analysis_corpus 임포트
import analysis_corpus as C


def _text(ws) -> str:
    return "\n".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None)


@pytest.mark.parametrize("case", C.CASES, ids=[c["name"] for c in C.CASES])
def test_corpus_case_runs_and_excel_structure(case) -> None:
    """전 케이스: strategy_from_spec 성공 + 엑셀 생성 + 기대 시트·문자열(헤더·값·수식) 포함."""
    _ir, _ds, res, xlsx, _mode = C.run_case(case)
    assert res.get("success"), f"[{case['name']}] 엔진 실패: {res.get('error')}"
    assert xlsx and len(xlsx) > 3000, f"[{case['name']}] xlsx 미생성"
    wb = load_workbook(io.BytesIO(xlsx))
    for sheet, needles in case["checks"].items():
        assert sheet in wb.sheetnames, f"[{case['name']}] 시트 '{sheet}' 없음 — {wb.sheetnames}"
        text = _text(wb[sheet])
        for n in needles:
            assert n in text, f"[{case['name']}] '{sheet}'에 '{n}' 없음"


def test_cross_asset_signal_panel_is_live_nested_if() -> None:
    """크로스에셋: 신호·결정 패널의 포지션 수식이 중첩 IF(엔진 신호 정의와 1:1·이번 세션 핵심)."""
    case = next(c for c in C.CASES if c["name"] == "cross_asset_intraday")
    _ir, _ds, _res, xlsx, _m = C.run_case(case)
    ps = load_workbook(io.BytesIO(xlsx))["신호·결정"]
    posf = next((ps.cell(r, ps.max_column).value for r in range(7, 60)
                 if isinstance(ps.cell(r, ps.max_column).value, str)
                 and ps.cell(r, ps.max_column).value.startswith("=IF(")), None)
    assert posf and "<=-0.1" in posf, f"포지션 라이브 수식 없음: {posf}"


def test_data_quality_warning_surfaces_in_result_and_excel() -> None:
    """stale 신호종목 → assess_data_quality stale_data 경고가 결과 + 엑셀(지표·설명)에 표면화."""
    case = next(c for c in C.CASES if c["name"] == "data_quality_stale")
    _ir, _ds, res, xlsx, _m = C.run_case(case)
    codes = [w.get("code") for w in (res.get("warnings") or []) if isinstance(w, dict)]
    assert "stale_data" in codes, f"stale_data 경고 없음: {codes}"
    assert "⚠" in _text(load_workbook(io.BytesIO(xlsx))["지표·설명"])


def test_corpus_covers_all_result_shapes() -> None:
    """코퍼스가 13 result_shape를 빠짐없이 덮는지(다양성 보장) — 회귀로 고정."""
    shapes = set()
    for case in C.CASES:
        _ir, _ds, res, _x, _m = C.run_case(case)
        if res.get("success"):
            shapes.add(C.case_shape(res))
    expected = {"simulate", "sweep", "extremize", "select", "describe_single",
                "describe_portfolio", "signal_dist", "relate_ic", "relate_regression", "event_study"}
    missing = expected - shapes
    assert not missing, f"코퍼스가 안 덮는 형상: {missing}"
