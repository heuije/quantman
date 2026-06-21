"""신호·결정 패널 라이브수식 **값 검증** (F2/F3) — Excel 셀이 엔진 지표값과 일치하는지.

formula evaluator(formulas/pycel) 미설치 환경이므로, 패널이 emit하는 좁은 수식 grammar
(AVERAGE/STDEV.S/MAX/MIN/SUM/AND/OR/IF·산술·비교·원자료·패널 셀 참조)만 다루는 미니 평가기로
생성된 셀을 실제 데이터에 평가→엔진 indicators.py 산출값과 대조한다. '틀린 라이브수식은 없느니만
못함'(원칙1)을 회귀로 잠근다. 미지원 지표/연산은 패널 생략 + 구체 사유(F4) 표면화 검증.
"""
from __future__ import annotations

import io
import math
import re
import sys
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from analysis_corpus import _noisy, _ohlc  # noqa: E402

from quant_core import indicators  # noqa: E402
from quant_core.ir_engine import StrategyIR, build_strategy_excel, strategy_from_spec  # noqa: E402


# ── 미니 Excel 평가기 (이 패널이 emit하는 grammar 전용; 미지토큰은 raise → 거짓통과 불가) ──
def _make_eval(wb):
    raw, panel = wb["원자료"], wb["신호·결정"]
    memo: dict = {}

    def cell(sheet: str, col: str, row: int):
        key = (sheet, col, row)
        if key in memo:
            return memo[key]
        v = (raw if sheet == "원자료" else panel)[f"{col}{row}"].value
        if isinstance(v, bool):
            out = v
        elif isinstance(v, (int, float)):
            out = float(v)
        elif v is None:
            out = float("nan")
        elif isinstance(v, str) and v.startswith("="):
            out = ev(v)
        else:
            raise ValueError(f"예상 밖 셀 {sheet}!{col}{row}={v!r}")
        memo[key] = out
        return out

    def rng(sheet: str, col: str, r1: int, r2: int):
        return [cell(sheet, col, r) for r in range(int(r1), int(r2) + 1)]

    def _flat(args):
        out = []
        for a in args:
            out.extend(a if isinstance(a, list) else [a])
        return out

    ns = {
        "_C": cell, "_R": rng,
        "AVERAGE": lambda *a: float(np.mean(_flat(a))),
        "SUM": lambda *a: float(np.sum(_flat(a))),
        "MAX": lambda *a: float(np.max(_flat(a))),
        "MIN": lambda *a: float(np.min(_flat(a))),
        "STDEV_S": lambda *a: float(np.std(_flat(a), ddof=1)),
        "AND": lambda *a: all(a), "OR": lambda *a: any(a),
        "IF": lambda c, a, b: a if c else b,
        "LN": math.log, "SQRT": math.sqrt, "ABS": abs,
    }

    def ev(formula: str):
        s = formula[1:] if formula.startswith("=") else formula
        s = s.replace("STDEV.S", "STDEV_S")
        s = re.sub(r"(원자료!)?([A-Z]+)(\d+):(원자료!)?([A-Z]+)(\d+)",
                   lambda m: f'_R("{"원자료" if m.group(1) else "P"}","{m.group(2)}",{m.group(3)},{m.group(6)})', s)
        s = re.sub(r"(원자료!)?([A-Z]+)(\d+)",
                   lambda m: f'_C("{"원자료" if m.group(1) else "P"}","{m.group(2)}",{m.group(3)})', s)
        s = s.replace("<>", "!=")
        s = re.sub(r"(?<![<>=!])=(?!=)", "==", s)
        return eval(s, {"__builtins__": {}}, ns)  # noqa: S307 — 좁은 ns·미지토큰 raise

    return ev


def _single_ds(n=340, seed=7):
    return {"AAA": indicators.compute_all(_ohlc(_noisy(n, 0.013, seed=seed, drift=0.0004)))}


def _run(signal, ds):
    ir = {"name": "panel-test", "universe": {"kind": "single", "symbols": ["AAA"]}, "signal": signal,
          "position": {"direction": "long", "sizing": {"mode": "pct_cash", "amount_pct": 20},
                       "entry": {"mode": "on_signal"}, "exit": {"hold_days": 5}},
          "simulation": {"delay": 1, "fill": "next_open"}, "query": "simulate", "study": {"axis": "none"}}
    res = strategy_from_spec(ir, ds)
    assert res.get("success"), res.get("error")
    wb = load_workbook(io.BytesIO(build_strategy_excel(StrategyIR.model_validate(ir), ds, res)))
    return wb, res


def _cmp(ref, op, val):
    return {"op": "compare", "params": {"op": op},
            "inputs": {"left": {"op": "data", "params": {"ref": ref}},
                       "right": {"op": "const", "params": {"value": val}}}}


_C_CLOSE = {"op": "data", "params": {"ref": "__SELF__.Close"}}


def _spot_rows(wb, count=4):
    """패널 포지션열에 수식이 있는 행 중 끝에서 count개(룩백 충분 구간)."""
    panel = wb["신호·결정"]
    pos_col = panel.max_column
    rows = [r for r in range(6, panel.max_row + 1)
            if isinstance(panel.cell(r, pos_col).value, str)
            and panel.cell(r, pos_col).value.startswith("=")]
    return rows[-count:]


# ── F2: 윈도우 지표 셀 == 엔진 indicators.py 산출값 ──────────────────────────────
@pytest.mark.parametrize("ind", ["ma_dev_20d", "ma_dev_60d", "bb_pct", "bb_width", "ma_gap_20_60"])
def test_window_indicator_cell_matches_engine(ind):
    ds = _single_ds()
    wb, res = _run(_cmp(f"__SELF__.{ind}", ">", -999.0), ds)   # 항상참 → 거래·패널 생성
    ev = _make_eval(wb)
    panel = wb["신호·결정"]
    eng = ds["AAA"][ind]
    idx = list(res["equity"].index)
    checked = 0
    for r in _spot_rows(wb):
        got = ev(panel.cell(r, 2).value)              # 신호값 열(B) = 지표 셀
        want = float(eng.loc[idx[r - 6]])
        assert math.isclose(got, want, rel_tol=1e-9, abs_tol=1e-9), f"{ind} r{r}: 셀 {got} != 엔진 {want}"
        checked += 1
    assert checked >= 2


# ── F3: ts_mean(MA) 위 비교 == 엔진 정의 ────────────────────────────────────────
def test_ts_mean_ma_cross_matches_engine():
    ds = _single_ds()
    sig = {"op": "compare", "params": {"op": ">"},
           "inputs": {"left": _C_CLOSE,
                      "right": {"op": "ts_mean", "params": {"window": 20}, "inputs": {"signal": _C_CLOSE}}}}
    wb, res = _run(sig, ds)
    ev = _make_eval(wb)
    panel = wb["신호·결정"]
    close = ds["AAA"]["Close"]
    ma20 = close.rolling(20).mean()                    # 엔진 ts_mean 정의(expression_parser)
    idx = list(res["equity"].index)
    for r in _spot_rows(wb):
        got = ev(panel.cell(r, panel.max_column).value)   # 포지션(룰) = Close>MA20 (bool)
        d = idx[r - 6]
        assert bool(got) == bool(close.loc[d] > ma20.loc[d]), f"r{r}: 셀 {got} != 엔진 {close.loc[d]>ma20.loc[d]}"


# ── F3: 골든크로스(cross) == 엔진 정의(오늘 교차 + 어제 반대) ────────────────────
def test_cross_golden_matches_engine():
    ds = _single_ds()

    def ma(w):
        return {"op": "ts_mean", "params": {"window": w}, "inputs": {"signal": _C_CLOSE}}

    sig = {"op": "cross", "params": {"direction": "up"}, "inputs": {"left": ma(5), "right": ma(20)}}
    wb, res = _run(sig, ds)
    ev = _make_eval(wb)
    panel = wb["신호·결정"]
    close = ds["AAA"]["Close"]
    ma5, ma20 = close.rolling(5).mean(), close.rolling(20).mean()
    idx = list(res["equity"].index)
    for r in _spot_rows(wb):
        got = ev(panel.cell(r, panel.max_column).value)
        d, dp = idx[r - 6], idx[r - 7]
        want = (ma5.loc[d] > ma20.loc[d]) and (ma5.loc[dp] < ma20.loc[dp])
        assert bool(got) == bool(want), f"r{r}: 셀 {got} != 엔진 {want}"


# ── F3: logic AND == 엔진(요소별 & ) ────────────────────────────────────────────
def test_logic_and_matches_engine():
    ds = _single_ds()
    sig = {"op": "logic", "params": {"logic": "AND"},
           "inputs": {"0": _cmp("__SELF__.pct_change_1d", "<=", 0.5),
                      "1": _cmp("__SELF__.pct_change_5d", ">", -5.0)}}
    wb, res = _run(sig, ds)
    ev = _make_eval(wb)
    panel = wb["신호·결정"]
    pc1, pc5 = ds["AAA"]["pct_change_1d"], ds["AAA"]["pct_change_5d"]
    idx = list(res["equity"].index)
    for r in _spot_rows(wb):
        got = ev(panel.cell(r, panel.max_column).value)
        d = idx[r - 6]
        assert bool(got) == bool((pc1.loc[d] <= 0.5) and (pc5.loc[d] > -5.0)), f"r{r}"


# ── F4: 미지원 지표/연산 → 패널 생략 + 구체 사유 표면화 ─────────────────────────
def test_unsupported_indicator_omits_panel_with_specific_reason():
    ds = _single_ds()
    res = strategy_from_spec(
        {"name": "rsi", "universe": {"kind": "single", "symbols": ["AAA"]},
         "signal": _cmp("__SELF__.rsi_14", "<", 30.0),
         "position": {"direction": "long", "sizing": {"mode": "pct_cash", "amount_pct": 20},
                      "entry": {"mode": "on_signal"}, "exit": {"hold_days": 5}},
         "simulation": {"fill": "next_open"}, "query": "simulate", "study": {"axis": "none"}}, ds)
    assert res.get("success")
    wb = load_workbook(io.BytesIO(build_strategy_excel(
        StrategyIR.model_validate({"name": "rsi", "universe": {"kind": "single", "symbols": ["AAA"]},
                                   "signal": _cmp("__SELF__.rsi_14", "<", 30.0),
                                   "position": {"direction": "long", "sizing": {"mode": "pct_cash", "amount_pct": 20},
                                                "entry": {"mode": "on_signal"}, "exit": {"hold_days": 5}},
                                   "simulation": {"fill": "next_open"}, "query": "simulate", "study": {"axis": "none"}}),
        ds, res)))
    assert "신호·결정" not in wb.sheetnames           # 미지원 → 패널 생략
    txt = "\n".join(str(v) for row in wb["지표·설명"].iter_rows(values_only=True) for v in row if v)
    assert "패널 생략" in txt and ("미지원 지표" in txt or "rsi" in txt.lower()), txt[-400:]
