"""증빙 엑셀 — 모든 분석 형상 MECE 검증 (엑셀 *구조* · run_query 경로).

형상별 IR·합성데이터는 analysis_corpus(단일 진실원천)에서 가져온다. 여기선 run_query로 실행해
.xlsx 시트 구성·헤더·값·수식 *문자열*을 직접 검수(구조 회귀). 챗봇 정본 경로
(strategy_from_spec)+데이터품질 경고·신호패널은 test_analysis_corpus가 담당한다.

라이브 수식의 수치 평가는 openpyxl이 못 하므로, 수식 *문자열*과 엔진 값의 정합(예: 포트 HHI
수식 옆 엔진값)으로 검증한다.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))          # analysis_corpus
from analysis_corpus import BASE_CASES, EVENT, SIG_C, ds_factor, ds_pf   # 단일 진실원천(중복 제거)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))   # quant_core
from quant_core.ir_engine import StrategyIR, build_strategy_excel, run_query


def _build(ds_fn, ir_dict):
    ds = ds_fn() if callable(ds_fn) else ds_fn
    ir = StrategyIR.model_validate(ir_dict)
    res = run_query(ir, ds)
    assert res.get("success"), f"엔진 실행 실패: {res.get('error')}"
    return ir, ds, res, build_strategy_excel(ir, ds, res)


def _sheet_text(ws) -> str:
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                out.append(str(v))
    return "\n".join(out)


@pytest.mark.parametrize("case", BASE_CASES, ids=[c["name"] for c in BASE_CASES])
def test_shape_excel_structure(case) -> None:
    """형상별: .xlsx 로드 + 기대 시트 존재 + 시트별 필수 문자열(헤더·값·수식) 포함."""
    _ir, _ds, _res, xlsx = _build(case["ds"], case["ir"])
    assert isinstance(xlsx, (bytes, bytearray)) and len(xlsx) > 4000
    wb = load_workbook(io.BytesIO(xlsx))
    for sheet, needles in case["checks"].items():
        assert sheet in wb.sheetnames, f"[{case['name']}] 시트 '{sheet}' 없음 — {wb.sheetnames}"
        text = _sheet_text(wb[sheet])
        for n in needles:
            assert n in text, f"[{case['name']}] 시트 '{sheet}'에 '{n}' 없음"


def test_portfolio_hhi_formula_matches_engine() -> None:
    """포트 HHI 라이브 수식(=SUMPRODUCT)이 엔진 HHI 값과 일치(증빙 수식 정확성)."""
    _ir, _ds, res, xlsx = _build(ds_pf, {"universe": {"kind": "portfolio",
        "symbols": ["AAA", "BBB", "CCC"], "weights": {"AAA": 0.5, "BBB": 0.3, "CCC": 0.2}},
        "signal": SIG_C, "query": "describe"})
    ws = load_workbook(io.BytesIO(xlsx))["보유진단"]
    # B9=수식, C9=엔진값. 가중치 0.5/0.3/0.2 → HHI = 0.25+0.09+0.04 = 0.38.
    assert ws["B9"].value == "=SUMPRODUCT(B4:B6,B4:B6)"
    assert ws["C9"].value == pytest.approx(res["concentration"]["hhi"])
    assert ws["C9"].value == pytest.approx(0.38, abs=1e-9)


def test_all_shapes_have_methodology_sheet() -> None:
    """모든 P2 형상에 전략정의(IR)+방법론을 담은 '설명' 시트가 붙는다(증빙 꼬리표)."""
    for case in BASE_CASES:
        if case["name"] in ("simulate", "period_split"):
            continue   # 단일 자산곡선 형상은 '지표·설명'(simulate 시트 재사용)으로 동일 역할
        _ir, _ds, _res, xlsx = _build(case["ds"], case["ir"])
        wb = load_workbook(io.BytesIO(xlsx))
        assert "설명" in wb.sheetnames, f"[{case['name']}] '설명' 시트 누락"
        assert "전략 정의 (IR)" in _sheet_text(wb["설명"])


def test_event_study_excel_has_raw_evidence_and_recompute_formula() -> None:
    """T8(#4b): 이벤트 엑셀이 집계 전 원자료('이벤트원자료')와 원자료 재계산 라이브 수식
    (=AVERAGE·=COUNTIF — 엔진 요약 검산)을 담아 '요약만'이던 증빙을 raw+수식 패리티로 끌어올린다."""
    case = next(c for c in BASE_CASES if c["name"] == "relate_event")
    _ir, _ds, _res, xlsx = _build(case["ds"], case["ir"])
    wb = load_workbook(io.BytesIO(xlsx))
    assert "이벤트원자료" in wb.sheetnames, wb.sheetnames
    raw_text = _sheet_text(wb["이벤트원자료"])
    assert "종목" in raw_text and "일자" in raw_text          # 이벤트별 원자료 헤더
    summary_text = _sheet_text(wb["이벤트분석"])
    assert "라이브 검증" in summary_text                       # 재계산 블록
    assert "AVERAGE(이벤트원자료!" in summary_text             # 평균 재계산이 원자료 참조
    assert "COUNTIF(이벤트원자료!" in summary_text             # 양(+)비율 재계산이 원자료 참조


# ── P3-b (E 뿌리): raw 증빙 엑셀 — GAP-1 전체시계열+신호수식 · GAP-2 코호트 · GAP-3 연도편중 ──
def _single_event_ir():
    """단일종목 이벤트스터디(Close>MA5) — 셀 재현 가능한 단순 이벤트(신호·발생 패널 렌더)."""
    return {"universe": {"kind": "single", "symbols": ["AAA"]}, "signal": SIG_C, "query": "relate",
            "study": {"event": EVENT, "windows": [5, 10], "event_basis": "close"}}


def test_single_event_excel_has_full_series_and_signal_formula():
    """GAP-1: 단일 이벤트스터디가 simulate처럼 전체 일별 시계열('원자료')과 '어느 행이 신호를 켰나'를
    셀 수식으로 매일 평가하는 '신호·발생'(=IF(조건,1,0)) 시트를 갖는다 — 추출 이벤트 행만 보이던 갭 종결."""
    ir, ds, res, xlsx = _build(ds_factor, _single_event_ir())
    wb = load_workbook(io.BytesIO(xlsx))
    assert {"원자료", "신호·발생"} <= set(wb.sheetnames), wb.sheetnames
    raw_text = _sheet_text(wb["원자료"])
    assert "AAA 종가(대상)" in raw_text                        # 전체 OHLCV 시계열(신호값 아님)
    assert len(list(wb["원자료"].iter_rows())) > 200            # 추출 이벤트가 아닌 전체 일수
    panel = _sheet_text(wb["신호·발생"])
    assert "이벤트(1=발생)" in panel                            # 결정 컬럼 = 발생 여부
    assert "=IF(" in panel and "원자료!" in panel               # 이벤트 조건을 원자료 참조 셀 수식으로


def test_cohort_excel_has_per_symbol_comparison_not_blank_sweep():
    """GAP-2: 종목 코호트(axis=asset)가 _build_sweep의 빈 백테스트 표가 아니라 전용 '종목비교'
    (종목×윈도 forward 수익 + n_events + 최다연도)로 렌더된다 — P1 shape의 엑셀 배선 누락 종결."""
    syms = ["AAA", "BBB", "CCC"]
    ir_d = {"universe": {"kind": "list", "symbols": syms}, "signal": SIG_C, "query": "relate",
            "study": {"axis": "entity", "assets": syms, "event": EVENT,
                      "windows": [5, 10], "event_basis": "close"}}
    ir, ds, res, xlsx = _build(ds_factor, ir_d)
    assert res.get("shape") == "cohort"
    wb = load_workbook(io.BytesIO(xlsx))
    assert "종목비교" in wb.sheetnames and "스윕결과" not in wb.sheetnames, wb.sheetnames
    text = _sheet_text(wb["종목비교"])
    assert "표본(이벤트수)" in text and "+5일 평균%" in text and "최다연도" in text
    # 실제 per-symbol 데이터가 채워졌는지(빈 표 회귀 차단) — 옛 오라우팅은 전 셀 None이었다.
    # 이제 n_events가 숫자로 채워지고(0-이벤트 종목 포함), 최소 2종목은 이벤트>0.
    ws = wb["종목비교"]
    filled = [ws.cell(r, 2).value for r in range(4, 4 + len(syms))]
    assert all(isinstance(v, (int, float)) for v in filled), filled   # None(빈표) 회귀 차단
    assert sum(1 for v in filled if v and v > 0) >= 2, filled


def test_event_excel_surfaces_year_concentration_block():
    """GAP-3(P3-a 연계): 이벤트 엑셀이 composition.by_year를 '연도별 이벤트 분포(레짐 편중 점검)'
    블록으로 실어, P3-a 함정 verdict의 raw 근거를 사용자가 직접 검산하게 한다."""
    ir, ds, res, xlsx = _build(ds_factor, _single_event_ir())
    wb = load_workbook(io.BytesIO(xlsx))
    text = _sheet_text(wb["이벤트분석"])
    assert "연도별 이벤트 분포" in text                         # 편중 점검 블록
    assert "2020" in text                                      # ds_factor는 2020~ 데이터 → 연도 표기
