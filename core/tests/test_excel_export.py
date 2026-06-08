"""선물 라이브 수식 엑셀 export 회귀 테스트 (excel_export.build_oil_excel).

PR #7(커밋 3196e39)로 머지된 build_oil_excel 에 커밋된 자동 테스트가 없어 추가.

검증 범위:
- 빌더가 에러 없이 .xlsx 바이트를 반환하고 openpyxl 로 로드된다.
- 시트 구성·순서가 ["백테스트","거래내역","만기일","로직설명"] 로 고정.
- "거래내역" 시트의 정적 값이 *같은 파라미터로 독립 호출한* run_backtest 의
  trades 와 행·열 단위로 일치 (진입가·청산가·롤횟수·수익률·PnL).

라이브 수식 시트("백테스트")는 셀이 Excel 함수라 파이썬으로 재계산할 수 없어
범위 밖이다. 정적 "거래내역" 시트가 앱 백테스트와 일치하는지가 회귀 기준.

per-symbol 일반화(PR #6) 회귀 방지를 위해 spec 3종(WTI $/USD, 코스피형
multiplier 250000/KRW, tick 0.10·multiplier 100/USD long)을 parametrize 하고
roll_cost_pct>0 케이스를 1개 포함한다. 외부 데이터 의존을 없애기 위해 임계
100 을 반복 돌파하는 결정론적 삼각파 OHLC(~220 영업일)를 합성한다.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

# 기존 core 테스트 컨벤션: 패키지 루트를 sys.path 에 추가 (editable install 우회).
_CORE_DIR = Path(__file__).resolve().parent.parent
if str(_CORE_DIR) not in sys.path:
    sys.path.insert(0, str(_CORE_DIR))

from quant_core.oil_futures.excel_export import build_oil_excel
from quant_core.oil_futures.backtest import (
    ContractSpec,
    CostModel,
    ExitRules,
    RollModel,
    WTI_SPEC,
    run_backtest,
)
from quant_core.oil_futures.signals import generate_signals


# ───── 합성 데이터 ───────────────────────────────────────────────────────────
# 임계 100 을 반복 돌파하는 삼각파. 상승 다리에서 high 가 100 을 위로 첫 터치
# (short 신호), 하강 다리에서 low 가 100 을 아래로 첫 터치(long 신호) → 주기마다
# 양방향 신호가 1회씩 발생. open=close=center, high=center+1, low=center-1 이라
# 모든 OHLC 가 정수 → 진입가/청산가 round 비교에 부동소수 잡음이 없다.
EXPECTED_SHEETS = ["백테스트", "거래내역", "만기일", "로직설명"]
_THRESHOLD = 100.0
_N_ROWS = 220


def _triangle_center(i: int, period: int = 30, lo: float = 85.0, hi: float = 115.0) -> float:
    """주기 period 로 lo↔hi 를 오가는 삼각파 (i=0 에서 lo, 중간에서 hi)."""
    half = period // 2
    pos = i % period
    if pos <= half:
        return lo + (hi - lo) * pos / half
    return hi - (hi - lo) * (pos - half) / half


def _synth_ohlc(n: int = _N_ROWS, start: str = "2020-01-03") -> pd.DataFrame:
    """결정론적 합성 일봉 (date ASC, open/high/low/close/volume)."""
    center = [_triangle_center(i) for i in range(n)]
    return pd.DataFrame({
        "date": pd.date_range(start, periods=n, freq="B"),
        "open": center,
        "high": [c + 1.0 for c in center],
        "low": [c - 1.0 for c in center],
        "close": center,
        "volume": [1000] * n,
    })


# ───── 케이스 (≥3 spec, roll>0 1개, long 1개) ────────────────────────────────

CASES = [
    {
        "id": "wti_short_usd",
        "expect_rolls": False,
        "params": dict(
            side="short", threshold=_THRESHOLD, horizon_days=30,
            spec=WTI_SPEC, commission_per_contract=2.5, slippage_ticks=1,
            roll_cost_pct=0.0, name="WTI", currency="USD", unit="배럴",
        ),
    },
    {
        # 코스피200형: multiplier 250000·KRW·roll_cost>0 (per-symbol 일반화 회귀 가드)
        "id": "kospi_short_krw_roll",
        "expect_rolls": True,
        "params": dict(
            side="short", threshold=_THRESHOLD, horizon_days=40,
            spec=ContractSpec(tick=0.05, multiplier=250000.0),
            commission_per_contract=2.5, slippage_ticks=1,
            roll_cost_pct=0.005, name="코스피200", currency="KRW", unit="포인트",
        ),
    },
    {
        "id": "micro_long_usd",
        "expect_rolls": False,
        "params": dict(
            side="long", threshold=_THRESHOLD, horizon_days=25,
            spec=ContractSpec(tick=0.10, multiplier=100.0),
            commission_per_contract=1.0, slippage_ticks=2,
            roll_cost_pct=0.0, name="미니", currency="USD", unit="",
        ),
    },
]


def _reference_trades(df: pd.DataFrame, p: dict) -> list:
    """build_oil_excel 내부와 동일한 신호·백테스트를 독립 재현해 오라클 trades 생성.

    excel_export 가 거래내역 시트에 쓰는 run_backtest 호출과 인자가 정확히 같아야
    행·열 단위 비교가 의미를 가진다. (결정론적 — 두 호출은 bit-identical.)
    """
    side = p["side"]
    sigs = generate_signals(
        df,
        short_thresholds=[p["threshold"]] if side == "short" else [],
        long_thresholds=[p["threshold"]] if side == "long" else [],
    )
    bt = run_backtest(
        df, sigs, p["horizon_days"],
        CostModel(p["commission_per_contract"], p["slippage_ticks"]),
        ExitRules(),
        RollModel(roll_cost_pct=p["roll_cost_pct"]),
        spec=p["spec"],
    )
    return bt.trades


@pytest.fixture(scope="module", params=CASES, ids=[c["id"] for c in CASES])
def built(request):
    """케이스별로 1회 빌드: (case, xlsx_bytes, 독립 run_backtest trades)."""
    case = request.param
    df = _synth_ohlc()
    xlsx = build_oil_excel(df, **case["params"])
    trades = _reference_trades(df, case["params"])
    return case, xlsx, trades


# ───── 테스트 ────────────────────────────────────────────────────────────────

def test_builder_returns_loadable_xlsx_bytes(built) -> None:
    """빌더가 에러 없이 충분한 크기의 .xlsx 바이트를 반환하고 로드된다."""
    _case, xlsx, _trades = built
    assert isinstance(xlsx, (bytes, bytearray))
    assert len(xlsx) > 5000
    wb = load_workbook(io.BytesIO(xlsx))   # 손상 없이 로드
    assert wb is not None


def test_sheetnames_exact_order(built) -> None:
    """시트 구성·순서 고정 (4시트)."""
    _case, xlsx, _trades = built
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_trade_sheet_matches_independent_backtest(built) -> None:
    """거래내역 시트의 정적 값이 독립 run_backtest trades 와 행·열 단위로 일치.

    데이터 5행~, 진입가(4)·청산가(5)·롤횟수(6)·수익률(7)·PnL(8).
    진입가/청산가 round4, num_rollovers, return_pct, round(net_pnl_usd,0).
    """
    case, xlsx, trades = built
    assert len(trades) > 0, "합성 데이터가 거래를 만들지 못함 (케이스 설계 오류)"

    wb = load_workbook(io.BytesIO(xlsx))
    tr = wb["거래내역"]
    assert tr["B2"].value == len(trades)   # 헤더의 거래 수

    for i, t in enumerate(trades):
        r = 5 + i
        assert tr.cell(row=r, column=4).value == round(t.entry_price, 4)
        assert tr.cell(row=r, column=5).value == round(t.exit_price, 4)
        assert tr.cell(row=r, column=6).value == t.num_rollovers
        assert tr.cell(row=r, column=7).value == pytest.approx(t.return_pct)
        assert tr.cell(row=r, column=8).value == round(t.net_pnl_usd, 0)

    # 거래 수만큼만 기록 — 그 다음 행은 비어 있어야 (off-by-one·잔여행 가드)
    assert tr.cell(row=5 + len(trades), column=4).value is None

    # roll_cost_pct>0 케이스는 실제로 롤오버를 거친 거래가 있어야 의미가 있다
    if case["expect_rolls"]:
        assert any(t.num_rollovers > 0 for t in trades)
