"""인사이트 엔진 증빙 엑셀 export 회귀 테스트 (excel_export.build_strategy_excel).

검증 범위:
- 빌더가 에러 없이 .xlsx 바이트를 반환하고 openpyxl 로 로드된다.
- 시트 구성·순서가 ["백테스트","거래내역","원자료","일별비중","지표·설명"] 로 고정.
- "거래내역" 시트의 정적 값이 엔진 result["trades"] 와 행·열 단위로 일치(진입가·청산가·
  보유일·수익률) — 직렬화 정합 앵커.
- "백테스트" 시트의 자산(엔진) 열이 result["equity"] 와 일치하고, 일수익·조정수익·조정자산·
  낙폭 열은 *수식 문자열*("="로 시작)이다.
- 지표 비교: 엔진 칸은 result["metrics"] 값, 엑셀 칸은 수식.

라이브 수식 시트는 셀이 Excel 함수라 파이썬으로 *수치* 재계산할 수 없다(openpyxl 미평가).
정적 값(거래·자산·지표) 일치 + 수식 문자열 정합이 회귀 기준 — 선물 export 테스트와 동일.
라이브 수식의 수치 검증(bps=0 ⇒ 엔진과 일치)은 수동 Excel 1회 확인.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

_CORE = Path(__file__).resolve().parent.parent
if str(_CORE) not in sys.path:
    sys.path.insert(0, str(_CORE))

from quant_core.blocks import const, data
from quant_core.blocks.node import Node
from quant_core.ir_engine import (Entry, Exit, PositionSpec, Sizing, SimSpec, StrategyIR,
                                   Universe, build_strategy_excel, run_strategy_ir)

EXPECTED_SHEETS = ["백테스트", "거래내역", "원자료", "일별비중", "지표·설명"]
_ROW_D0 = 13   # excel_export 백테스트 데이터 첫 행


def _bars(drift: float, n: int = 120) -> pd.DataFrame:
    idx = pd.date_range("2020-01-01", periods=n, freq="B")
    close = 100 * (1 + drift) ** np.arange(n)
    open_ = np.concatenate([[close[0]], close[:-1]])   # 전일 종가=당일 시가(갭 손익)
    return pd.DataFrame({"Open": open_, "High": np.maximum(open_, close) * 1.001,
                         "Low": np.minimum(open_, close) * 0.999, "Close": close,
                         "Volume": 1e6}, index=idx)


def _dataset() -> dict[str, pd.DataFrame]:
    """결정론적 3종목 일봉 (drift 다른 기하 시계열)."""
    return {"AAA": _bars(0.002), "BBB": _bars(-0.001), "CCC": _bars(0.001)}


def _always_true() -> Node:
    return Node(op="compare", params={"op": ">"},
                inputs={"left": data("Close"), "right": const(0)})


def _ir() -> StrategyIR:
    """3종목 매수후보유(on_signal 항상참·기간말 청산) — 비중 패널·거래 모두 발생."""
    return StrategyIR(
        name="3종목 매수후보유 (테스트)",
        signal=_always_true(),
        universe=Universe(kind="list", symbols=["AAA", "BBB", "CCC"]),
        position=PositionSpec(direction="long",
                              sizing=Sizing(mode="pct_cash", amount_pct=10.0),
                              entry=Entry(mode="on_signal"), exit=Exit()),
        simulation=SimSpec(initial_capital=1e7, commission=0.0, slippage=0.0, sell_tax=0.0))


@pytest.fixture(scope="module")
def built():
    """1회 빌드: (ir, dataset, result, xlsx_bytes)."""
    ir, ds = _ir(), _dataset()
    res = run_strategy_ir(ir, ds)
    assert res["success"], res.get("error")
    xlsx = build_strategy_excel(ir, ds, res)
    return ir, ds, res, xlsx


def test_builder_returns_loadable_xlsx_bytes(built) -> None:
    *_, xlsx = built
    assert isinstance(xlsx, (bytes, bytearray))
    assert len(xlsx) > 5000
    assert load_workbook(io.BytesIO(xlsx)) is not None


def test_sheetnames_exact_order(built) -> None:
    *_, xlsx = built
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_trade_sheet_matches_engine_result(built) -> None:
    """거래내역 정적 값이 엔진 result['trades'] 와 행·열 단위로 일치."""
    _ir_, _ds, res, xlsx = built
    trades = res["trades"]
    assert isinstance(trades, pd.DataFrame) and len(trades) > 0, "합성 전략이 거래를 안 만듦"

    tr = load_workbook(io.BytesIO(xlsx))["거래내역"]
    assert tr["B2"].value == len(trades)
    for i, (_, t) in enumerate(trades.iterrows()):
        r = 5 + i
        assert tr.cell(r, 4).value == int(t["보유일"])
        assert tr.cell(r, 5).value == round(float(t["진입가"]), 4)
        assert tr.cell(r, 6).value == round(float(t["청산가"]), 4)
        assert tr.cell(r, 7).value == round(float(t["수익률(%)"]), 4)
    # 거래 수만큼만 — 다음 행은 비어 있어야(off-by-one·잔여행 가드)
    assert tr.cell(5 + len(trades), 4).value is None


def test_backtest_sheet_equity_values_and_formulas(built) -> None:
    """백테스트 시트: 자산(엔진) 열 = result['equity'] 값, 파생열은 수식 문자열."""
    _ir_, _ds, res, xlsx = built
    equity = res["equity"]
    bt = load_workbook(io.BytesIO(xlsx))["백테스트"]

    # B열(자산) = 엔진 자산곡선 값 (앞·중간·끝 샘플)
    for i in (0, len(equity) // 2, len(equity) - 1):
        assert bt.cell(_ROW_D0 + i, 2).value == pytest.approx(float(equity.iloc[i]))

    # 입력칸 bps = 0 (기본 = 엔진 그대로)
    assert bt.cell(3, 2).value == 0

    # 둘째 데이터 행의 파생열(일수익 C·조정수익 D·조정자산 E·낙폭 F)은 수식
    for col in (3, 4, 5, 6):
        v = bt.cell(_ROW_D0 + 1, col).value
        assert isinstance(v, str) and v.startswith("="), f"col {col} 수식 아님: {v!r}"


def test_summary_engine_values_vs_excel_formulas(built) -> None:
    """지표 비교: 엔진 칸은 metrics 값, 엑셀 칸은 수식."""
    _ir_, _ds, res, xlsx = built
    m = res["metrics"]
    bt = load_workbook(io.BytesIO(xlsx))["백테스트"]

    # 누적수익률 행(6): 엔진 칸(2) = total_return 값, 엑셀 칸(3) = 수식
    assert bt.cell(6, 2).value == pytest.approx(float(m["total_return"]))
    assert str(bt.cell(6, 3).value).startswith("=")
    # 샤프 행(8): 엔진 칸 = sharpe 값
    assert bt.cell(8, 2).value == pytest.approx(float(m["sharpe"]))


def test_raw_and_weight_sheets_have_held_symbols(built) -> None:
    """원자료(역할태깅 대상 OHLC)·일별비중 헤더에 대상 종목이 노출된다(데이터·포지션 증빙)."""
    _ir_, _ds, res, xlsx = built
    held = set(str(s) for s in res["weight"].columns)
    wb = load_workbook(io.BytesIO(xlsx))
    # 원자료: 역할 태깅 헤더("{종목} 시가/…/종가(대상)") — 종목명이 부분문자열로 등장.
    raw_hdr = [wb["원자료"].cell(5, c).value for c in range(2, wb["원자료"].max_column + 1)]
    for s in held:
        assert any(s in str(h) and "대상" in str(h) for h in raw_hdr), f"원자료에 대상 {s} 없음: {raw_hdr}"
    # 일별비중: weight 패널 있으면 생성 — 헤더에 보유종목.
    wgt_hdr = {wb["일별비중"].cell(4, c).value for c in range(2, 2 + len(held))}
    assert held & wgt_hdr, f"일별비중 헤더에 보유종목 없음: {wgt_hdr}"


def test_input_roles_include_signal_refs_and_empty_weight_drops_sheet() -> None:
    """★ 재설계 핵심(입력 주도): 신호로만 참조되는 심볼(크로스에셋·거래 안 함)이 입력 역할로
    잡혀 원자료에 등장하고, weight 패널 없는 단일종목 백테스트는 일별비중 시트를 생략한다."""
    from quant_core.ir_engine.excel_export import _input_symbols_by_role
    ds = {"코스피200선물": _bars(0.001), "S&P500": _bars(-0.001)}
    # 헬퍼: subject=universe, signal=dataset−universe(신호로만 참조된 심볼)
    roles = _input_symbols_by_role({"universe": {"kind": "single", "symbols": ["코스피200선물"]}}, ds)
    assert roles["subject"] == ["코스피200선물"]
    assert "S&P500" in roles["signal"]
    # 통합: 크로스에셋 신호 단일종목 시뮬 → 원자료에 신호종목 등장, weight 없어 일별비중 생략
    sig = Node(op="select", inputs={
        "cond": Node(op="compare", params={"op": "<="},
                     inputs={"left": data("S&P500.pct_change_1d"), "right": const(-0.1)}),
        "a": const(1), "b": const(0)})
    ir = StrategyIR(name="신호종목 원자료 테스트", signal=sig,
                    universe=Universe(kind="single", symbols=["코스피200선물"]),
                    position=PositionSpec(direction="long_short", sizing=Sizing(mode="equal_weight"),
                                          entry=Entry(mode="on_signal"), exit=Exit(hold_days=0)),
                    simulation=SimSpec(initial_capital=1e8, commission=0.0))
    res = run_strategy_ir(ir, ds)
    wb = load_workbook(io.BytesIO(build_strategy_excel(ir, ds, res)))
    raw_hdr = [wb["원자료"].cell(5, c).value for c in range(2, wb["원자료"].max_column + 1)]
    assert any("S&P500" in str(h) and "신호" in str(h) for h in raw_hdr), f"신호종목 누락: {raw_hdr}"
    assert "일별비중" not in wb.sheetnames


def test_signal_decision_panel_formulas_match_engine_indicators() -> None:
    """★ Phase 2 (L2): 신호·결정 패널의 신호값·포지션 수식이 엔진 지표 정의(indicators.py)와
    1:1 일치한다 — 틀린 라이브 수식 방지(4원칙 검증). 미지원 op는 폴백(None)."""
    from quant_core.ir_engine.excel_export import _ind_excel, _signal_to_excel
    # 지표 셀 수식 = 엔진 정의(pct_change_1d = Close.pct_change(1)*100 / price_level = Close)
    assert _ind_excel("pct_change_1d", "F", 7) == "=(원자료!F7/원자료!F6-1)*100"
    assert _ind_excel("pct_change_5d", "F", 7) == "=(원자료!F7/원자료!F2-1)*100"
    assert _ind_excel("price_level", "F", 7) == "=원자료!F7"
    assert _ind_excel("rsi_14", "F", 7) is None    # 미지원 → 폴백
    # 트리 컴파일(select/compare/const/data) — data ref는 지표열 cell로 치환
    tree = {"op": "select", "inputs": {
        "cond": {"op": "compare", "params": {"op": "<="}, "inputs": {
            "left": {"op": "data", "params": {"ref": "S&P500.pct_change_1d"}},
            "right": {"op": "const", "params": {"value": -0.1}}}},
        "a": {"op": "const", "params": {"value": 1}}, "b": {"op": "const", "params": {"value": 0}}}}
    assert _signal_to_excel(tree, {"S&P500.pct_change_1d": "B"}) == "IF((B{r}<=-0.1),1.0,0.0)"
    assert _signal_to_excel({"op": "unknown_xyz"}, {}) is None   # 미지원 op → 폴백

    # 통합: 단일대상+크로스에셋 신호 → 패널 생성, 포지션 = 중첩 IF
    ds = {"코스피200선물": _bars(0.001), "S&P500": _bars(-0.0008)}
    sig = Node(op="select", inputs={
        "cond": Node(op="compare", params={"op": "<="},
                     inputs={"left": data("S&P500.pct_change_1d"), "right": const(-0.1)}),
        "a": const(1), "b": const(0)})
    ir = StrategyIR(name="패널테스트", signal=sig,
                    universe=Universe(kind="single", symbols=["코스피200선물"]),
                    position=PositionSpec(direction="long_short", sizing=Sizing(mode="equal_weight"),
                                          entry=Entry(mode="on_signal"), exit=Exit(hold_days=0)),
                    simulation=SimSpec(initial_capital=1e8, commission=0.0))
    res = run_strategy_ir(ir, ds)
    wb = load_workbook(io.BytesIO(build_strategy_excel(ir, ds, res)))
    assert "신호·결정" in wb.sheetnames
    ps = wb["신호·결정"]
    posf = next(ps.cell(r, 3).value for r in range(7, 30) if ps.cell(r, 3).value)
    assert posf.startswith("=IF(") and "<=-0.1" in posf, posf


def test_signal_panel_skipped_for_multi_symbol_portfolio(built) -> None:
    """다중 대상 포트는 종목별 결정이 1열로 표현 불가 → 신호·결정 패널 생략(폴백)."""
    _ir_, _ds, _res, xlsx = built   # 3종목 fixture
    wb = load_workbook(io.BytesIO(xlsx))
    assert "신호·결정" not in wb.sheetnames


def test_data_quality_warnings_surface_in_excel() -> None:
    """★ Phase 3 (원칙 #5): 데이터 품질·실행 경고가 엑셀 설명/지표 시트에 표면화된다 — 전 유형 공통.
    챗 응답뿐 아니라 엑셀(오프라인 검토 surface)에도 데이터 품질 계약이 닿아야 함."""
    from openpyxl import Workbook as _WB

    from quant_core.ir_engine.excel_export import _styles, _warning_block
    # 단위: _warning_block이 경고 메시지를 시트에 쓴다 / 없으면 무동작
    ws0 = _WB().active
    nxt = _warning_block(ws0, 3, {"warnings": [{"code": "stale_data", "message": "S&P500: 데이터 결손"}]}, _styles())
    flat = [ws0.cell(r, c).value for r in range(1, nxt + 1) for c in (1, 2)]
    assert any("S&P500: 데이터 결손" in str(v) for v in flat)
    assert _warning_block(_WB().active, 3, {"warnings": []}, _styles()) == 3   # 경고 없음→무동작

    # 통합: 경고 주입한 simulate 결과 → '지표·설명' 시트에 ⚠ 메시지 등장
    res = dict(run_strategy_ir(_ir(), _dataset()))
    res["warnings"] = [{"code": "data_gap", "message": "BBB: 밀도 부족 테스트경고"}]
    wb = load_workbook(io.BytesIO(build_strategy_excel(_ir(), _dataset(), res)))
    info = wb["지표·설명"]
    cells = [info.cell(r, c).value for r in range(1, info.max_row + 1) for c in (1, 2)]
    assert any("BBB: 밀도 부족 테스트경고" in str(v) for v in cells), "simulate 지표·설명에 경고 누락"


def test_deterministic_equity(built) -> None:
    """엔진 결정론 — 동일 IR·데이터 재실행 시 자산곡선 동일(증빙 앵커의 전제)."""
    ir2, ds2 = _ir(), _dataset()
    res2 = run_strategy_ir(ir2, ds2)
    _ir_, _ds, res, _xlsx = built
    pd.testing.assert_series_equal(res["equity"], res2["equity"])


def test_excel_formula_math_reproduces_engine_metrics(built) -> None:
    """백테스트 시트 라이브 수식을 파이썬으로 *동일하게* 모사해 bps=0에서 엔진 metrics를
    재현하는지 검증한다. openpyxl은 수식을 평가하지 않으므로(셀 문자열만), 수식 '설계'가
    엔진 정본과 일치함을 이 모사로 고정한다 — 증빙의 핵심 불변식.

    모사 대상(엑셀 수식과 1:1):
      C_t = B_t/B_{t-1}-1 ;  D_t = C_t - bps/10000 ;  E_13=B_13, E_t=E_{t-1}(1+D_t)
      누적=(E_last/E_13-1)*100 ; CAGR=((E_last/E_13)^(252/COUNT(E))-1)*100
      샤프=AVERAGE(D[2:])/STDEV.S(D[2:])*SQRT(252) ; MDD=MIN(E_t/cummax-1)*100
    """
    _ir_, _ds, res, _xlsx = built
    eq = res["equity"].to_numpy(dtype=float)
    m = res["metrics"]
    bps = 0.0

    c = eq[1:] / eq[:-1] - 1.0                 # 일수익률(엑셀 C, r>=2)
    d = c - bps / 10000.0                       # 조정수익률(엑셀 D)
    e = eq[0] * np.concatenate([[1.0], np.cumprod(1.0 + d)])  # 조정자산(엑셀 E), E_13=B_13
    n_eq = len(e)                               # COUNT(E)

    cum = (e[-1] / e[0] - 1.0) * 100
    cagr = ((e[-1] / e[0]) ** (252.0 / n_eq) - 1.0) * 100
    sharpe = d.mean() / d.std(ddof=1) * np.sqrt(252)        # STDEV.S = 표본(ddof=1)
    cummax = np.maximum.accumulate(e)
    mdd = (e / cummax - 1.0).min() * 100

    assert cum == pytest.approx(float(m["total_return"]), rel=1e-9, abs=1e-6)
    assert cagr == pytest.approx(float(m["cagr"]), rel=1e-9, abs=1e-6)
    assert sharpe == pytest.approx(float(m["sharpe"]), rel=1e-9, abs=1e-6)
    assert mdd == pytest.approx(float(m["mdd"]), rel=1e-9, abs=1e-6)
