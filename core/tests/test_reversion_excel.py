"""build_oil_reversion_excel — 라이브수식 엑셀 빌더 테스트.

핵심 게이트:
- 엑셀 청산/회귀 수식 로직을 파이썬으로 재현해 reversion_events(엔진)와 **정확히 동치** 단언
  (시트2 라이브 수식이 같은 트리거 집합에서 엔진과 일치함을 보장).
- openpyxl 재오픈: 4시트·raw 데이터·이벤트 행·스냅샷 값 일치.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

_CORE_DIR = Path(__file__).resolve().parent.parent
if str(_CORE_DIR) not in sys.path:
    sys.path.insert(0, str(_CORE_DIR))

from quant_core.oil_futures import (  # noqa: E402
    DOWN_EXHAUSTION,
    build_oil_reversion_excel,
    reversion_events,
    reversion_summary,
)


def _mk(closes) -> pd.DataFrame:
    n = len(closes)
    return pd.DataFrame({
        "date": pd.date_range("2020-01-01", periods=n, freq="B"),
        "open": list(closes),
        "high": [c + 1 for c in closes],
        "low": [c - 1 for c in closes],
        "close": list(closes),
        "volume": [1000] * n,
    })


def _sine_df() -> pd.DataFrame:
    t = np.arange(200)
    return _mk(110.0 + 20.0 * np.sin(t / 8.0))


def _excel_formula(df: pd.DataFrame, e, leverage: float, horizon: int):
    """시트2 라이브 수식(청산=윈도우 MIN/MAX 돌파, 회귀=close[trig+N] 환산)의 파이썬 재현."""
    closes = df["close"].to_numpy()
    di = {pd.Timestamp(d): i for i, d in enumerate(df["date"].to_numpy())}
    ti = di[pd.Timestamp(e.trigger_date)]
    tc = float(closes[ti])
    win = closes[ti + 1: ti + horizon + 1]
    if e.direction == DOWN_EXHAUSTION:   # 롱: 추가 하락이 역행
        liq = float(win.min()) <= tc * (1 - 1 / leverage)
        pos = 1.0
    else:                                # 숏: 추가 상승이 역행
        liq = float(win.max()) >= tc * (1 + 1 / leverage)
        pos = -1.0
    rev = -1.0 if liq else pos * leverage * (float(closes[ti + horizon]) / tc - 1)
    return liq, rev


# ───── 수식↔엔진 동치 (핵심) ─────────────────────────────────────────────

def test_excel_formula_matches_engine_multi():
    df = _sine_df()
    evs = reversion_events(
        df, reversal_account_pct=20, run_account_pct=80, horizon=10, leverage=10,
    )
    assert len(evs) >= 2
    for e in evs:
        liq, rev = _excel_formula(df, e, 10, 10)
        assert liq == e.liquidated
        assert rev == pytest.approx(e.reversion_account, abs=1e-9)


def test_excel_formula_matches_engine_liquidation():
    # 트리거 97 진입(롱) 후 86으로 -11.3% → 청산. 수식 재현도 청산·-1.0.
    df = _mk([101, 110, 104, 97, 86, 93])
    evs = reversion_events(
        df, reversal_account_pct=50, run_account_pct=100, horizon=2, leverage=10,
    )
    assert len(evs) == 1 and evs[0].liquidated is True
    liq, rev = _excel_formula(df, evs[0], 10, 2)
    assert liq is True
    assert rev == pytest.approx(-1.0)


# ───── openpyxl 재오픈 ────────────────────────────────────────────────────

def _build(df, **kw):
    defaults = dict(reversal=50, run=100, horizon=2, leverage=10, gap=0,
                    name="원유", price_sym="$")
    defaults.update(kw)
    evs = reversion_events(
        df, reversal_account_pct=defaults["reversal"], run_account_pct=defaults["run"],
        horizon=defaults["horizon"], leverage=defaults["leverage"], gap=defaults["gap"],
    )
    data = build_oil_reversion_excel(df, evs, **defaults)
    return evs, load_workbook(io.BytesIO(data))


def test_excel_sheets_and_snapshot():
    df = _mk([101, 110, 104, 97, 100, 104])
    evs, wb = _build(df)
    assert wb.sheetnames == ["데이터(raw)", "회귀계산", "요약", "현재 결과(스냅샷)"]
    # raw 데이터: 헤더(2행) + n행
    raw = wb["데이터(raw)"]
    assert raw["E3"].value == 101 and raw[f"E{2 + len(df)}"].value == 104
    # 스냅샷 down 표본수(B8) = 엔진 down 이벤트 수
    s = reversion_summary(evs)
    snap = wb["현재 결과(스냅샷)"]
    assert snap["B8"].value == str(s[DOWN_EXHAUSTION]["n"])
    # 회귀계산 시트에 이벤트 행 존재(트리거행 헬퍼 I열 = 시트1 행번호)
    cal = wb["회귀계산"]
    assert cal["A10"].value is not None        # 첫 이벤트(피벗일)
    assert isinstance(cal["I10"].value, int)   # 트리거행(엑셀 행번호)


def test_excel_empty_events_ok():
    # 트리거 안 잡히는 단조 시리즈 → 이벤트 0건이어도 빌드 성공·4시트.
    df = _mk([100, 101, 102, 103, 104, 105])
    evs, wb = _build(df, run=500)
    assert evs == []
    assert wb.sheetnames == ["데이터(raw)", "회귀계산", "요약", "현재 결과(스냅샷)"]
