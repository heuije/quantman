#!/usr/bin/env python
"""LLM-free 분석·엑셀 진단 CLI.

챗봇 결정적 코어(IR→엔진→결과→요약/엑셀)를 코퍼스(core/tests/analysis_corpus.py)로 전수
실행해 ① 실제 .xlsx 파일과 ② REPORT.md(형상·핵심지표·경고·시트·라이브수식 spot-check 표)를
생성한다. LLM(자연어→IR)을 거치지 않으므로 토큰 비용 0·완전 재현가능 — 다양한 분석/엑셀
생성 품질을 무료로 무제한 진단·개선하기 위한 도구.

실행:
  PYTHONUTF8=1 PYTHONPATH=core python scripts/analysis_diag.py            # 합성 데이터
  PYTHONUTF8=1 PYTHONPATH=core python scripts/analysis_diag.py --real     # frozen 실데이터(있으면)
  ... --only cross_asset   # 이름 부분일치 필터
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
for p in (_ROOT / "core", _ROOT / "core" / "tests"):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

import openpyxl

import analysis_corpus as C
from quant_core.ir_engine.summarize import summarize_result


def _headline(res: dict) -> str:
    """summarize_result(챗봇과 동일 투영)의 첫 줄 = 핵심 지표 한 줄."""
    try:
        return summarize_result(res).splitlines()[0][:90]
    except Exception as e:  # noqa: BLE001
        return f"(요약 실패: {e})"


def _spot_check(wb) -> str:
    """라이브 수식·신호패널이 실제로 살아있는지 짧은 표식."""
    out = []
    if "신호·결정" in wb.sheetnames:
        ps = wb["신호·결정"]
        f = next((ps.cell(r, ps.max_column).value for r in range(6, 40)
                  if isinstance(ps.cell(r, ps.max_column).value, str)
                  and ps.cell(r, ps.max_column).value.startswith("=")), None)
        out.append(f"신호패널:{'IF수식' if f and 'IF(' in f else (f or '없음')}")
    if "백테스트" in wb.sheetnames:
        bt = wb["백테스트"]
        has = any(isinstance(bt.cell(r, 5).value, str) and str(bt.cell(r, 5).value).startswith("=")
                  for r in range(13, 30))
        out.append(f"자산곡선:{'라이브' if has else '값'}")
    if "연도별성과" in wb.sheetnames:
        yr = wb["연도별성과"]
        f = next((yr.cell(r, 3).value for r in range(5, 30)
                  if isinstance(yr.cell(r, 3).value, str) and yr.cell(r, 3).value.startswith("=")), None)
        out.append(f"연수익:{'라이브' if f else '없음'}")
    return " · ".join(out) or "—"


def main() -> int:
    ap = argparse.ArgumentParser(description="LLM-free 분석·엑셀 진단")
    ap.add_argument("--real", action="store_true", help="frozen 실데이터 사용(없으면 합성 폴백)")
    ap.add_argument("--out", default=str(_ROOT / "analysis_diag_out"))
    ap.add_argument("--only", default=None, help="케이스 이름 부분일치 필터")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    rows: list[dict] = []

    for case in C.CASES:
        if args.only and args.only not in case["name"]:
            continue
        rec = {"name": case["name"], "desc": case["desc"], "mode": "?", "shape": "", "ok": False,
               "headline": "", "sheets": [], "n_warn": 0, "warns": "", "spot": "", "xlsx": ""}
        try:
            _ir, _ds, res, xlsx, mode = C.run_case(case, real=args.real)
            rec["mode"] = mode
            rec["ok"] = bool(res.get("success"))
            if not rec["ok"]:
                rec["shape"] = "❌ " + str(res.get("error"))[:70]
            else:
                rec["shape"] = C.case_shape(res)
                rec["headline"] = _headline(res)
                warns = res.get("warnings") or []
                rec["n_warn"] = len(warns)
                rec["warns"] = "; ".join((w.get("message") if isinstance(w, dict) else str(w)) for w in warns)[:140]
                if xlsx:
                    fp = out / f"{case['name']}.xlsx"
                    fp.write_bytes(xlsx)
                    rec["xlsx"] = fp.name
                    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
                    rec["sheets"] = wb.sheetnames
                    rec["spot"] = _spot_check(wb)
        except Exception as e:  # noqa: BLE001 — 진단 도구는 어떤 케이스가 깨져도 끝까지 돈다
            rec["shape"] = f"💥 EXCEPTION: {type(e).__name__}: {str(e)[:70]}"
        rows.append(rec)

    _write_report(out, rows, args.real)
    n_ok = sum(r["ok"] for r in rows)
    print(f"\n진단 완료: {n_ok}/{len(rows)} 성공 · 출력 → {out}")
    print(f"리포트: {out / 'REPORT.md'}")
    for r in rows:
        flag = "✅" if r["ok"] else "❌"
        print(f"  {flag} {r['name']:26} {r['shape'][:40]:40} warns={r['n_warn']} {r['spot']}")
    return 0 if n_ok == len(rows) else 1


def _write_report(out: Path, rows: list[dict], real: bool) -> None:
    lines = [f"# 분석·엑셀 진단 리포트 ({'real' if real else 'synthetic'} 데이터)", "",
             f"케이스 {len(rows)}개 · 성공 {sum(r['ok'] for r in rows)}개. LLM 미경유(토큰 0).", "",
             "| 케이스 | 모드 | 형상 | 핵심지표(요약) | 시트 | 경고 | 라이브수식 |",
             "|---|---|---|---|---|---|---|"]
    for r in rows:
        sheets = ", ".join(r["sheets"]) if r["sheets"] else "—"
        warn = f"{r['n_warn']}건: {r['warns']}" if r["n_warn"] else "—"
        lines.append(f"| **{r['name']}** | {r['mode']} | {r['shape']} | {r['headline']} | {sheets} | {warn} | {r['spot']} |")
    lines += ["", "## 케이스 설명", ""]
    for r in rows:
        lines.append(f"- **{r['name']}** — {r['desc']}" + (f" → `{r['xlsx']}`" if r["xlsx"] else ""))
    (out / "REPORT.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
